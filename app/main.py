import os
import json
import base64
import hashlib
import time
import uuid
import logging
import requests
from io import BytesIO
from PIL import Image
from datetime import datetime
from flask import Flask, request, jsonify, Response, send_file, session, redirect, url_for
from pywa.utils import default_flow_request_decryptor, default_flow_response_encryptor

app = Flask(__name__, static_folder="static", static_url_path="/static")
app.secret_key = os.environ.get("SECRET_KEY", "raihmimpi-secret-2024-xK9mP")
import functools
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

MIDTRANS_SERVER_KEY = os.environ.get("MIDTRANS_SERVER_KEY", "")
MIDTRANS_IS_PRODUCTION = os.environ.get("MIDTRANS_IS_PRODUCTION", "false").lower() == "true"
WA_PHONE_NUMBER_ID = os.environ.get("WA_PHONE_NUMBER_ID", "")
WA_ACCESS_TOKEN = os.environ.get("WA_ACCESS_TOKEN", "")
WA_TOKEN = WA_ACCESS_TOKEN  # alias untuk kompatibilitas
WABA_ID = os.environ.get("WABA_ID", "855788613793352")
WEBHOOK_VERIFY_TOKEN = os.environ.get("WEBHOOK_VERIFY_TOKEN", "raihmimpi2024")
TELEGRAM_BOT_TOKEN = os.environ.get("TELEGRAM_BOT_TOKEN", "")
TELEGRAM_CHAT_ID = os.environ.get("TELEGRAM_CHAT_ID", "")
FLOW_PRIVATE_KEY_PEM = os.environ.get("FLOW_PRIVATE_KEY", "")

# Meta Pixel / Conversions API (Raihmimpi)
META_PIXEL_ID = os.environ.get("META_PIXEL_ID", "404823950728687")
META_PIXEL_ACCESS_TOKEN = os.environ.get("META_PIXEL_ACCESS_TOKEN", "")
META_PAGE_ID = os.environ.get("META_PAGE_ID", "")

MIDTRANS_BASE_URL = (
    "https://app.midtrans.com/snap/v1/transactions"
    if MIDTRANS_IS_PRODUCTION
    else "https://app.sandbox.midtrans.com/snap/v1/transactions"
)
RAIHMIMPI_API = "https://api.raihmimpi.id/campaign"

def get_private_key_pem():
    import base64 as b64
    key_b64 = os.environ.get("FLOW_PRIVATE_KEY_B64", "")
    if key_b64:
        pem = b64.b64decode(key_b64).decode()
    else:
        pem = FLOW_PRIVATE_KEY_PEM.replace("\\n", "\n")
    return pem

def decrypt_request(body):
    decrypted, aes_key, iv = default_flow_request_decryptor(
        body["encrypted_flow_data"],
        body["encrypted_aes_key"],
        body["initial_vector"],
        get_private_key_pem()
    )
    return decrypted, aes_key, iv

def encrypt_response(response_data, aes_key, iv):
    return default_flow_response_encryptor(response_data, aes_key, iv)

_campaigns_cache = {"data": None, "ts": 0}

def get_campaigns(full=False):
    """Ambil kampanye dari API Raihmimpi.
    - full=False (default): return 5 default (page=1) — untuk Flow donasi, cepat
    - full=True: return SEMUA kampanye (page=100 → plateau ~455), dengan cache 5 menit"""
    import time
    if full:
        now = time.time()
        if _campaigns_cache["data"] is not None and (now - _campaigns_cache["ts"]) < 300:
            return _campaigns_cache["data"]
        try:
            resp = requests.get(RAIHMIMPI_API + "?page=100", timeout=30)
            resp.raise_for_status()
            data = resp.json()
            _campaigns_cache["data"] = data
            _campaigns_cache["ts"] = now
            logger.info(f"get_campaigns(full=True) refreshed: {len(data)} kampanye")
            return data
        except Exception as e:
            logger.error(f"Error ambil semua kampanye: {e}")
            return _campaigns_cache["data"] or []
    try:
        resp = requests.get(RAIHMIMPI_API, timeout=10)
        resp.raise_for_status()
        return resp.json()
    except Exception as e:
        logger.error(f"Error ambil kampanye: {e}")
        return []

def fetch_and_resize_image(url, max_size_kb=60, target_dim=120):
    """Fetch gambar dari URL, resize, dan convert ke base64 (max ~90KB)"""
    try:
        resp = requests.get(url, timeout=8)
        resp.raise_for_status()
        img = Image.open(BytesIO(resp.content))
        img = img.convert("RGB")

        # Resize ke target dimension (square crop center)
        img.thumbnail((target_dim, target_dim))

        quality = 80
        while quality > 20:
            buf = BytesIO()
            img.save(buf, format="JPEG", quality=quality)
            size_kb = buf.tell() / 1024
            if size_kb <= max_size_kb:
                break
            quality -= 10

        return base64.b64encode(buf.getvalue()).decode("utf-8")
    except Exception as e:
        logger.error(f"Error resize image {url}: {e}")
        # 1x1 transparent pixel fallback (base64 PNG)
        return "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mNk+A8AAQUAAarVyFEAAAAASUVORK5CYII="

def format_rupiah(amount):
    try:
        return f"Rp {int(amount):,}".replace(",", ".")
    except:
        return str(amount)

def filter_kampanye_aktif(campaigns):
    """Filter kampanye berdasarkan kampanye_aktif di settings.json.
    Kalau kosong/tidak ada, return semua (fallback default)."""
    try:
        settings = load_settings()
        aktif_ids = settings.get("kampanye_aktif", [])
        if not aktif_ids:
            return campaigns
        aktif_ids_str = [str(x) for x in aktif_ids]
        filtered = [c for c in campaigns if str(c.get("ID_CAMPAIGN", "")) in aktif_ids_str]
        return filtered if filtered else campaigns
    except Exception as e:
        logger.error(f"Error filter kampanye aktif: {e}")
        return campaigns

def format_campaigns_for_flow(campaigns, limit=10):
    campaigns = filter_kampanye_aktif(campaigns)
    result = []
    for c in campaigns[:limit]:
        campaign_id = str(c.get("ID_CAMPAIGN", ""))
        name = c.get("CAMPAIGN_NAME", "")[:72]
        terkumpul = format_rupiah(c.get("TOTAL_DONASI", 0))
        target = format_rupiah(c.get("TARGET_DONASI_UANG", 0))
        result.append({"id": campaign_id, "title": name, "description": f"Terkumpul: {terkumpul} dari {target}"[:72]})
    return result

def format_campaigns_with_images(campaigns, limit=5, tipe_donasi=""):
    """Format kampanye dengan gambar base64 untuk NavigationList.
    tipe_donasi arg di-accept untuk backward compat tapi tidak dipakai (NavigationList Flow 7.x handle navigation sendiri)."""
    campaigns = filter_kampanye_aktif(campaigns)
    result = []
    for c in campaigns[:limit]:
        campaign_id = str(c.get("ID_CAMPAIGN", ""))
        name = c.get("CAMPAIGN_NAME", "")[:30]
        terkumpul_raw = c.get("TOTAL_DONASI", 0)
        target_raw = c.get("TARGET_DONASI_UANG", 0)
        try:
            pct = min(100, round(int(terkumpul_raw) / int(target_raw) * 100))
        except:
            pct = 0
        terkumpul = format_rupiah(terkumpul_raw)

        img_url = c.get("IMG_MOBILE") or c.get("IMG_CAMPAIGNER") or c.get("IMG_BIG", "")
        image_b64 = fetch_and_resize_image(img_url) if img_url else ""

        result.append({
            "id": campaign_id,
            "main-content": {
                "title": name,
                "description": f"Terkumpul {pct}%"[:20]
            },
            "start": {
                "image": image_b64
            }
        })
    return result

def create_midtrans_payment(order_id, amount, donatur_name, phone, campaign_name):
    import base64 as b64, re
    auth = b64.b64encode(f"{MIDTRANS_SERVER_KEY}:".encode()).decode()
    # Sanitize: Midtrans tidak terima beberapa karakter khusus di item.name
    safe_campaign = re.sub(r"[^A-Za-z0-9 \-]", "", campaign_name)[:40]
    safe_donatur = re.sub(r"[^A-Za-z0-9 ]", "", donatur_name)[:60] or "Donatur"
    payload = {
        "transaction_details": {"order_id": order_id, "gross_amount": int(amount)},
        "customer_details": {"first_name": safe_donatur, "phone": phone},
        "item_details": [{"id": "DONASI", "price": int(amount), "quantity": 1, "name": f"Donasi {safe_campaign}"[:50]}],
        "callbacks": {"finish": f"https://raihmimpi.id/donasi-sukses?order_id={order_id}"}
    }
    logger.info(f"Midtrans PAYLOAD: {json.dumps(payload)}")
    resp = requests.post(MIDTRANS_BASE_URL, json=payload,
        headers={"Authorization": f"Basic {auth}", "Content-Type": "application/json"}, timeout=15)
    logger.info(f"Midtrans RESPONSE status={resp.status_code} body={resp.text[:500]}")
    resp.raise_for_status()
    return resp.json().get("redirect_url")

def send_wa_message(to_phone, message):
    url = f"https://graph.facebook.com/v22.0/{WA_PHONE_NUMBER_ID}/messages"
    resp = requests.post(url, json={"messaging_product": "whatsapp", "to": to_phone, "type": "text", "text": {"body": message}},
        headers={"Authorization": f"Bearer {WA_ACCESS_TOKEN}"}, timeout=10)
    logger.info(f"send_wa_message to={to_phone} status={resp.status_code} body={resp.text[:200]}")
    return resp

def send_wa_buttons(to_phone, body_text, buttons):
    """Kirim interactive reply button message via Graph API.
    buttons: list of {"id": "...", "title": "..."} (max 3, title max 20 char)"""
    url = f"https://graph.facebook.com/v22.0/{WA_PHONE_NUMBER_ID}/messages"
    payload = {
        "messaging_product": "whatsapp",
        "to": to_phone,
        "type": "interactive",
        "interactive": {
            "type": "button",
            "body": {"text": body_text},
            "action": {
                "buttons": [
                    {"type": "reply", "reply": {"id": b["id"], "title": b["title"][:20]}}
                    for b in buttons
                ]
            }
        }
    }
    resp = requests.post(url, json=payload,
        headers={"Authorization": f"Bearer {WA_ACCESS_TOKEN}"}, timeout=10)
    logger.info(f"send_wa_buttons to={to_phone} status={resp.status_code} body={resp.text[:200]}")
    return resp

def send_menu_utama(to_phone):
    """Kirim pesan Menu Utama (sapaan + tombol aksi) ke kontak baru."""
    settings = load_settings()
    menu = settings.get("menu_utama", {})
    message = menu.get("message", "Halo! Yuk donasi via Raihmimpi.")
    buttons_cfg = menu.get("buttons", [])
    buttons = []
    for i, b in enumerate(buttons_cfg):
        if not b.get("enabled"):
            continue
        action = b.get("action", "donasi")
        btn_id = f"btn_{action}_{i}"
        buttons.append({"id": btn_id, "title": b.get("label", "")})
    if not buttons:
        buttons = [{"id": "btn_donasi", "title": "Mulai Donasi"}]
    resp = send_wa_buttons(to_phone, message, buttons[:3])
    # Simpan ke inbox sebagai outgoing message interactive (untuk render di chat dashboard)
    try:
        inbox = load_inbox()
        contacts = inbox.setdefault("contacts", {})
        messages = inbox.setdefault("messages", {})
        contact = contacts.get(to_phone, {"phone": to_phone, "name": to_phone, "labels": [], "unread": 0})
        contact["last_message"] = message
        contact["last_message_at"] = datetime.now().isoformat()
        contacts[to_phone] = contact
        msg_list = messages.setdefault(to_phone, [])
        msg_list.append({
            "direction": "out",
            "type": "interactive",
            "text": message,
            "buttons": [b["title"] for b in buttons[:3]],
            "timestamp": datetime.now().isoformat(),
        })
        save_inbox(inbox)
    except Exception as e:
        logger.error(f"Gagal record Menu Utama outgoing: {e}")
    return resp

# ID Flow donasi (didapat dari WhatsApp Manager > Flows)
DONASI_FLOW_ID = os.environ.get("DONASI_FLOW_ID", "")

# Kata kunci yang men-trigger Flow donasi
DONASI_KEYWORDS = ["donasi", "infak", "infaq", "sedekah", "zakat", "wakaf", "berdonasi", "donatur"]

def send_wa_flow_message(to_phone, body_text="Yuk mulai donasi via Raihmimpi 🤲", cta_text="Mulai Donasi", screen="PILIH_TIPE"):
    """Kirim interactive Flow message langsung ke user (dalam window 24 jam, tanpa template)."""
    url = f"https://graph.facebook.com/v19.0/{WA_PHONE_NUMBER_ID}/messages"
    flow_token = f"phone_{to_phone}"
    payload = {
        "messaging_product": "whatsapp",
        "to": to_phone,
        "type": "interactive",
        "interactive": {
            "type": "flow",
            "body": {"text": body_text},
            "action": {
                "name": "flow",
                "parameters": {
                    "flow_message_version": "3",
                    "flow_token": flow_token,
                    "flow_id": DONASI_FLOW_ID,
                    "flow_cta": cta_text,
                    "flow_action": "data_exchange"
                }
            }
        }
    }
    resp = requests.post(url, json=payload,
        headers={"Authorization": f"Bearer {WA_ACCESS_TOKEN}", "Content-Type": "application/json"}, timeout=10)
    logger.info(f"send_wa_flow_message status={resp.status_code} body={resp.text}")
    return resp

# ============================================================
def notify_telegram(text):
    if not TELEGRAM_BOT_TOKEN or not TELEGRAM_CHAT_ID:
        return
    requests.post(f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage",
        json={"chat_id": TELEGRAM_CHAT_ID, "text": text, "parse_mode": "HTML"}, timeout=10)

def hash_sha256(value):
    """Hash nilai (lowercase, trimmed) dengan SHA256 sesuai standar Meta CAPI."""
    if not value:
        return None
    return hashlib.sha256(str(value).strip().lower().encode("utf-8")).hexdigest()

def get_ctwa_for_phone(phone):
    """Lookup ctwa_clid yang tersimpan di inbox contact berdasarkan phone.
    Return ctwa_clid string atau None."""
    if not phone:
        return None
    try:
        clean = phone.replace("+", "").replace(" ", "").replace("-", "")
        inbox = load_inbox()
        contact = inbox.get("contacts", {}).get(clean, {})
        clid = contact.get("ctwa_clid", "")
        # Validasi: minimal 50 char dan bukan test
        if clid and len(clid) >= 50 and not clid.startswith("test_") and not clid.startswith("clid_test"):
            return clid
    except Exception as e:
        logger.error(f"get_ctwa_for_phone error: {e}")
    return None

def send_pixel_event(event_name, phone=None, value=None, currency="IDR", event_id=None,
                      content_name=None, content_ids=None, source_url=None, ctwa_clid=None):
    """
    Kirim event ke Meta Conversions API (Pixel Raihmimpi).
    - phone: nomor WA donatur (62...), dipakai sebagai 'ph' (hashed) untuk matching.
    - event_id: untuk deduplikasi jika nanti dikombinasikan dengan Pixel browser-side.
    - source_url: action_source dianggap 'business_messaging' karena ini dari WhatsApp Flow.
    """
    if not META_PIXEL_ACCESS_TOKEN:
        logger.warning(f"send_pixel_event SKIP (no META_PIXEL_ACCESS_TOKEN): {event_name}")
        return None

    # Skip kalau business_messaging tapi tidak ada ctwa_clid (Meta akan reject 400)
    # Donatur non-CTWA (organic/manual) tidak punya ctwa_clid, attribution tidak applicable
    if not ctwa_clid:
        logger.info(f"send_pixel_event SKIP (no ctwa_clid, non-CTWA traffic): {event_name} phone={phone}")
        return None

    user_data = {}
    if phone:
        # Normalisasi: pastikan format 62xxxxxxxxxx tanpa '+'
        clean = phone.replace("+", "").replace(" ", "").replace("-", "")
        user_data["ph"] = [hash_sha256(clean)]
    # ctwa_clid wajib untuk CTWA attribution
    if ctwa_clid:
        user_data["ctwa_clid"] = ctwa_clid
    # page_id di user_data (per dokumentasi Meta)
    if META_PAGE_ID:
        user_data["page_id"] = META_PAGE_ID

    custom_data = {"currency": currency}
    if value is not None:
        custom_data["value"] = float(value)
    if content_name:
        custom_data["content_name"] = content_name
    if content_ids:
        custom_data["content_ids"] = content_ids

    event_data = {
        "event_name": event_name,
        "event_time": int(time.time()),
        "event_id": event_id or str(uuid.uuid4()),
        "action_source": "business_messaging",
        "messaging_channel": "whatsapp",
        "user_data": user_data,
        "custom_data": custom_data,
    }
    payload = {"data": [event_data]}

    try:
        resp = requests.post(
            f"https://graph.facebook.com/v22.0/{META_PIXEL_ID}/events",
            params={"access_token": META_PIXEL_ACCESS_TOKEN},
            json=payload, timeout=10)
        logger.info(f"send_pixel_event {event_name} status={resp.status_code} body={resp.text[:300]}")
        return resp
    except Exception as e:
        logger.error(f"send_pixel_event {event_name} failed: {e}", exc_info=True)
        return None

DATA_FILE = os.environ.get("DATA_DIR", "/tmp") + "/transaksi.json"

def load_data():
    if not os.path.exists(DATA_FILE):
        return []
    with open(DATA_FILE, "r") as f:
        return json.load(f)

def save_data(data):
    with open(DATA_FILE, "w") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

# ============================================================
# Inbox / Percakapan WhatsApp
# ============================================================
INBOX_FILE = os.environ.get("DATA_DIR", "/tmp") + "/inbox.json"

def load_inbox():
    """Struktur: {"contacts": {phone: {"name":..., "phone":..., "labels":[...], "status":"perlu_dibalas|otomatis|selesai",
       "last_message":..., "last_message_at":..., "unread": int}}, "messages": {phone: [ {direction, type, text, timestamp} ]}}"""
    if not os.path.exists(INBOX_FILE):
        return {"contacts": {}, "messages": {}}
    with open(INBOX_FILE, "r") as f:
        return json.load(f)

def save_inbox(inbox):
    with open(INBOX_FILE, "w") as f:
        json.dump(inbox, f, ensure_ascii=False, indent=2)

# ============================================================
# Settings (Menu Utama, dll)
# ============================================================
SETTINGS_FILE = os.environ.get("DATA_DIR", "/tmp") + "/settings.json"
SHORTCUTS_FILE = os.environ.get("DATA_DIR", "/tmp") + "/shortcuts.json"
SHORTCUTS_MEDIA_DIR = os.environ.get("DATA_DIR", "/tmp") + "/shortcut_files"

DEFAULT_SETTINGS = {
    "menu_utama": {
        "message": "Halo! kak, sekarang kakak bisa donasi via WhatsApp di Raihmimpi.\n\nYuk, coba donasi kak!",
        "buttons": [
            {"enabled": True, "label": "Saya mau donasi", "action": "donasi"},
            {"enabled": True, "label": "Saya mau WA admin", "action": "admin"},
            {"enabled": False, "label": "Kembali ke Donasi", "action": "donasi"},
        ],
    }
}

def load_shortcuts():
    if not os.path.exists(SHORTCUTS_FILE):
        return []
    try:
        with open(SHORTCUTS_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return []

def save_shortcuts(data):
    with open(SHORTCUTS_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def load_settings():
    if not os.path.exists(SETTINGS_FILE):
        return DEFAULT_SETTINGS.copy()
    with open(SETTINGS_FILE, "r") as f:
        data = json.load(f)
    merged = DEFAULT_SETTINGS.copy()
    merged.update(data)
    return merged

def save_settings(settings):
    with open(SETTINGS_FILE, "w") as f:
        json.dump(settings, f, ensure_ascii=False, indent=2)

def record_incoming_message(phone, text, msg_type="text", name=None):
    """Simpan pesan masuk ke inbox, update info kontak (unread, last_message, status)."""
    inbox = load_inbox()
    contacts = inbox.setdefault("contacts", {})
    messages = inbox.setdefault("messages", {})

    contact = contacts.get(phone, {})
    contact["phone"] = phone
    if name and not contact.get("name"):
        contact["name"] = name
    elif not contact.get("name"):
        contact["name"] = phone
    contact.setdefault("labels", [])
    contact["status"] = "perlu_dibalas"
    contact["last_message"] = text if msg_type == "text" else f"[{msg_type.upper()}]"
    contact["last_message_at"] = datetime.now().isoformat()
    contact["unread"] = contact.get("unread", 0) + 1
    contacts[phone] = contact

    msg_list = messages.setdefault(phone, [])
    msg_list.append({
        "direction": "in",
        "type": msg_type,
        "text": text,
        "timestamp": datetime.now().isoformat(),
    })

    save_inbox(inbox)
    return contact

def record_outgoing_message(phone, text, msg_type="text"):
    """Simpan pesan keluar (balasan admin) ke inbox."""
    inbox = load_inbox()
    contacts = inbox.setdefault("contacts", {})
    messages = inbox.setdefault("messages", {})

    contact = contacts.get(phone, {"phone": phone, "name": phone, "labels": [], "unread": 0})
    contact["last_message"] = text if msg_type == "text" else f"[{msg_type.upper()}]"
    contact["last_message_at"] = datetime.now().isoformat()
    contacts[phone] = contact

    msg_list = messages.setdefault(phone, [])
    msg_list.append({
        "direction": "out",
        "type": msg_type,
        "text": text,
        "timestamp": datetime.now().isoformat(),
    })

    save_inbox(inbox)
    return contact

def handle_flow_request(decrypted_body):
    action = decrypted_body.get("action")
    screen = decrypted_body.get("screen")
    data = decrypted_body.get("data", {})
    flow_token = decrypted_body.get("flow_token", "")

    logger.info(f"Flow action={action} screen={screen}")

    if action == "ping":
        return {"version": "3.0", "data": {"status": "active"}}

    if action == "INIT":
        campaigns = get_campaigns(full=True)
        phone_init = flow_token.replace("phone_", "") if flow_token.startswith("phone_") else ""
        send_pixel_event("ViewContent", phone=phone_init, currency="IDR",
                          event_id=f"vc_{flow_token}_{int(time.time())}",
                          content_name="Donasi via WA Raihmimpi",
                          ctwa_clid=get_ctwa_for_phone(phone_init))
        # Simpan flag view_content ke contact
        if phone_init:
            try:
                inbox_vc = load_inbox()
                if phone_init in inbox_vc.get("contacts", {}):
                    inbox_vc["contacts"][phone_init]["view_content_at"] = datetime.now().isoformat()
                    save_inbox(inbox_vc)
            except Exception as vce:
                logger.error(f"Gagal simpan view_content_at: {vce}")
        return {"screen": "PILIH_TIPE", "data": {"kampanye_list": format_campaigns_with_images(campaigns)}}

    if action == "data_exchange":
        if screen == "PILIH_KAMPANYE":
            logger.info(f"PILIH_KAMPANYE FULL_BODY: {json.dumps(decrypted_body)[:1000]}")
            tipe_donasi = str(data.get("tipe_donasi", "sekali"))
            # NavigationList Flow 7.x kirim ID item yang diklik via ${form.kampanye_nav}
            # yang sudah di-resolve ke field kampanye_id di payload
            kampanye_id = ""
            for key in ["kampanye_id", "kampanye_nav", "id", "selected_id"]:
                val = data.get(key)
                if isinstance(val, list) and val:
                    kampanye_id = str(val[0])
                    break
                elif isinstance(val, str) and val and not val.startswith("${"):
                    kampanye_id = val
                    break

            logger.info(f"PILIH_KAMPANYE: data_keys={list(data.keys())} kampanye_id={kampanye_id} tipe={tipe_donasi}")

            if kampanye_id:
                # User pilih kampanye -> fetch nama dari API berdasarkan ID
                kampanye_nama = "Kampanye Raihmimpi"
                try:
                    campaigns = get_campaigns(full=True)
                    for c in campaigns:
                        if str(c.get("ID_CAMPAIGN", "")) == kampanye_id:
                            kampanye_nama = c.get("CAMPAIGN_NAME", "Kampanye Raihmimpi")
                            break
                except Exception as e:
                    logger.error(f"Error fetch nama kampanye: {e}")
                logger.info(f"PILIH_KAMPANYE resolved: id={kampanye_id} nama={kampanye_nama}")
                return {"screen": "PILIH_NOMINAL", "data": {
                    "tipe_donasi": tipe_donasi,
                    "kampanye_id": kampanye_id,
                    "kampanye_nama": kampanye_nama,
                }}
            else:
                # Belum ada kampanye dipilih -> return list kampanye (fallback, biasanya tidak terjadi)
                campaigns = get_campaigns(full=True)
                return {"screen": "PILIH_KAMPANYE", "data": {
                    "kampanye_list": format_campaigns_with_images(campaigns, tipe_donasi=tipe_donasi),
                    "tipe_donasi": tipe_donasi
                }}

        if screen == "PILIH_NOMINAL":
            # User klik "Isi Data Donatur" -> trigger AddToCart, lanjut ke screen DATA_DONATUR
            kampanye_id = str(data.get("kampanye_id", ""))
            kampanye_nama = str(data.get("kampanye_nama", ""))
            # Fix: kalau kampanye_nama masih template variable, fetch dari API berdasarkan ID
            if not kampanye_nama or kampanye_nama.startswith("${"):
                try:
                    campaigns = get_campaigns(full=True)
                    for c in campaigns:
                        if str(c.get("id")) == kampanye_id:
                            kampanye_nama = c.get("main-content", {}).get("title", "Kampanye Raihmimpi")
                            break
                    if not kampanye_nama or kampanye_nama.startswith("${"):
                        kampanye_nama = campaigns[0].get("main-content", {}).get("title", "Kampanye Raihmimpi") if campaigns else "Kampanye Raihmimpi"
                except Exception:
                    kampanye_nama = "Kampanye Raihmimpi"
            logger.info(f"PILIH_NOMINAL: kampanye_id={kampanye_id} kampanye_nama={kampanye_nama}")
            nominal = data.get("nominal", "50000")
            nominal_lain = data.get("nominal_lain", 0)
            try:
                final_nominal = int(nominal_lain) if nominal_lain and int(nominal_lain) > 0 else int(nominal)
            except (ValueError, TypeError):
                final_nominal = int(nominal) if str(nominal).isdigit() else 50000


            try:
                nominal_lain_int = int(nominal_lain) if nominal_lain else 0
            except (ValueError, TypeError):
                nominal_lain_int = 0

            return {"screen": "DATA_DONATUR", "data": {
                "tipe_donasi": data.get("tipe_donasi", "sekali"),
                "kampanye_id": str(kampanye_id),
                "kampanye_nama": str(kampanye_nama),
                "nominal": str(nominal),
                "nominal_lain": nominal_lain_int,
            }}

        if screen == "DATA_DONATUR":
            # User klik Lihat Konfirmasi -> return data ke screen KONFIRMASI
            try:
                nominal_lain_val = int(data.get("nominal_lain", 0) or 0)
            except (ValueError, TypeError):
                nominal_lain_val = 0
            nominal_raw = str(data.get("nominal", "50000"))
            try:
                nominal_int = nominal_lain_val if nominal_lain_val > 0 else int(nominal_raw)
            except (ValueError, TypeError):
                nominal_int = 50000
            nominal_display = f"Rp {nominal_int:,}".replace(",", ".")
            return {"screen": "KONFIRMASI", "data": {
                "tipe_donasi": str(data.get("tipe_donasi", "sekali")),
                "kampanye_id": str(data.get("kampanye_id", "")),
                "kampanye_nama": str(data.get("kampanye_nama", "Kampanye Raihmimpi")),
                "nominal": nominal_raw,
                "nominal_lain": nominal_lain_val,
                "nominal_display": nominal_display,
                "nama_donatur": str(data.get("nama_donatur", "Donatur")),
                "atas_nama": str(data.get("atas_nama", "")),
            }}

        if screen == "KONFIRMASI":
            logger.info(f"KONFIRMASI data: {data}")
            logger.info(f"KONFIRMASI flow_token: {flow_token}")
            try:
                nama_donatur = str(data.get("nama_donatur", "Donatur"))
                kampanye_id = str(data.get("kampanye_id", ""))
                kampanye_nama = str(data.get("kampanye_nama", "Kampanye Raihmimpi"))
                nominal = str(data.get("nominal", "50000"))
                nominal_lain = data.get("nominal_lain", 0)
                atas_nama = str(data.get("atas_nama", nama_donatur))
                tipe = str(data.get("tipe_donasi", "sekali"))
                logger.info(f"KONFIRMASI step1 OK: nominal={nominal} kampanye_nama={kampanye_nama}")

                # Fix template variables yang tidak ter-resolve dari Flow JSON lama
                if not kampanye_id or kampanye_id.startswith("${"):
                    kampanye_id = "unknown"
                if kampanye_nama.startswith("${"):
                    try:
                        kampanye_list = get_campaigns(full=True)
                        if kampanye_list:
                            kampanye_nama = kampanye_list[0].get("main-content", {}).get("title", "Kampanye Raihmimpi")
                            kampanye_id = str(kampanye_list[0].get("id", "unknown"))
                        logger.info(f"KONFIRMASI kampanye fetch OK: {kampanye_nama}")
                    except Exception as e2:
                        logger.error(f"KONFIRMASI kampanye fetch error: {e2}")
                        kampanye_nama = "Kampanye Raihmimpi"

                try:
                    nominal_lain_int = int(nominal_lain) if nominal_lain else 0
                except (ValueError, TypeError):
                    nominal_lain_int = 0

                try:
                    final_nominal = nominal_lain_int if nominal_lain_int > 0 else int(nominal)
                except (ValueError, TypeError):
                    final_nominal = 50000

                logger.info(f"KONFIRMASI step2 OK: final_nominal={final_nominal}")
                phone = flow_token.replace("phone_", "") if flow_token and flow_token.startswith("phone_") else ""
                kid_clean = kampanye_id[-6:].replace(".", "").replace("}", "").replace("{", "")
                order_id = f"RM-{datetime.now().strftime('%Y%m%d%H%M%S')}-{kid_clean}"
                logger.info(f"KONFIRMASI step3 OK: order_id={order_id} phone={phone}")

                payment_url = create_midtrans_payment(order_id, final_nominal, nama_donatur, phone, kampanye_nama)
                logger.info(f"KONFIRMASI step4 OK: payment_url={payment_url[:50] if payment_url else 'None'}")

                transaksi = load_data()
                transaksi.append({"order_id": order_id, "donatur": nama_donatur, "atas_nama": atas_nama, "phone": phone,
                    "kampanye_id": kampanye_id, "kampanye": kampanye_nama, "nominal": final_nominal, "tipe": tipe,
                    "status": "pending", "payment_url": payment_url, "created_at": datetime.now().isoformat()})
                save_data(transaksi)
                logger.info("KONFIRMASI step5 OK: data saved")

                if phone:
                    send_wa_message(phone, f"Assalamu'alaikum *{nama_donatur}*! \U0001f932\n\nTerima kasih berniat berdonasi untuk:\n*{kampanye_nama}*\n\nNominal: *{format_rupiah(final_nominal)}*\n\nSelesaikan donasi di:\n{payment_url}\n\n_Link berlaku 24 jam. Semoga berkah._ \U0001f64f")
                    logger.info(f"KONFIRMASI step6 OK: WA sent to {phone}")

                notify_telegram(f"\U0001f514 <b>Donasi Baru!</b>\n\U0001f464 {nama_donatur} ({phone})\n\U0001f4cb {kampanye_nama}\n\U0001f4b0 {format_rupiah(final_nominal)}\n\U0001f194 {order_id}")
                nominal_display = f"Rp {final_nominal:,}".replace(",", ".")
                return {"screen": "SELESAI", "data": {
                    "payment_url": payment_url or "",
                    "order_id": order_id,
                    "nama_donatur": nama_donatur,
                    "kampanye_nama": kampanye_nama,
                    "nominal_display": nominal_display
                }}
            except Exception as e:
                logger.error(f"KONFIRMASI handler FATAL ERROR: {e}", exc_info=True)
                return {"screen": "SELESAI", "data": {
                    "payment_url": "",
                    "order_id": "ERROR-" + datetime.now().strftime("%H%M%S"),
                    "nama_donatur": "Donatur",
                    "kampanye_nama": "Kampanye Raihmimpi",
                    "nominal_display": "Rp 0"
                }}


    return {"screen": "PILIH_TIPE", "data": {}}

@app.route("/wa-flow", methods=["GET"])
def wa_flow_verify():
    """Verifikasi webhook Meta untuk App subscription."""
    mode = request.args.get("hub.mode")
    token = request.args.get("hub.verify_token")
    challenge = request.args.get("hub.challenge")
    logger.info(f"Webhook verify: mode={mode} token={token} challenge={challenge}")
    if mode == "subscribe" and token == WEBHOOK_VERIFY_TOKEN:
        logger.info("Webhook verified successfully")
        return challenge, 200
    logger.warning(f"Webhook verify FAILED: token mismatch atau mode salah")
    return jsonify({"error": "Forbidden"}), 403

@app.route("/wa-flow", methods=["POST"])
def wa_flow_endpoint():
    try:
        body = request.get_json()
        logger.info(f"RAW BODY KEYS: {list(body.keys()) if body else 'None'}")

        if body and "encrypted_aes_key" in body:
            decrypted_body, aes_key, iv = decrypt_request(body)
            response_data = handle_flow_request(decrypted_body)
            encrypted_response = encrypt_response(response_data, aes_key, iv)
            return Response(encrypted_response, mimetype="text/plain")

        if body and "entry" in body:
            try:
                for entry in body.get("entry", []):
                    for change in entry.get("changes", []):
                        value = change.get("value", {})
                        messages = value.get("messages", [])
                        contacts_meta = value.get("contacts", [])
                        for msg in messages:
                            phone = msg.get("from")
                            if not phone:
                                continue
                            msg_type = msg.get("type", "text")
                            button_reply_id = None
                            if msg_type == "text":
                                text = msg.get("text", {}).get("body", "")
                            elif msg_type == "interactive":
                                interactive = msg.get("interactive", {})
                                itype = interactive.get("type")
                                if itype == "button_reply":
                                    button_reply_id = interactive.get("button_reply", {}).get("id")
                                    text = interactive.get("button_reply", {}).get("title", "")
                                elif itype == "nfm_reply":
                                    # Flow submission — tampilkan persis seperti di WhatsApp
                                    text = "📄 Mulai Donasi · Jawaban terkirim"
                                else:
                                    text = json.dumps(interactive)[:500]
                            else:
                                text = f"[{msg_type.upper()}]"

                            contact_name = None
                            for c in contacts_meta:
                                if c.get("wa_id") == phone:
                                    contact_name = c.get("profile", {}).get("name")

                            # Extract CTWA referral (dari iklan Meta) - struktur Meta native
                            referral = msg.get("referral") or {}
                            ctwa_clid = referral.get("ctwa_clid", "")

                            logger.info(f"WA WEBHOOK MSG from={phone} name={contact_name} type={msg_type} button_id={button_reply_id} text={text[:50] if text else ''} ctwa={ctwa_clid[:30] if ctwa_clid else 'no'}")

                            inbox_before = load_inbox()
                            existing_contact = inbox_before.get("contacts", {}).get(phone)

                            record_incoming_message(phone, text, msg_type="text" if msg_type == "text" else msg_type, name=contact_name)

                            # Auto-label dari pesan yang mengandung [nama_label]
                            if msg_type == "text" and text:
                                import re as _re
                                bracket_labels = _re.findall(r'\[([^\[\]]+)\]', text)
                                if bracket_labels:
                                    try:
                                        labels_cfg = load_labels()
                                        existing_label_names = [l["name"].lower() for l in labels_cfg]
                                        inbox_al = load_inbox()
                                        contact_al = inbox_al.get("contacts", {}).get(phone, {})
                                        contact_labels = contact_al.get("labels", [])
                                        changed = False
                                        for bl in bracket_labels:
                                            bl = bl.strip()
                                            if not bl:
                                                continue
                                            # Buat label baru jika belum ada
                                            if bl.lower() not in existing_label_names:
                                                import uuid as _uuid
                                                new_label = {
                                                    "id": str(_uuid.uuid4()),
                                                    "name": bl,
                                                    "category": "Meta Ads",
                                                    "bg_color": "#e0d9ff",
                                                    "text_color": "#5b3df0"
                                                }
                                                labels_cfg.append(new_label)
                                                settings_al = load_settings()
                                                settings_al["labels"] = labels_cfg
                                                save_settings(settings_al)
                                                existing_label_names.append(bl.lower())
                                                logger.info(f"Auto-created label [{bl}] kategori Meta Ads")
                                            # Assign ke kontak jika belum ada
                                            if bl not in contact_labels:
                                                contact_labels.append(bl)
                                                changed = True
                                        if changed:
                                            contact_al["labels"] = contact_labels
                                            inbox_al["contacts"][phone] = contact_al
                                            save_inbox(inbox_al)
                                            logger.info(f"Auto-label {bracket_labels} assigned ke {phone}")
                                    except Exception as ale:
                                        logger.error(f"Auto-label error: {ale}", exc_info=True)

                            # Simpan metadata CTWA + nama jika ada
                            if referral or contact_name:
                                inbox = load_inbox()
                                contact_rec = inbox.get("contacts", {}).get(phone, {})
                                if contact_name and contact_name != phone:
                                    contact_rec["name"] = contact_name
                                if referral:
                                    contact_rec["ctwa_clid"] = ctwa_clid
                                    contact_rec["ad_source_url"] = referral.get("source_url", "")
                                    contact_rec["ad_source_id"] = referral.get("source_id", "")
                                    contact_rec["ad_source_type"] = referral.get("source_type", "")
                                    contact_rec["ad_headline"] = referral.get("headline", "")
                                    contact_rec["ad_body"] = referral.get("body", "")
                                    contact_rec["ad_media_type"] = referral.get("media_type", "")
                                    contact_rec["ad_image_url"] = referral.get("image_url", "")
                                    contact_rec["ad_thumbnail_url"] = referral.get("thumbnail_url", "")
                                inbox["contacts"][phone] = contact_rec
                                save_inbox(inbox)
                                if referral:
                                    logger.info(f"CTWA referral disimpan {phone}: headline={referral.get('headline','')[:40]} img={(referral.get('image_url') or referral.get('thumbnail_url') or '')[:60]}")

                                    # Pixel LeadSubmitted untuk CTWA (validasi ctwa_clid asli)
                                    had_ctwa_before = existing_contact and existing_contact.get("ctwa_clid")
                                    is_valid_clid = ctwa_clid and len(ctwa_clid) >= 50 and not ctwa_clid.startswith("clid_test") and not ctwa_clid.startswith("test_")
                                    if not had_ctwa_before and is_valid_clid:
                                        try:
                                            send_pixel_event("LeadSubmitted", phone=phone, currency="IDR",
                                                              event_id=f"lead_{phone}_{int(datetime.now().timestamp())}",
                                                              content_name=referral.get("headline", "CTWA Raihmimpi"),
                                                              ctwa_clid=ctwa_clid)
                                        except Exception as pe:
                                            logger.error(f"Pixel Lead event gagal: {pe}")

                            # Handle klik tombol Menu Utama
                            if button_reply_id and (button_reply_id == "btn_donasi" or button_reply_id.startswith("btn_donasi")):
                                try:
                                    send_wa_flow_message(phone)
                                    logger.info(f"Flow donasi dikirim ke {phone} (klik tombol Mulai Donasi)")
                                except Exception as fe:
                                    logger.error(f"Gagal kirim Flow ke {phone}: {fe}", exc_info=True)
                                continue
                            elif button_reply_id and (button_reply_id == "btn_admin" or button_reply_id.startswith("btn_admin")):
                                try:
                                    send_wa_message(phone, "Admin akan segera membalas pesan Anda. Mohon tunggu sebentar 🙏")
                                    inbox = load_inbox()
                                    inbox["contacts"][phone]["status"] = "perlu_dibalas"
                                    save_inbox(inbox)
                                    logger.info(f"Notifikasi admin dikirim ke {phone}")
                                except Exception as ae:
                                    logger.error(f"Gagal proses btn_admin untuk {phone}: {ae}", exc_info=True)
                                continue

                            text_lower = (text or "").lower()
                            if msg_type == "text" and any(kw in text_lower for kw in DONASI_KEYWORDS):
                                try:
                                    send_wa_flow_message(phone)
                                    logger.info(f"Flow donasi dikirim ke {phone} (keyword match)")
                                except Exception as fe:
                                    logger.error(f"Gagal kirim Flow ke {phone}: {fe}", exc_info=True)
                                continue

                            # Kirim Menu Utama jika pesan masuk pertama & belum di-resolve
                            if msg_type == "text":
                                is_new_contact = existing_contact is None
                                already_resolved = existing_contact and existing_contact.get("status") == "selesai"
                                menu_already_sent = existing_contact and existing_contact.get("menu_sent")
                                if (is_new_contact or not already_resolved) and not menu_already_sent:
                                    try:
                                        send_menu_utama(phone)
                                        inbox = load_inbox()
                                        inbox["contacts"][phone]["menu_sent"] = True
                                        save_inbox(inbox)
                                        logger.info(f"Menu Utama dikirim ke {phone}")
                                    except Exception as me:
                                        logger.error(f"Gagal kirim Menu Utama ke {phone}: {me}", exc_info=True)
            except Exception as we:
                logger.error(f"Error parsing webhook entry: {we}", exc_info=True)
            return jsonify({"status": "ok"}), 200

        response_data = handle_flow_request(body)
        return jsonify(response_data)
    except Exception as e:
        logger.error(f"Error wa-flow: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500

@app.route("/midtrans-callback", methods=["POST"])
def midtrans_callback():
    try:
        data = request.get_json()
        order_id = data.get("order_id")
        status = data.get("transaction_status")
        fraud = data.get("fraud_status", "accept")
        final_status = "lunas" if status in ("capture", "settlement") and fraud == "accept" else ("gagal" if status in ("cancel", "deny", "expire") else status)
        transaksi = load_data()
        donatur, kampanye, kampanye_id, nominal, phone = "", "", "", 0, ""
        for t in transaksi:
            if t["order_id"] == order_id:
                t["status"] = final_status
                t["paid_at"] = datetime.now().isoformat()
                donatur, kampanye, kampanye_id, nominal, phone = t.get("donatur",""), t.get("kampanye",""), t.get("kampanye_id",""), t.get("nominal",0), t.get("phone","")
                break
        save_data(transaksi)
        # AddToCart = transaksi masuk Midtrans (pending + lunas)
        if phone and final_status not in ("gagal",):
            try:
                send_pixel_event("AddToCart", phone=phone, value=nominal, currency="IDR",
                                  event_id="atc_midtrans_" + str(order_id), content_name=kampanye,
                                  content_ids=[kampanye_id] if kampanye_id else None,
                                  ctwa_clid=get_ctwa_for_phone(phone))
            except Exception as atce:
                logger.error("Pixel AddToCart midtrans gagal: " + str(atce))
        if final_status == "lunas" and phone:
            send_wa_message(phone, f"✅ *Donasi Berhasil!*\n\nAlhamdulillah donasi kakak sudah kami terima untuk,\n📋 *{kampanye}*\n💰 {format_rupiah(nominal)}\n🆔 {order_id}\n\nSemoga kakak beserta keluarga diberikan kesehatan selalu, segala urusannya selalu dilancarkan, dan apa yang telah di berikan membawa keberkahan untuk kakak beserta keluarga. Aamiin. 🤲\n_Raihmimpi.id_")
            notify_telegram(f"✅ <b>LUNAS!</b>\n👤 {donatur} ({phone})\n📋 {kampanye}\n💰 {format_rupiah(nominal)}\n🆔 {order_id}")
            send_pixel_event("Purchase", phone=phone, value=nominal, currency="IDR",
                              event_id=f"purchase_{order_id}", content_name=kampanye,
                              content_ids=[kampanye_id] if kampanye_id else None,
                              ctwa_clid=get_ctwa_for_phone(phone))
        return jsonify({"status": "ok"})
    except Exception as e:
        logger.error(f"Error midtrans-callback: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500

def fetch_midtrans_status(order_id):
    """Cek status transaksi dari Midtrans API Status endpoint."""
    import base64 as b64
    if not MIDTRANS_SERVER_KEY:
        return {"error": "MIDTRANS_SERVER_KEY belum di-set"}
    if "sandbox" in MIDTRANS_BASE_URL:
        status_url = f"https://api.sandbox.midtrans.com/v2/{order_id}/status"
    else:
        status_url = f"https://api.midtrans.com/v2/{order_id}/status"
    auth = b64.b64encode(f"{MIDTRANS_SERVER_KEY}:".encode()).decode()
    try:
        resp = requests.get(status_url, headers={"Authorization": f"Basic {auth}", "Accept": "application/json"}, timeout=15)
        return resp.json()
    except Exception as e:
        logger.error(f"fetch_midtrans_status error {order_id}: {e}")
        return {"error": str(e)}

def apply_midtrans_status(transaksi_list, order_id, mt_data):
    """Apply status Midtrans ke transaksi. Return (found, t_record, final_status, changed)."""
    status = mt_data.get("transaction_status")
    fraud = mt_data.get("fraud_status", "accept")
    if not status:
        return False, None, None, False
    final_status = "lunas" if status in ("capture", "settlement") and fraud == "accept" else ("gagal" if status in ("cancel", "deny", "expire", "failure") else "pending")
    for t in transaksi_list:
        if t.get("order_id") == order_id:
            old_status = t.get("status", "pending")
            if old_status == final_status:
                return True, t, final_status, False
            t["status"] = final_status
            if final_status == "lunas":
                t["paid_at"] = datetime.now().isoformat()
            t["midtrans_synced_at"] = datetime.now().isoformat()
            return True, t, final_status, True
    return False, None, final_status, False

def _send_lunas_notif(t, order_id, source="sync"):
    """Kirim WA + Telegram + Pixel saat transaksi jadi lunas."""
    if not t.get("phone"):
        return
    try:
        msg_wa = "✅ *Donasi Berhasil!*\n\nAlhamdulillah donasi kakak sudah kami terima untuk,\n📋 *" + str(t.get("kampanye","")) + "*\n💰 " + format_rupiah(t.get("nominal",0)) + "\n🆔 " + order_id + "\n\nSemoga kakak beserta keluarga diberikan kesehatan selalu, segala urusannya selalu dilancarkan, dan apa yang telah di berikan membawa keberkahan untuk kakak beserta keluarga. Aamiin. 🤲\n_Raihmimpi.id_"
        send_wa_message(t["phone"], msg_wa)
        msg_tg = "✅ <b>LUNAS (" + source + ")!</b>\n👤 " + str(t.get("donatur","")) + " (" + str(t.get("phone","")) + ")\n📋 " + str(t.get("kampanye","")) + "\n💰 " + format_rupiah(t.get("nominal",0)) + "\n🆔 " + order_id
        notify_telegram(msg_tg)
        send_pixel_event("Purchase", phone=t["phone"], value=t.get("nominal",0), currency="IDR",
                          event_id=f"purchase_{order_id}", content_name=t.get("kampanye",""),
                          content_ids=[t.get("kampanye_id")] if t.get("kampanye_id") else None,
                          ctwa_clid=get_ctwa_for_phone(t["phone"]))
    except Exception as e:
        logger.error(f"_send_lunas_notif error {order_id}: {e}")

def sync_all_pending(source="manual"):
    """Loop semua transaksi pending, sync dengan Midtrans."""
    transaksi = load_data()
    pending_list = [t for t in transaksi if t.get("status") == "pending"]
    if not pending_list:
        return {"total_pending": 0, "checked": 0, "changed": 0, "lunas": 0, "gagal": 0, "errors": 0, "details": []}
    logger.info(f"sync_all_pending ({source}): {len(pending_list)} pending")
    results = {"total_pending": len(pending_list), "checked": 0, "changed": 0, "lunas": 0, "gagal": 0, "errors": 0, "details": []}
    any_change = False
    for t in pending_list:
        order_id = t.get("order_id")
        if not order_id:
            continue
        results["checked"] += 1
        mt_data = fetch_midtrans_status(order_id)
        if mt_data.get("error") and not mt_data.get("transaction_status"):
            results["errors"] += 1
            results["details"].append({"order_id": order_id, "error": mt_data.get("error")})
            continue
        found, t_ref, final_status, changed = apply_midtrans_status(transaksi, order_id, mt_data)
        if changed:
            any_change = True
            results["changed"] += 1
            if final_status == "lunas":
                results["lunas"] += 1
                _send_lunas_notif(t_ref, order_id, source=source)
            elif final_status == "gagal":
                results["gagal"] += 1
        results["details"].append({"order_id": order_id, "status": final_status, "changed": changed})
    if any_change:
        save_data(transaksi)
    logger.info(f"sync_all_pending ({source}) done: {results['changed']} changed, {results['lunas']} lunas, {results['gagal']} gagal, {results['errors']} errors")
    return results

@app.route("/api/sync-midtrans/<order_id>", methods=["POST"])
def api_sync_midtrans_one(order_id):
    mt_data = fetch_midtrans_status(order_id)
    if mt_data.get("error") and not mt_data.get("transaction_status"):
        return jsonify({"order_id": order_id, "error": mt_data.get("error"), "midtrans_raw": mt_data}), 400
    transaksi = load_data()
    found, t, final_status, changed = apply_midtrans_status(transaksi, order_id, mt_data)
    if not found:
        return jsonify({"order_id": order_id, "error": "Transaksi tidak ditemukan", "midtrans_raw": mt_data}), 404
    if changed:
        save_data(transaksi)
        logger.info(f"sync-midtrans {order_id}: {final_status} (changed)")
        if final_status == "lunas":
            _send_lunas_notif(t, order_id, source="manual")
    return jsonify({"order_id": order_id, "status": final_status, "changed": changed, "midtrans_transaction_status": mt_data.get("transaction_status"), "midtrans_fraud_status": mt_data.get("fraud_status")})

@app.route("/api/sync-midtrans-all", methods=["POST"])
def api_sync_midtrans_all():
    return jsonify(sync_all_pending(source="manual"))

def _midtrans_poll_worker():
    import time
    time.sleep(60)
    while True:
        try:
            sync_all_pending(source="auto-poll")
        except Exception as e:
            logger.error(f"_midtrans_poll_worker error: {e}", exc_info=True)
        time.sleep(300)

def start_midtrans_poller():
    import threading
    t = threading.Thread(target=_midtrans_poll_worker, daemon=True, name="midtrans-poll")
    t.start()
    logger.info("Midtrans auto-polling started (interval: 5 menit)")

if os.environ.get("ENABLE_MIDTRANS_POLL", "1") == "1":
    start_midtrans_poller()

@app.route("/api/kampanye-source", methods=["GET"])
def list_kampanye():
    campaigns = get_campaigns(full=True)
    return jsonify({"total": len(campaigns), "formatted": format_campaigns_for_flow(campaigns)})

@app.route("/", methods=["GET"])
def health():
    return jsonify({"status": "ok", "service": "Raihmimpi WA Flow Backend", "version": "3.2.0"})



@app.route("/api/inbox", methods=["GET"])
def api_inbox_list():
    """List semua kontak/percakapan, diurutkan dari pesan terbaru."""
    inbox = load_inbox()
    contacts = list(inbox.get("contacts", {}).values())
    contacts_sorted = sorted(contacts, key=lambda c: c.get("last_message_at", ""), reverse=True)
    return jsonify({"contacts": contacts_sorted})

@app.route("/api/inbox/<phone>", methods=["GET"])
def api_inbox_messages(phone):
    """Ambil semua pesan untuk satu kontak, dan reset unread counter."""
    inbox = load_inbox()
    messages = inbox.get("messages", {}).get(phone, [])
    contact = inbox.get("contacts", {}).get(phone, {"phone": phone, "name": phone, "labels": []})

    # Reset unread saat dibuka
    if inbox.get("contacts", {}).get(phone, {}).get("unread", 0) > 0:
        inbox["contacts"][phone]["unread"] = 0
        save_inbox(inbox)

    return jsonify({"contact": contact, "messages": messages})

@app.route("/api/inbox/<phone>/reply", methods=["POST"])
def api_inbox_reply(phone):
    """Kirim balasan teks ke kontak via WhatsApp Cloud API (Graph), dan simpan ke inbox."""
    body = request.get_json(silent=True) or {}
    text = body.get("text", "").strip()
    if not text:
        return jsonify({"error": "text wajib diisi"}), 400
    try:
        resp = send_wa_message(phone, text)
        record_outgoing_message(phone, text, msg_type="text")
        return jsonify({"status": "sent", "wa_status": resp.status_code, "wa_body": resp.text[:300]})
    except Exception as e:
        logger.error(f"api_inbox_reply error: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500

@app.route("/api/send-flow/<phone>", methods=["POST"])
def api_send_flow(phone):
    """Kirim Flow donasi ke kontak via attachment menu."""
    try:
        resp = send_wa_flow_message(phone)
        record_outgoing_message(phone, "📋 Flow Donasi · Dikirim", msg_type="text")
        return jsonify({"status": "sent", "wa_status": resp.status_code})
    except Exception as e:
        logger.error(f"api_send_flow error: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500

@app.route("/api/send-media/<phone>", methods=["POST"])
def api_send_media(phone):
    """Upload dan kirim media (document/image/video/audio) ke kontak via WhatsApp Cloud API."""
    try:
        file = request.files.get("file")
        media_type = request.form.get("media_type", "document")
        if not file:
            return jsonify({"error": "file wajib diisi"}), 400

        filename = file.filename or "file"
        mime = file.content_type or "application/octet-stream"
        file_bytes = file.read()

        # Step 1: Upload media ke Meta
        upload_url = f"https://graph.facebook.com/v22.0/{WA_PHONE_NUMBER_ID}/media"
        upload_resp = requests.post(upload_url,
            headers={"Authorization": f"Bearer {WA_TOKEN}"},
            files={"file": (filename, file_bytes, mime)},
            data={"messaging_product": "whatsapp", "type": mime},
            timeout=30)
        if not upload_resp.ok:
            logger.error(f"Media upload failed: {upload_resp.text[:300]}")
            return jsonify({"error": "Upload media gagal", "detail": upload_resp.text[:200]}), 500
        media_id = upload_resp.json().get("id")

        # Step 2: Kirim pesan media
        if media_type == "image":
            msg_payload = {"type": "image", "image": {"id": media_id}}
            label = f"🖼 {filename}"
        elif media_type == "audio":
            msg_payload = {"type": "audio", "audio": {"id": media_id}}
            label = f"🎵 {filename}"
        else:
            msg_payload = {"type": "document", "document": {"id": media_id, "filename": filename}}
            label = f"📄 {filename}"

        send_url = f"https://graph.facebook.com/v22.0/{WA_PHONE_NUMBER_ID}/messages"
        send_resp = requests.post(send_url,
            headers={"Authorization": f"Bearer {WA_TOKEN}", "Content-Type": "application/json"},
            json={"messaging_product": "whatsapp", "to": phone, "recipient_type": "individual", **msg_payload},
            timeout=15)
        logger.info(f"api_send_media to={phone} type={media_type} file={filename} status={send_resp.status_code}")
        record_outgoing_message(phone, label, msg_type=media_type)
        return jsonify({"status": "sent", "wa_status": send_resp.status_code})
    except Exception as e:
        logger.error(f"api_send_media error: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500

@app.route("/api/inbox/<phone>/label", methods=["POST"])
def api_inbox_label(phone):
    """Update label dan/atau status kontak."""
    body = request.get_json(silent=True) or {}
    inbox = load_inbox()
    contact = inbox.get("contacts", {}).get(phone)
    if not contact:
        return jsonify({"error": "contact not found"}), 404
    if "labels" in body:
        contact["labels"] = body["labels"]
    if "status" in body:
        contact["status"] = body["status"]
    if "name" in body:
        contact["name"] = body["name"]
    inbox["contacts"][phone] = contact
    save_inbox(inbox)
    return jsonify(contact)

@app.route("/api/inbox/<phone>/reset-menu", methods=["POST"])
def api_inbox_reset_menu(phone):
    inbox = load_inbox()
    contact = inbox.get("contacts", {}).get(phone)
    if not contact:
        return jsonify({"error": "contact not found"}), 404
    contact["menu_sent"] = False
    save_inbox(inbox)
    logger.info(f"menu_sent direset untuk {phone}")
    return jsonify({"status": "ok", "phone": phone})

@app.route("/api/inbox/reset-menu-all", methods=["POST"])
def api_inbox_reset_menu_all():
    inbox = load_inbox()
    count = 0
    for phone, contact in inbox.get("contacts", {}).items():
        if contact.get("menu_sent"):
            contact["menu_sent"] = False
            count += 1
    save_inbox(inbox)
    logger.info(f"menu_sent direset untuk {count} kontak")
    return jsonify({"status": "ok", "reset_count": count})

@app.route("/api/settings/menu-utama", methods=["GET"])
def api_get_menu_utama():
    settings = load_settings()
    return jsonify(settings.get("menu_utama", DEFAULT_SETTINGS["menu_utama"]))

@app.route("/api/settings/menu-utama", methods=["POST"])
def api_save_menu_utama():
    body = request.get_json(silent=True) or {}
    settings = load_settings()
    settings["menu_utama"] = {
        "message": body.get("message", ""),
        "buttons": body.get("buttons", []),
    }
    save_settings(settings)
    return jsonify(settings["menu_utama"])

# ============================================================
# BLAST HISTORY FILE
# ============================================================
BLAST_FILE = os.environ.get("DATA_DIR", "/tmp") + "/blast_history.json"

def load_blast_history():
    if not os.path.exists(BLAST_FILE):
        return []
    try:
        with open(BLAST_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return []

def save_blast_history(data):
    with open(BLAST_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

@app.route("/api/blast", methods=["POST"])
def api_blast():
    """Kirim WA Blast via Meta Template API."""
    try:
        body = request.get_json(silent=True) or {}
        phones = body.get("phones", [])
        template_name = body.get("template_name", "")
        template_language = body.get("template_language", "id")
        judul = body.get("judul", "")
        judul_campaign = body.get("judul_campaign", "")
        kategori = body.get("kategori", "MARKETING")
        header_type = body.get("header_type", "")
        body_params = body.get("body_params", [])

        if not phones:
            return jsonify({"error": "phones wajib diisi"}), 400
        if not template_name:
            return jsonify({"error": "template_name wajib diisi"}), 400

        blast_id = f"blast_{int(datetime.now().timestamp()*1000)}"
        results = []

        for phone in phones:
            try:
                # Build template payload
                components = []
                if body_params:
                    components.append({
                        "type": "body",
                        "parameters": [{"type": "text", "text": p} for p in body_params]
                    })

                payload = {
                    "messaging_product": "whatsapp",
                    "to": phone,
                    "type": "template",
                    "template": {
                        "name": template_name,
                        "language": {"code": template_language},
                    }
                }
                if components:
                    payload["template"]["components"] = components

                resp = requests.post(
                    f"https://graph.facebook.com/v22.0/{WA_PHONE_NUMBER_ID}/messages",
                    headers={"Authorization": f"Bearer {WA_ACCESS_TOKEN}", "Content-Type": "application/json"},
                    json=payload, timeout=15
                )
                wa_resp = resp.json()
                status = "sent" if resp.ok else "error"
                record_outgoing_message(phone, f"📢 Blast: {template_name}", msg_type="text")
                results.append({"phone": phone, "status": status, "wa_id": wa_resp.get("messages", [{}])[0].get("id", "")})
            except Exception as e:
                logger.error(f"api_blast error for {phone}: {e}", exc_info=True)
                results.append({"phone": phone, "status": "error", "error": str(e)})

        sent_count = sum(1 for r in results if r["status"] == "sent")

        # Simpan history
        history = load_blast_history()
        history.insert(0, {
            "id": blast_id,
            "judul": judul,
            "judul_campaign": judul_campaign,
            "kategori": kategori,
            "template_name": template_name,
            "template_language": template_language,
            "header_type": header_type,
            "phones": phones,
            "total": len(phones),
            "sent": sent_count,
            "in_progress": 0,
            "delivered": 0,
            "read": 0,
            "failed": len(phones) - sent_count,
            "status": "DONE",
            "created_at": datetime.now().isoformat(),
            "results": results
        })
        save_blast_history(history)

        return jsonify({"blast_id": blast_id, "total": len(phones), "sent": sent_count, "results": results})
    except Exception as e:
        logger.error(f"api_blast error: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500

# ============================================================
# USERS SYSTEM
# ============================================================
USERS_FILE = os.environ.get("DATA_DIR", "/tmp") + "/users.json"

def hash_password(password):
    return hashlib.sha256(password.encode()).hexdigest()

def load_users():
    if not os.path.exists(USERS_FILE):
        # Buat admin default pertama kali
        default = [{
            "id": "user_1",
            "email": "admin@raihmimpi.id",
            "password": hash_password("admin123"),
            "nama": "Administrator",
            "role": "ADMINISTRATOR",
            "status": "aktif",
            "created_at": datetime.now().isoformat(),
            "last_login": None
        }]
        save_users(default)
        return default
    try:
        with open(USERS_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return []

def save_users(data):
    with open(USERS_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def login_required(f):
    @functools.wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("user_id"):
            # API endpoints return JSON 401, page endpoints redirect
            if request.path.startswith("/api/"):
                return jsonify({"error": "Unauthorized", "redirect": "/login"}), 401
            return redirect("/login")
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @functools.wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("user_id"):
            return redirect("/login")
        if session.get("user_role") != "ADMINISTRATOR":
            return Response("Akses ditolak - hanya Administrator", status=403)
        return f(*args, **kwargs)
    return decorated

@app.route("/login", methods=["GET", "POST"])
def login_page():
    if session.get("user_id"):
        return redirect("/dashboard")
    error = ""
    if request.method == "POST":
        email = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "")
        users = load_users()
        user = next((u for u in users if u["email"].lower() == email and u["status"] == "aktif"), None)
        if user and user["password"] == hash_password(password):
            session["user_id"] = user["id"]
            session["user_email"] = user["email"]
            session["user_nama"] = user["nama"]
            session["user_role"] = user["role"]
            # Update last login
            users = load_users()
            for u in users:
                if u["id"] == user["id"]:
                    u["last_login"] = datetime.now().isoformat()
            save_users(users)
            return redirect("/dashboard")
        else:
            error = "Email atau password salah, atau akun tidak aktif."
    return Response(f"""<!DOCTYPE html>
<html lang="id">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0, maximum-scale=1.0, user-scalable=no">
<title>Masuk - Raihmimpi</title>
<style>
  * {{ box-sizing:border-box; margin:0; padding:0; }}
  body {{ min-height:100vh; background:#f0edff; display:flex; align-items:center; justify-content:center; font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif; padding:20px; }}
  .card {{ background:#fff; border-radius:24px; padding:40px 36px; width:100%; max-width:420px; box-shadow:0 4px 32px rgba(91,61,240,.12); }}
  .logo {{ text-align:center; margin-bottom:28px; }}
  .logo img {{ height:48px; object-fit:contain; }}
  h2 {{ text-align:center; font-size:24px; font-weight:800; color:#1f2937; margin-bottom:28px; }}
  .form-group {{ margin-bottom:16px; }}
  label {{ display:block; font-size:13px; font-weight:600; color:#374151; margin-bottom:6px; }}
  input[type=email], input[type=password], input[type=text] {{
    width:100%; padding:12px 16px; border:1.5px solid #e5e7eb; border-radius:12px;
    font-size:15px; outline:none; transition:border .2s; background:#f9fafb; color:#1f2937;
  }}
  input:focus {{ border-color:#5b3df0; background:#fff; }}
  .password-wrap {{ position:relative; }}
  .password-wrap input {{ padding-right:44px; }}
  .eye-btn {{ position:absolute; right:14px; top:50%; transform:translateY(-50%); background:none; border:none; cursor:pointer; color:#9ca3af; font-size:18px; }}
  .error {{ background:#fef2f2; border:1px solid #fca5a5; color:#dc2626; padding:10px 14px; border-radius:10px; font-size:13px; margin-bottom:16px; }}
  .remember {{ display:flex; align-items:center; gap:8px; font-size:13px; color:#6b7280; margin-bottom:20px; }}
  .btn-login {{ width:100%; padding:14px; background:#5b3df0; color:#fff; border:none; border-radius:14px; font-size:16px; font-weight:700; cursor:pointer; transition:background .2s; }}
  .btn-login:hover {{ background:#4c30d9; }}
  .footer-links {{ text-align:center; margin-top:20px; font-size:13px; color:#9ca3af; }}
</style>
</head>
<body>
  <div class="card">
    <div class="logo"><img src="/static/IconRM.png" alt="Raihmimpi"></div>
    <h2>Masuk</h2>
    {'<div class="error">' + error + '</div>' if error else ''}
    <form method="POST">
      <div class="form-group">
        <label>Email</label>
        <input type="email" name="email" placeholder="email@raihmimpi.id" required autofocus>
      </div>
      <div class="form-group">
        <label>Password</label>
        <div class="password-wrap">
          <input type="password" name="password" id="pwdInput" placeholder="Password" required>
          <button type="button" class="eye-btn" onclick="togglePwd()">👁</button>
        </div>
      </div>
      <div class="remember">
        <input type="checkbox" id="remember"> <label for="remember" style="margin:0;font-weight:400;">Ingatkan saya</label>
      </div>
      <button type="submit" class="btn-login">Masuk</button>
    </form>
    <div class="footer-links">Raihmimpi &copy; 2024</div>
  </div>
  <script>
    function togglePwd() {{
      const i = document.getElementById("pwdInput");
      i.type = i.type === "password" ? "text" : "password";
    }}
  </script>
</body>
</html>""", mimetype="text/html")

@app.route("/laporan/tracking-campaign", methods=["GET"])
@login_required
def laporan_tracking_campaign():
    today = datetime.now().strftime("%Y-%m-%d")
    first_month = datetime.now().strftime("%Y-%m-01")
    body = f"""
  <!-- Filter -->
  <div style="background:#fff;border-radius:12px;padding:20px;margin-bottom:16px;box-shadow:0 1px 3px rgba(0,0,0,.06);">
    <div style="display:flex;gap:12px;align-items:flex-end;flex-wrap:wrap;">
      <div>
        <div style="font-size:12px;font-weight:600;color:#6b7280;margin-bottom:6px;">Tracking Period</div>
        <div style="display:flex;gap:8px;align-items:center;">
          <input type="date" id="filterFrom" value="{first_month}" style="padding:8px 12px;border:1px solid #d1d5db;border-radius:8px;font-size:13px;">
          <span style="color:#9ca3af;">→</span>
          <input type="date" id="filterTo" value="{today}" style="padding:8px 12px;border:1px solid #d1d5db;border-radius:8px;font-size:13px;">
        </div>
      </div>
      <div>
        <div style="font-size:12px;font-weight:600;color:#6b7280;margin-bottom:6px;">Campaign Name</div>
        <input type="text" id="searchCampaign" placeholder="Cari campaign..." style="padding:8px 12px;border:1px solid #d1d5db;border-radius:8px;font-size:13px;min-width:180px;">
      </div>
      <button onclick="loadTracking()" style="padding:8px 20px;background:#5b3df0;color:#fff;border:none;border-radius:8px;font-size:13px;font-weight:600;cursor:pointer;">Submit</button>
      <button onclick="exportTracking()" style="padding:8px 20px;background:#fff;color:#5b3df0;border:1px solid #5b3df0;border-radius:8px;font-size:13px;font-weight:600;cursor:pointer;">Export to Excel</button>
    </div>
  </div>

  <!-- List View -->
  <div id="trackingListView">
    <div style="background:#fff;border-radius:12px;overflow:hidden;box-shadow:0 1px 3px rgba(0,0,0,.06);">
      <div style="overflow-x:auto;">
        <table style="width:100%;border-collapse:collapse;min-width:800px;">
          <thead>
            <tr style="background:#f9fafb;">
              <th style="padding:12px 16px;text-align:left;font-size:12px;color:#6b7280;font-weight:600;border-bottom:1px solid #e5e7eb;">Campaign Name</th>
              <th style="padding:12px 16px;text-align:left;font-size:12px;color:#6b7280;font-weight:600;border-bottom:1px solid #e5e7eb;">Ad ID</th>
              <th style="padding:12px 16px;text-align:left;font-size:12px;color:#6b7280;font-weight:600;border-bottom:1px solid #e5e7eb;">Type</th>
              <th style="padding:12px 16px;text-align:left;font-size:12px;color:#6b7280;font-weight:600;border-bottom:1px solid #e5e7eb;">Channel</th>
              <th style="padding:12px 16px;text-align:center;font-size:12px;color:#6b7280;font-weight:600;border-bottom:1px solid #e5e7eb;">Total Contacts</th>
              <th style="padding:12px 16px;text-align:center;font-size:12px;color:#6b7280;font-weight:600;border-bottom:1px solid #e5e7eb;">Lead Submitted</th>
              <th style="padding:12px 16px;text-align:center;font-size:12px;color:#6b7280;font-weight:600;border-bottom:1px solid #e5e7eb;">View Content</th>
              <th style="padding:12px 16px;text-align:center;font-size:12px;color:#6b7280;font-weight:600;border-bottom:1px solid #e5e7eb;">Add to Cart</th>
              <th style="padding:12px 16px;text-align:center;font-size:12px;color:#6b7280;font-weight:600;border-bottom:1px solid #e5e7eb;">Purchase</th>
            </tr>
          </thead>
          <tbody id="trackingBody">
            <tr><td colspan="8" style="padding:40px;text-align:center;color:#9ca3af;">Klik Submit untuk memuat data</td></tr>
          </tbody>
        </table>
      </div>
    </div>
  </div>

  <!-- Detail View -->
  <div id="trackingDetailView" style="display:none;">
    <div style="display:flex;align-items:center;gap:12px;margin-bottom:16px;">
      <button onclick="backToTracking()" style="background:none;border:none;cursor:pointer;color:#5b3df0;font-size:20px;font-weight:700;">←</button>
      <div>
        <div style="font-size:18px;font-weight:800;" id="detailCampaignName"></div>
        <div style="font-size:13px;color:#6b7280;" id="detailCampaignPeriode"></div>
      </div>
      <button onclick="exportDetailTracking()" style="margin-left:auto;padding:8px 16px;background:#fff;color:#5b3df0;border:1px solid #5b3df0;border-radius:8px;font-size:13px;font-weight:600;cursor:pointer;">Export to Excel</button>
    </div>
    <div style="background:#fff;border-radius:12px;overflow:hidden;box-shadow:0 1px 3px rgba(0,0,0,.06);">
      <table style="width:100%;border-collapse:collapse;">
        <thead>
          <tr style="background:#f9fafb;">
            <th style="padding:12px 16px;text-align:left;font-size:12px;color:#6b7280;font-weight:600;border-bottom:1px solid #e5e7eb;">Nama</th>
            <th style="padding:12px 16px;text-align:left;font-size:12px;color:#6b7280;font-weight:600;border-bottom:1px solid #e5e7eb;">No. Handphone</th>
            <th style="padding:12px 16px;text-align:left;font-size:12px;color:#6b7280;font-weight:600;border-bottom:1px solid #e5e7eb;">First Interaction Message</th>
            <th style="padding:12px 16px;text-align:left;font-size:12px;color:#6b7280;font-weight:600;border-bottom:1px solid #e5e7eb;">Tanggal Interaction</th>
          </tr>
        </thead>
        <tbody id="trackingDetailBody"></tbody>
      </table>
      <div style="padding:12px 16px;font-size:13px;color:#6b7280;border-top:1px solid #f3f4f6;" id="trackingDetailInfo"></div>
    </div>
  </div>

<script>
var _trackingData = [];
var _currentCampaign = null;

function loadTracking() {{
  var df = document.getElementById("filterFrom").value;
  var dt = document.getElementById("filterTo").value;
  var sc = document.getElementById("searchCampaign").value;
  var tbody = document.getElementById("trackingBody");
  tbody.innerHTML = "<tr><td colspan=8 style='padding:20px;text-align:center;color:#9ca3af;'>Memuat...</td></tr>";

  fetch("/api/laporan/tracking-campaign?from=" + df + "&to=" + dt + "&campaign=" + encodeURIComponent(sc))
  .then(function(r) {{ return r.json(); }})
  .then(function(json) {{
    _trackingData = json.campaigns || [];
    if (!_trackingData.length) {{
      tbody.innerHTML = "<tr><td colspan=8 style='padding:40px;text-align:center;color:#9ca3af;'>Tidak ada data campaign.</td></tr>";
      return;
    }}
    var html = "";
    var totLeads=0, totAtc=0, totPurchase=0, totContacts=0, totVc=0;
    for (var i=0; i<_trackingData.length; i++) {{
      var c = _trackingData[i];
      totLeads += c.total_leads;
      totAtc += c.total_add_to_cart;
      totPurchase += c.total_purchase;
      totContacts += c.total_contacts;
      totVc += (c.total_view_content||0);
      html += "<tr style='border-bottom:1px solid #f3f4f6;'>";
      html += "<td style='padding:12px 16px;'><span onclick='showTrackingDetail(" + i + ")' style='color:#5b3df0;cursor:pointer;font-weight:600;text-decoration:underline;'>" + c.campaign_name + "</span></td>";
      html += "<td style='padding:12px 16px;font-size:12px;color:#6b7280;'>" + (c.ad_id||"-") + "</td>";
      html += "<td style='padding:12px 16px;font-size:13px;'>" + c.type + "</td>";
      html += "<td style='padding:12px 16px;font-size:13px;'>" + c.channel + "</td>";
      html += "<td style='padding:12px 16px;text-align:center;font-size:13px;font-weight:600;'>" + c.total_contacts + "</td>";
      html += "<td style='padding:12px 16px;text-align:center;font-size:13px;color:#5b3df0;font-weight:600;'>" + c.total_leads + "</td>";
      html += "<td style='padding:12px 16px;text-align:center;font-size:13px;color:#6b7280;'>" + (c.total_view_content||0) + "</td>";
      html += "<td style='padding:12px 16px;text-align:center;font-size:13px;color:#d97706;'>" + c.total_add_to_cart + "</td>";
      html += "<td style='padding:12px 16px;text-align:center;font-size:13px;color:#16a34a;font-weight:600;'>" + c.total_purchase + "</td>";
      html += "</tr>";
    }}
    // Total
    html += "<tr style='background:#f9fafb;font-weight:700;border-top:2px solid #e5e7eb;'>";
    html += "<td style='padding:12px 16px;font-size:13px;' colspan=4>TOTAL</td>";
    html += "<td style='padding:12px 16px;text-align:center;font-size:13px;'>" + totContacts + "</td>";
    html += "<td style='padding:12px 16px;text-align:center;font-size:13px;color:#5b3df0;'>" + totLeads + "</td>";
    html += "<td style='padding:12px 16px;text-align:center;font-size:13px;color:#6b7280;'>" + totVc + "</td>";
    html += "<td style='padding:12px 16px;text-align:center;font-size:13px;color:#d97706;'>" + totAtc + "</td>";
    html += "<td style='padding:12px 16px;text-align:center;font-size:13px;color:#16a34a;'>" + totPurchase + "</td>";
    html += "</tr>";
    tbody.innerHTML = html;
  }})
  .catch(function() {{
    tbody.innerHTML = "<tr><td colspan=8 style='color:#dc2626;padding:20px;'>Gagal memuat.</td></tr>";
  }});
}}

function showTrackingDetail(idx) {{
  var c = _trackingData[idx];
  if (!c) return;
  _currentCampaign = c;
  document.getElementById("trackingListView").style.display = "none";
  document.getElementById("trackingDetailView").style.display = "block";
  document.getElementById("detailCampaignName").textContent = "Laporan Tracking Campaign - " + c.campaign_name;
  var df = document.getElementById("filterFrom").value;
  var dt = document.getElementById("filterTo").value;
  document.getElementById("detailCampaignPeriode").textContent = "Periode: " + df + " sampai " + dt;
  var contacts = c.contacts || [];
  var tbody = document.getElementById("trackingDetailBody");
  tbody.innerHTML = contacts.map(function(ct) {{
    return "<tr style='border-bottom:1px solid #f3f4f6;'>" +
      "<td style='padding:12px 16px;font-size:13px;font-weight:600;'>" + ct.nama + "</td>" +
      "<td style='padding:12px 16px;font-size:13px;color:#6b7280;'>" + ct.phone + "</td>" +
      "<td style='padding:12px 16px;font-size:13px;color:#6b7280;'>" + (ct.first_msg||"-") + "</td>" +
      "<td style='padding:12px 16px;font-size:13px;color:#6b7280;'>" + (ct.first_date||"-") + "</td>" +
      "</tr>";
  }}).join("");
  document.getElementById("trackingDetailInfo").textContent = "Menampilkan " + contacts.length + " kontak";
}}

function backToTracking() {{
  document.getElementById("trackingDetailView").style.display = "none";
  document.getElementById("trackingListView").style.display = "block";
}}

function exportTracking() {{
  if (!_trackingData.length) {{ alert("Tidak ada data. Klik Submit dulu."); return; }}
  var rows = [["Campaign Name","Ad ID","Type","Channel","Total Contacts","Lead Submitted","View Content","Add to Cart","Purchase"]];
  _trackingData.forEach(function(c) {{
    rows.push([c.campaign_name, c.ad_id, c.type, c.channel, c.total_contacts, c.total_leads, c.total_view_content||0, c.total_add_to_cart, c.total_purchase]);
  }});
  var csv = rows.map(function(r) {{ return r.map(function(v) {{ return '"'+String(v).replace(/"/g,'""')+'"'; }}).join(","); }}).join(String.fromCharCode(10));
  var blob = new Blob([csv], {{type:"text/csv;charset=utf-8;"}});
  var url = URL.createObjectURL(blob);
  var a = document.createElement("a"); a.href=url; a.download="tracking_campaign.csv"; a.click();
  URL.revokeObjectURL(url);
}}

function exportDetailTracking() {{
  if (!_currentCampaign) return;
  var contacts = _currentCampaign.contacts || [];
  var rows = [["Nama","No. Handphone","First Interaction Message","Tanggal Interaction"]];
  contacts.forEach(function(c) {{ rows.push([c.nama, c.phone, c.first_msg||"", c.first_date||""]); }});
  var csv = rows.map(function(r) {{ return r.map(function(v) {{ return '"'+String(v).replace(/"/g,'""')+'"'; }}).join(","); }}).join(String.fromCharCode(10));
  var blob = new Blob([csv], {{type:"text/csv;charset=utf-8;"}});
  var url = URL.createObjectURL(blob);
  var a = document.createElement("a"); a.href=url; a.download="detail_"+_currentCampaign.campaign_name+".csv"; a.click();
  URL.revokeObjectURL(url);
}}

loadTracking();
</script>
"""
    return Response(render_page("laporan-tracking", "Laporan Tracking Campaign", "", body), mimetype="text/html")

@app.route("/laporan/contact-label", methods=["GET"])
@login_required
def laporan_contact_label():
    today = datetime.now().strftime("%Y-%m-%d")
    first_month = datetime.now().strftime("%Y-%m-01")
    # Ambil kategori unik dari labels
    labels_cfg = load_labels()
    kategoris = sorted(set(l.get("category","") for l in labels_cfg if l.get("category")))
    kat_options = "".join(f'<option value="{k}">{k}</option>' for k in kategoris)
    body = f"""
  <!-- Header -->
  <div style="background:#fff;border-radius:12px;padding:20px;margin-bottom:16px;box-shadow:0 1px 3px rgba(0,0,0,.06);">
    <div style="display:flex;gap:16px;flex-wrap:wrap;align-items:flex-end;">
      <div>
        <div style="font-size:12px;font-weight:600;color:#6b7280;margin-bottom:6px;">Nama Contact Label</div>
        <input type="text" id="searchLabel" placeholder="Cari nama label..." oninput="filterTable()"
          style="padding:8px 12px;border:1px solid #d1d5db;border-radius:8px;font-size:13px;min-width:200px;">
      </div>
      <div>
        <div style="font-size:12px;font-weight:600;color:#6b7280;margin-bottom:6px;">Periode</div>
        <div style="display:flex;gap:8px;align-items:center;">
          <input type="date" id="filterFrom" value="{first_month}" style="padding:8px 12px;border:1px solid #d1d5db;border-radius:8px;font-size:13px;">
          <span style="color:#9ca3af;">→</span>
          <input type="date" id="filterTo" value="{today}" style="padding:8px 12px;border:1px solid #d1d5db;border-radius:8px;font-size:13px;">
        </div>
      </div>
      <div>
        <div style="font-size:12px;font-weight:600;color:#6b7280;margin-bottom:6px;">Kategori</div>
        <select id="filterKategori" onchange="loadContactLabel()" style="padding:8px 12px;border:1px solid #d1d5db;border-radius:8px;font-size:13px;min-width:140px;">
          <option value="">Semua</option>
          {kat_options}
        </select>
      </div>
      <button onclick="loadContactLabel()" style="padding:8px 20px;background:#5b3df0;color:#fff;border:none;border-radius:8px;font-size:13px;font-weight:600;cursor:pointer;">Cari</button>
      <button onclick="exportContactLabel()" style="padding:8px 20px;background:#fff;color:#5b3df0;border:1px solid #5b3df0;border-radius:8px;font-size:13px;font-weight:600;cursor:pointer;">Export to Excel</button>
    </div>
  </div>

  <!-- Summary card -->
  <div style="background:#fff;border-radius:10px;padding:16px 24px;box-shadow:0 1px 3px rgba(0,0,0,.06);margin-bottom:16px;display:inline-block;">
    <div style="font-size:12px;color:#6b7280;margin-bottom:4px;">Total Label</div>
    <div style="font-size:28px;font-weight:800;color:#5b3df0;" id="totalLabelCount">-</div>
  </div>

  <!-- Tabel List Label -->
  <div id="labelListView">
    <div style="background:#fff;border-radius:12px;overflow:hidden;box-shadow:0 1px 3px rgba(0,0,0,.06);">
      <table style="width:100%;border-collapse:collapse;">
        <thead>
          <tr style="background:#f9fafb;">
            <th style="padding:12px 16px;text-align:left;font-size:12px;color:#6b7280;font-weight:600;border-bottom:1px solid #e5e7eb;">Nama Contact Label</th>
            <th style="padding:12px 16px;text-align:left;font-size:12px;color:#6b7280;font-weight:600;border-bottom:1px solid #e5e7eb;">Kategori</th>
            <th style="padding:12px 16px;text-align:center;font-size:12px;color:#6b7280;font-weight:600;border-bottom:1px solid #e5e7eb;">Total Contact</th>
          </tr>
        </thead>
        <tbody id="labelListBody">
          <tr><td colspan="3" style="padding:40px;text-align:center;color:#9ca3af;">Klik Cari untuk memuat data</td></tr>
        </tbody>
      </table>
    </div>
  </div>

  <!-- Detail Label View -->
  <div id="labelDetailView" style="display:none;">
    <div style="display:flex;align-items:center;gap:12px;margin-bottom:16px;">
      <button onclick="backToList()" style="background:none;border:none;cursor:pointer;color:#5b3df0;font-size:20px;font-weight:700;">←</button>
      <div>
        <div style="font-size:14px;font-weight:700;display:flex;align-items:center;gap:8px;" id="detailLabelName"></div>
        <div style="font-size:13px;color:#6b7280;" id="detailPeriode"></div>
      </div>
      <button onclick="exportDetailLabel()" style="margin-left:auto;padding:8px 16px;background:#fff;color:#5b3df0;border:1px solid #5b3df0;border-radius:8px;font-size:13px;font-weight:600;cursor:pointer;">Export to Excel</button>
    </div>
    <div style="background:#fff;border-radius:12px;overflow:hidden;box-shadow:0 1px 3px rgba(0,0,0,.06);">
      <table style="width:100%;border-collapse:collapse;">
        <thead>
          <tr style="background:#f9fafb;">
            <th style="padding:12px 16px;text-align:left;font-size:12px;color:#6b7280;font-weight:600;border-bottom:1px solid #e5e7eb;">Nama</th>
            <th style="padding:12px 16px;text-align:left;font-size:12px;color:#6b7280;font-weight:600;border-bottom:1px solid #e5e7eb;">No WhatsApp</th>
            <th style="padding:12px 16px;text-align:left;font-size:12px;color:#6b7280;font-weight:600;border-bottom:1px solid #e5e7eb;">Type</th>
          </tr>
        </thead>
        <tbody id="labelDetailBody">
        </tbody>
      </table>
      <div style="padding:12px 16px;font-size:13px;color:#6b7280;border-top:1px solid #f3f4f6;" id="detailInfo"></div>
    </div>
  </div>

<script>
var _clData = [];
var _currentLabel = null;

function loadContactLabel() {{
  var df = document.getElementById("filterFrom").value;
  var dt = document.getElementById("filterTo").value;
  var kat = document.getElementById("filterKategori").value;
  var tbody = document.getElementById("labelListBody");
  tbody.innerHTML = "<tr><td colspan=3 style='padding:20px;text-align:center;color:#9ca3af;'>Memuat...</td></tr>";

  fetch("/api/laporan/contact-label?from=" + df + "&to=" + dt + "&kategori=" + encodeURIComponent(kat))
  .then(function(r) {{ return r.json(); }})
  .then(function(json) {{
    _clData = json.labels || [];
    document.getElementById("totalLabelCount").textContent = json.total_label || 0;
    filterTable();
  }})
  .catch(function() {{
    tbody.innerHTML = "<tr><td colspan=3 style='color:#dc2626;padding:20px;'>Gagal memuat.</td></tr>";
  }});
}}

function filterTable() {{
  var q = document.getElementById("searchLabel").value.toLowerCase();
  var filtered = q ? _clData.filter(function(l) {{ return l.name.toLowerCase().indexOf(q) >= 0; }}) : _clData;
  var tbody = document.getElementById("labelListBody");
  if (!filtered.length) {{
    tbody.innerHTML = "<tr><td colspan=3 style='padding:40px;text-align:center;color:#9ca3af;'>Tidak ada data.</td></tr>";
    return;
  }}
  var html = "";
  var totContact = 0;
  for (var i=0; i<filtered.length; i++) {{
    var l = filtered[i];
    totContact += l.total_contact;
    html += "<tr style='border-bottom:1px solid #f3f4f6;'>";
    html += "<td style='padding:12px 16px;'><span onclick='showDetail(" + i + ")' style='color:#5b3df0;cursor:pointer;font-weight:600;text-decoration:underline;'>" + l.name + "</span></td>";
    html += "<td style='padding:12px 16px;font-size:13px;color:#6b7280;'>" + (l.category||"-") + "</td>";
    html += "<td style='padding:12px 16px;text-align:center;font-size:13px;font-weight:600;'>" + l.total_contact + "</td>";
    html += "</tr>";
  }}
  // Total row
  html += "<tr style='background:#f9fafb;font-weight:700;border-top:2px solid #e5e7eb;'>";
  html += "<td style='padding:12px 16px;font-size:13px;' colspan=2>TOTAL</td>";
  html += "<td style='padding:12px 16px;text-align:center;font-size:13px;'>" + totContact + "</td>";
  html += "</tr>";
  tbody.innerHTML = html;
}}

function showDetail(idx) {{
  var l = _clData[idx];
  if (!l) return;
  _currentLabel = l;
  document.getElementById("labelListView").style.display = "none";
  document.getElementById("labelDetailView").style.display = "block";
  var df = document.getElementById("filterFrom").value;
  var dt = document.getElementById("filterTo").value;
  document.getElementById("detailLabelName").innerHTML = "<span style='background:" + l.bg_color + ";color:" + l.text_color + ";padding:3px 10px;border-radius:6px;font-size:13px;font-weight:700;'>" + l.name + "</span>";
  document.getElementById("detailPeriode").textContent = "Periode: " + df + " sampai " + dt;
  var contacts = l.contacts || [];
  var tbody = document.getElementById("labelDetailBody");
  var typeColor = {{"User Initiated":"#2563eb","Business Initiated":"#16a34a","Ads":"#f59e0b"}};
  tbody.innerHTML = contacts.map(function(c) {{
    var tc = typeColor[c.type] || "#6b7280";
    return "<tr style='border-bottom:1px solid #f3f4f6;'>" +
      "<td style='padding:12px 16px;font-size:13px;font-weight:600;'>" + c.nama + "</td>" +
      "<td style='padding:12px 16px;font-size:13px;color:#6b7280;'>+" + c.phone + "</td>" +
      "<td style='padding:12px 16px;'><span style='font-size:12px;font-weight:600;color:" + tc + ";'>" + c.type + "</span></td>" +
      "</tr>";
  }}).join("");
  document.getElementById("detailInfo").textContent = "Menampilkan " + contacts.length + " kontak";
}}

function backToList() {{
  document.getElementById("labelDetailView").style.display = "none";
  document.getElementById("labelListView").style.display = "block";
}}

function exportContactLabel() {{
  if (!_clData.length) {{ alert("Tidak ada data. Klik Cari dulu."); return; }}
  var rows = [["Nama Contact Label","Kategori","Total Contact"]];
  _clData.forEach(function(l) {{ rows.push([l.name, l.category||"-", l.total_contact]); }});
  var csv = rows.map(function(r) {{ return r.map(function(v) {{ return '"'+String(v).replace(/"/g,'""')+'"'; }}).join(","); }}).join(String.fromCharCode(10));
  var blob = new Blob([csv], {{type:"text/csv;charset=utf-8;"}});
  var url = URL.createObjectURL(blob);
  var a = document.createElement("a"); a.href=url; a.download="contact_label.csv"; a.click();
  URL.revokeObjectURL(url);
}}

function exportDetailLabel() {{
  if (!_currentLabel) return;
  var contacts = _currentLabel.contacts || [];
  var rows = [["Nama","No WhatsApp","Type"]];
  contacts.forEach(function(c) {{ rows.push([c.nama, "+"+c.phone, c.type]); }});
  var csv = rows.map(function(r) {{ return r.map(function(v) {{ return '"'+String(v).replace(/"/g,'""')+'"'; }}).join(","); }}).join(String.fromCharCode(10));
  var blob = new Blob([csv], {{type:"text/csv;charset=utf-8;"}});
  var url = URL.createObjectURL(blob);
  var a = document.createElement("a"); a.href=url; a.download="detail_"+_currentLabel.name+".csv"; a.click();
  URL.revokeObjectURL(url);
}}

loadContactLabel();
</script>
"""
    return Response(render_page("laporan-contact-label", "Laporan Contact Label", "", body), mimetype="text/html")

@app.route("/laporan/summary", methods=["GET"])
@login_required
def laporan_summary():
    today = datetime.now().strftime("%Y-%m-%d")
    first_month = datetime.now().strftime("%Y-%m-01")
    body = f"""
  <!-- Filter -->
  <div style="background:#fff;border-radius:12px;padding:20px;margin-bottom:16px;box-shadow:0 1px 3px rgba(0,0,0,.06);">
    <div style="display:flex;gap:12px;align-items:flex-end;flex-wrap:wrap;">
      <div>
        <div style="font-size:12px;font-weight:600;color:#6b7280;margin-bottom:6px;">Group By</div>
        <select id="groupBy" style="padding:8px 12px;border:1px solid #d1d5db;border-radius:8px;font-size:13px;min-width:140px;">
          <option value="day">Day (Hari)</option>
          <option value="month">Month (Bulan)</option>
          <option value="hour">Hour (Jam)</option>
        </select>
      </div>
      <div>
        <div style="font-size:12px;font-weight:600;color:#6b7280;margin-bottom:6px;">Periode</div>
        <div style="display:flex;gap:8px;align-items:center;">
          <input type="date" id="filterFrom" value="{first_month}" style="padding:8px 12px;border:1px solid #d1d5db;border-radius:8px;font-size:13px;">
          <span style="color:#9ca3af;">→</span>
          <input type="date" id="filterTo" value="{today}" style="padding:8px 12px;border:1px solid #d1d5db;border-radius:8px;font-size:13px;">
        </div>
      </div>
      <button onclick="loadSummary()" style="padding:8px 20px;background:#5b3df0;color:#fff;border:none;border-radius:8px;font-size:13px;font-weight:600;cursor:pointer;">Cari</button>
      <button onclick="exportSummary()" style="padding:8px 20px;background:#fff;color:#5b3df0;border:1px solid #5b3df0;border-radius:8px;font-size:13px;font-weight:600;cursor:pointer;">Export to Excel</button>
      <div style="display:flex;border:1px solid #d1d5db;border-radius:8px;overflow:hidden;">
        <button id="btnTabel" onclick="setView('tabel')" style="padding:8px 16px;border:none;background:#5b3df0;color:#fff;font-size:13px;font-weight:600;cursor:pointer;">📋 Tabel</button>
        <button id="btnGrafik" onclick="setView('grafik')" style="padding:8px 16px;border:none;background:#fff;color:#6b7280;font-size:13px;font-weight:600;cursor:pointer;">📈 Grafik</button>
      </div>
    </div>
  </div>

  <!-- Grafik -->
  <div id="grafikView" style="display:none;background:#fff;border-radius:12px;padding:24px;box-shadow:0 1px 3px rgba(0,0,0,.06);margin-bottom:16px;">
    <div style="font-size:15px;font-weight:700;margin-bottom:16px;color:#374151;">Tren Jumlah Contact</div>
    <canvas id="summaryChart" style="max-height:350px;"></canvas>
  </div>

  <!-- Tabel -->
  <div id="tabelView" style="background:#fff;border-radius:12px;overflow:hidden;box-shadow:0 1px 3px rgba(0,0,0,.06);">
    <div style="overflow-x:auto;">
      <table style="width:100%;border-collapse:collapse;min-width:1000px;" id="summaryTable">
        <thead id="summaryHead">
          <tr style="background:#f9fafb;">
            <th style="padding:12px 16px;text-align:left;font-size:12px;color:#6b7280;font-weight:600;border-bottom:1px solid #e5e7eb;white-space:nowrap;">Periode</th>
            <th style="padding:12px 16px;text-align:center;font-size:12px;color:#6b7280;font-weight:600;border-bottom:1px solid #e5e7eb;white-space:nowrap;">Jumlah Contact</th>
            <th style="padding:12px 16px;text-align:center;font-size:12px;color:#6b7280;font-weight:600;border-bottom:1px solid #e5e7eb;white-space:nowrap;">Jumlah Unik Contact</th>
            <th style="padding:12px 16px;text-align:center;font-size:12px;color:#6b7280;font-weight:600;border-bottom:1px solid #e5e7eb;white-space:nowrap;">% Sudah Diberikan Label</th>
            <th style="padding:12px 16px;text-align:center;font-size:12px;color:#6b7280;font-weight:600;border-bottom:1px solid #e5e7eb;white-space:nowrap;background:#fef9c3;">Non Donatur</th>
            <th style="padding:12px 16px;text-align:center;font-size:12px;color:#6b7280;font-weight:600;border-bottom:1px solid #e5e7eb;white-space:nowrap;background:#fef9c3;">No Respon</th>
            <th style="padding:12px 16px;text-align:center;font-size:12px;color:#6b7280;font-weight:600;border-bottom:1px solid #e5e7eb;white-space:nowrap;background:#fef9c3;">Respon</th>
            <th style="padding:12px 16px;text-align:center;font-size:12px;color:#6b7280;font-weight:600;border-bottom:1px solid #e5e7eb;white-space:nowrap;background:#fef9c3;">Donatur Rutin</th>
            <th style="padding:12px 16px;text-align:center;font-size:12px;color:#6b7280;font-weight:600;border-bottom:1px solid #e5e7eb;white-space:nowrap;background:#dcfce7;">Keluhan</th>
            <th style="padding:12px 16px;text-align:center;font-size:12px;color:#6b7280;font-weight:600;border-bottom:1px solid #e5e7eb;white-space:nowrap;background:#dcfce7;">Kerjasama</th>
            <th style="padding:12px 16px;text-align:center;font-size:12px;color:#6b7280;font-weight:600;border-bottom:1px solid #e5e7eb;white-space:nowrap;background:#dcfce7;">Laporan</th>
            <th style="padding:12px 16px;text-align:center;font-size:12px;color:#6b7280;font-weight:600;border-bottom:1px solid #e5e7eb;white-space:nowrap;background:#dcfce7;">Galang Dana</th>
            <th style="padding:12px 16px;text-align:center;font-size:12px;color:#6b7280;font-weight:600;border-bottom:1px solid #e5e7eb;white-space:nowrap;background:#dcfce7;">Donasi</th>
            <th style="padding:12px 16px;text-align:center;font-size:12px;color:#6b7280;font-weight:600;border-bottom:1px solid #e5e7eb;white-space:nowrap;background:#dcfce7;">Lainnya</th>
          </tr>
        </thead>
        <tbody id="summaryBody">
          <tr><td colspan="14" style="padding:40px;text-align:center;color:#9ca3af;">Klik Cari untuk memuat data</td></tr>
        </tbody>
      </table>
    </div>
  </div>
  </div>

<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.min.js"></script>
<script>
var _summaryData = [];
var _summaryChart = null;
var _currentView = "tabel";

function setView(view) {{
  _currentView = view;
  document.getElementById("tabelView").style.display = view === "tabel" ? "block" : "none";
  document.getElementById("grafikView").style.display = view === "grafik" ? "block" : "none";
  document.getElementById("btnTabel").style.background = view === "tabel" ? "#5b3df0" : "#fff";
  document.getElementById("btnTabel").style.color = view === "tabel" ? "#fff" : "#6b7280";
  document.getElementById("btnGrafik").style.background = view === "grafik" ? "#5b3df0" : "#fff";
  document.getElementById("btnGrafik").style.color = view === "grafik" ? "#fff" : "#6b7280";
  if (view === "grafik" && _summaryData.length) renderChart();
}}

function renderChart() {{
  var labels = _summaryData.map(function(r) {{ return r.group_label; }});
  var dataContact = _summaryData.map(function(r) {{ return r.jumlah_contact; }});
  var dataUnik = _summaryData.map(function(r) {{ return r.jumlah_unik; }});

  var ctx = document.getElementById("summaryChart").getContext("2d");
  if (_summaryChart) {{ _summaryChart.destroy(); }}
  _summaryChart = new Chart(ctx, {{
    type: "line",
    data: {{
      labels: labels,
      datasets: [
        {{
          label: "Jumlah Contact",
          data: dataContact,
          borderColor: "#5b3df0",
          backgroundColor: "rgba(91,61,240,0.1)",
          tension: 0.3,
          fill: true,
          pointRadius: 4,
          pointBackgroundColor: "#5b3df0"
        }},
        {{
          label: "Jumlah Unik Contact",
          data: dataUnik,
          borderColor: "#16a34a",
          backgroundColor: "rgba(22,163,74,0.1)",
          tension: 0.3,
          fill: true,
          pointRadius: 4,
          pointBackgroundColor: "#16a34a"
        }}
      ]
    }},
    options: {{
      responsive: true,
      plugins: {{
        legend: {{ position: "top" }},
        tooltip: {{ mode: "index", intersect: false }}
      }},
      scales: {{
        y: {{ beginAtZero: true, ticks: {{ stepSize: 1 }} }}
      }}
    }}
  }});
}}
var _labelCols = ["Non Donatur","No Respon","Respon","Donatur Rutin","Keluhan","Kerjasama","Laporan","Galang dana","Donasi","Lainnya"];

function loadSummary() {{
  var df = document.getElementById("filterFrom").value;
  var dt = document.getElementById("filterTo").value;
  var gb = document.getElementById("groupBy").value;
  var tbody = document.getElementById("summaryBody");
  tbody.innerHTML = "<tr><td colspan=14 style='padding:20px;text-align:center;color:#9ca3af;'>Memuat...</td></tr>";

  fetch("/api/laporan/summary?from=" + df + "&to=" + dt + "&group_by=" + gb)
  .then(function(r) {{ return r.json(); }})
  .then(function(json) {{
    _summaryData = json.rows || [];
    if (!_summaryData.length) {{
      tbody.innerHTML = "<tr><td colspan=14 style='padding:40px;text-align:center;color:#9ca3af;'>Tidak ada data.</td></tr>";
      return;
    }}
    var html = "";
    var totContact=0, totUnik=0, totLabel={{}};
    _labelCols.forEach(function(l) {{ totLabel[l]=0; }});

    for (var i=0; i<_summaryData.length; i++) {{
      var r = _summaryData[i];
      totContact += r.jumlah_contact;
      totUnik += r.jumlah_unik;
      _labelCols.forEach(function(l) {{ totLabel[l] += (r.labels[l]||0); }});

      var pct = r.pct_label + "%";
      var pctColor = r.pct_label >= 70 ? "#16a34a" : r.pct_label >= 40 ? "#d97706" : "#dc2626";
      html += "<tr style='border-bottom:1px solid #f3f4f6;'>";
      html += "<td style='padding:8px 12px;font-size:12px;font-weight:600;white-space:nowrap;'>" + r.group_label + "</td>";
      html += "<td style='padding:12px 16px;font-size:13px;text-align:center;'>" + r.jumlah_contact + "</td>";
      html += "<td style='padding:12px 16px;font-size:13px;text-align:center;'>" + r.jumlah_unik + "</td>";
      html += "<td style='padding:12px 16px;font-size:13px;text-align:center;font-weight:700;color:" + pctColor + ";'>" + pct + "</td>";
      _labelCols.forEach(function(l) {{
        var bg = l === "Non Donatur" || l === "No Respon" || l === "Respon" || l === "Donatur Rutin" ? "#fef9c3" : "#f0fdf4";
        html += "<td style='padding:12px 16px;font-size:13px;text-align:center;background:" + bg + ";'>" + (r.labels[l]||0) + "</td>";
      }});
      html += "</tr>";
    }}

    totUnik = json.total_unik_all || totUnik;
    var totPct = json.total_pct_all || 0;
    html += "<tr style='background:#f9fafb;font-weight:700;border-top:2px solid #e5e7eb;'>";
    html += "<td style='padding:12px 16px;font-size:13px;'>TOTAL</td>";
    html += "<td style='padding:12px 16px;font-size:13px;text-align:center;'>" + totContact + "</td>";
    html += "<td style='padding:12px 16px;font-size:13px;text-align:center;'>" + totUnik + "</td>";
    html += "<td style='padding:12px 16px;font-size:13px;text-align:center;'>" + totPct + "%</td>";
    _labelCols.forEach(function(l) {{
      html += "<td style='padding:12px 16px;font-size:13px;text-align:center;'>" + (totLabel[l]||0) + "</td>";
    }});
    html += "</tr>";

    tbody.innerHTML = html;
    if (_currentView === "grafik") renderChart();
  }})
  .catch(function() {{
    tbody.innerHTML = "<tr><td colspan=14 style='color:#dc2626;padding:20px;'>Gagal memuat.</td></tr>";
  }});
}}

function exportSummary() {{
  if (!_summaryData.length) {{ alert("Tidak ada data. Klik Cari dulu."); return; }}
  var headers = ["Periode","Jumlah Contact","Jumlah Unik Contact","% Sudah Label"].concat(_labelCols);
  var rows = [headers];
  for (var i=0; i<_summaryData.length; i++) {{
    var r = _summaryData[i];
    var row = [r.group_label, r.jumlah_contact, r.jumlah_unik, r.pct_label+"%"];
    _labelCols.forEach(function(l) {{ row.push(r.labels[l]||0); }});
    rows.push(row);
  }}
  var csv = rows.map(function(r) {{ return r.map(function(v) {{ return '"'+String(v).replace(/"/g,'""')+'"'; }}).join(","); }}).join(String.fromCharCode(10));
  var blob = new Blob([csv], {{type:"text/csv;charset=utf-8;"}});
  var url = URL.createObjectURL(blob);
  var a = document.createElement("a");
  a.href=url; a.download="laporan_summary_"+document.getElementById("filterFrom").value+".csv"; a.click();
  URL.revokeObjectURL(url);
}}

loadSummary();
</script>
"""
    return Response(render_page("laporan-summary", "Laporan Summary", "Ringkasan percakapan berdasarkan periode", body), mimetype="text/html")

@app.route("/laporan/chat-harian", methods=["GET"])
@login_required
def laporan_chat_harian():
    today = datetime.now().strftime("%Y-%m-%d")
    body = f"""
  <!-- Filter -->
  <div style="background:#fff;border-radius:12px;padding:20px;margin-bottom:16px;box-shadow:0 1px 3px rgba(0,0,0,.06);">
    <div style="display:flex;gap:12px;align-items:flex-end;flex-wrap:wrap;">
      <div>
        <div style="font-size:12px;font-weight:600;color:#6b7280;margin-bottom:6px;">Filter Tanggal</div>
        <div style="display:flex;gap:8px;align-items:center;">
          <input type="date" id="filterFrom" value="{today}" style="padding:8px 12px;border:1px solid #d1d5db;border-radius:8px;font-size:13px;">
          <span style="color:#9ca3af;">→</span>
          <input type="date" id="filterTo" value="{today}" style="padding:8px 12px;border:1px solid #d1d5db;border-radius:8px;font-size:13px;">
        </div>
      </div>
      <div>
        <div style="font-size:12px;font-weight:600;color:#6b7280;margin-bottom:6px;">Tipe</div>
        <select id="filterTipe" style="padding:8px 12px;border:1px solid #d1d5db;border-radius:8px;font-size:13px;min-width:180px;">
          <option value="">(Semua)</option>
          <option value="User Initiated">User Initiated</option>
          <option value="Business Initiated">Business Initiated</option>
          <option value="Ads">Ads</option>
        </select>
      </div>
      <button onclick="loadLaporan()" style="padding:8px 20px;background:#5b3df0;color:#fff;border:none;border-radius:8px;font-size:13px;font-weight:600;cursor:pointer;">Cari</button>
      <button onclick="exportExcel()" style="padding:8px 20px;background:#fff;color:#5b3df0;border:1px solid #5b3df0;border-radius:8px;font-size:13px;font-weight:600;cursor:pointer;">Export to Excel</button>
    </div>
  </div>

  <!-- Summary -->
  <div style="display:flex;gap:16px;margin-bottom:16px;flex-wrap:wrap;">
    <div style="background:#fff;border-radius:10px;padding:16px 24px;box-shadow:0 1px 3px rgba(0,0,0,.06);flex:1;min-width:160px;">
      <div style="font-size:12px;color:#6b7280;margin-bottom:4px;">Total User Initiated</div>
      <div style="font-size:28px;font-weight:800;color:#5b3df0;" id="totalUserInit">-</div>
    </div>
    <div style="background:#fff;border-radius:10px;padding:16px 24px;box-shadow:0 1px 3px rgba(0,0,0,.06);flex:1;min-width:160px;">
      <div style="font-size:12px;color:#6b7280;margin-bottom:4px;">Total Business Initiated</div>
      <div style="font-size:28px;font-weight:800;color:#16a34a;" id="totalBizInit">-</div>
    </div>
    <div style="background:#fff;border-radius:10px;padding:16px 24px;box-shadow:0 1px 3px rgba(0,0,0,.06);flex:1;min-width:160px;">
      <div style="font-size:12px;color:#6b7280;margin-bottom:4px;">Total Ads</div>
      <div style="font-size:28px;font-weight:800;color:#f59e0b;" id="totalAds">-</div>
    </div>
    <div style="background:#fff;border-radius:10px;padding:16px 24px;box-shadow:0 1px 3px rgba(0,0,0,.06);flex:1;min-width:160px;">
      <div style="font-size:12px;color:#6b7280;margin-bottom:4px;">Total Contact</div>
      <div style="font-size:28px;font-weight:800;color:#374151;" id="totalContact">-</div>
    </div>
    <div style="background:#fff;border-radius:10px;padding:16px 24px;box-shadow:0 1px 3px rgba(0,0,0,.06);flex:1;min-width:160px;">
      <div style="font-size:12px;color:#6b7280;margin-bottom:4px;">Total Unik Contact</div>
      <div style="font-size:28px;font-weight:800;color:#0891b2;" id="totalUnik">-</div>
    </div>
  </div>

  <!-- Grafik -->
  <div id="grafikView" style="display:none;background:#fff;border-radius:12px;padding:24px;box-shadow:0 1px 3px rgba(0,0,0,.06);margin-bottom:16px;">
    <div style="font-size:15px;font-weight:700;margin-bottom:16px;color:#374151;">Tren Jumlah Contact</div>
    <canvas id="summaryChart" style="max-height:350px;"></canvas>
  </div>

  <!-- Tabel -->
  <div id="tabelView" style="background:#fff;border-radius:12px;overflow:hidden;box-shadow:0 1px 3px rgba(0,0,0,.06);">
    <div style="overflow-x:auto;">
      <table style="width:100%;border-collapse:collapse;min-width:700px;">
        <thead>
          <tr style="background:#f9fafb;">
            <th style="padding:12px 16px;text-align:left;font-size:12px;color:#6b7280;font-weight:600;border-bottom:1px solid #e5e7eb;white-space:nowrap;">Chat Date</th>
            <th style="padding:12px 16px;text-align:left;font-size:12px;color:#6b7280;font-weight:600;border-bottom:1px solid #e5e7eb;white-space:nowrap;">Nama User</th>
            <th style="padding:12px 16px;text-align:left;font-size:12px;color:#6b7280;font-weight:600;border-bottom:1px solid #e5e7eb;white-space:nowrap;">Nama</th>
            <th style="padding:12px 16px;text-align:left;font-size:12px;color:#6b7280;font-weight:600;border-bottom:1px solid #e5e7eb;white-space:nowrap;">Cell Phone</th>
            <th style="padding:12px 16px;text-align:left;font-size:12px;color:#6b7280;font-weight:600;border-bottom:1px solid #e5e7eb;white-space:nowrap;">Type</th>
            <th style="padding:12px 16px;text-align:left;font-size:12px;color:#6b7280;font-weight:600;border-bottom:1px solid #e5e7eb;white-space:nowrap;">Kontak Label</th>
          </tr>
        </thead>
        <tbody id="laporanBody">
          <tr><td colspan="6" style="padding:40px;text-align:center;color:#9ca3af;">Klik Cari untuk memuat data</td></tr>
        </tbody>
      </table>
    </div>
    <div style="padding:12px 16px;border-top:1px solid #f3f4f6;font-size:13px;color:#6b7280;" id="laporanInfo"></div>
  </div>

<script>
var _laporanData = [];
function loadLaporan() {{
  var df = document.getElementById("filterFrom").value;
  var dt = document.getElementById("filterTo").value;
  var tbody = document.getElementById("laporanBody");
  tbody.innerHTML = "<tr><td colspan=6 style='padding:20px;text-align:center;color:#9ca3af;'>Memuat...</td></tr>";
  var tipe = document.getElementById("filterTipe") ? document.getElementById("filterTipe").value : "";
  fetch("/api/laporan/chat-harian?from=" + df + "&to=" + dt + "&tipe=" + encodeURIComponent(tipe))
  .then(function(r) {{ return r.json(); }})
  .then(function(json) {{
    _laporanData = json.rows || [];
    document.getElementById("totalUserInit").textContent = json.user_initiated || 0;
    document.getElementById("totalBizInit").textContent = json.business_initiated || 0;
    document.getElementById("totalAds").textContent = json.ads || 0;
    document.getElementById("totalContact").textContent = json.total_contact || 0;
    document.getElementById("totalUnik").textContent = json.total_unik || 0;
    if (!_laporanData.length) {{
      tbody.innerHTML = "<tr><td colspan=6 style='padding:40px;text-align:center;color:#9ca3af;'>Tidak ada data.</td></tr>";
      document.getElementById("laporanInfo").textContent = "";
      return;
    }}
    var seen = {{}}, unique = [];
    for (var i=0;i<_laporanData.length;i++) {{
      var r=_laporanData[i], key=r.chat_date+"_"+r.phone;
      if (!seen[key]) {{ seen[key]=true; unique.push(r); }}
    }}
    var html = "";
    for (var j=0;j<unique.length;j++) {{
      var r=unique[j];
      var tc=r.type==="User Initiated"?"#2563eb":r.type==="Ads"?"#f59e0b":"#16a34a";
      var la=(r.labels||[]).map(function(l){{return "<span style='padding:2px 8px;border-radius:10px;font-size:11px;background:#e0d9ff;color:#5b3df0;'>"+l+"</span>";}}).join("")||"-";
      html+="<tr style='border-bottom:1px solid #f3f4f6;'>";
      html+="<td style='padding:12px 16px;font-size:13px;white-space:nowrap;'>"+r.chat_date+"</td>";
      html+="<td style='padding:12px 16px;font-size:13px;'>"+r.nama_user+"</td>";
      html+="<td style='padding:12px 16px;font-size:13px;font-weight:600;'>"+r.nama+"</td>";
      html+="<td style='padding:12px 16px;font-size:13px;color:#6b7280;'>+"+r.phone+"</td>";
      html+="<td style='padding:12px 16px;'><span style='font-size:12px;font-weight:600;color:"+tc+";'>"+r.type+"</span></td>";
      html+="<td style='padding:12px 16px;'>"+la+"</td></tr>";
    }}
    tbody.innerHTML = html;
    document.getElementById("laporanInfo").textContent = "Menampilkan "+unique.length+" kontak dari "+_laporanData.length+" total pesan";
  }})
  .catch(function() {{
    tbody.innerHTML = "<tr><td colspan=6 style='color:#dc2626;padding:20px;'>Gagal memuat.</td></tr>";
  }});
}}
function exportExcel() {{
  if (!_laporanData.length) {{ alert("Tidak ada data. Klik Cari dulu."); return; }}
  var seen={{}}, unique=[];
  for (var i=0;i<_laporanData.length;i++) {{
    var r=_laporanData[i], key=r.chat_date+"_"+r.phone;
    if (!seen[key]) {{ seen[key]=true; unique.push(r); }}
  }}
  var hdr=["Chat Date","Nama User","Nama","Cell Phone","Type","Kontak Label"];
  var rows=[hdr];
  for (var j=0;j<unique.length;j++) {{
    var r=unique[j];
    rows.push([r.chat_date,r.nama_user,r.nama,"+"+r.phone,r.type,(r.labels||[]).join("; ")]);
  }}
  var csv=rows.map(function(row){{return row.map(function(v){{return '"'+String(v).replace(/"/g,'""')+'"';}}).join(",");}}).join(String.fromCharCode(10));
  var blob=new Blob([csv],{{type:"text/csv;charset=utf-8;"}});
  var url=URL.createObjectURL(blob);
  var a=document.createElement("a");
  a.href=url;a.download="laporan_chat_"+document.getElementById("filterFrom").value+".csv";a.click();
  URL.revokeObjectURL(url);
}}
loadLaporan();
</script>
"""
    return Response(render_page("laporan", "Laporan Chat Harian", "", body), mimetype="text/html")

@app.route("/pengaturan/pengguna", methods=["GET"])
@admin_required
def pengguna_page():
    user = {"nama": session.get("user_nama",""), "email": session.get("user_email",""), "role": session.get("user_role","")}
    body = f"""
  <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:24px;">
    <div></div>
    <button onclick="showUserForm()" class="btn">+ Tambah Pengguna</button>
  </div>

  <div style="background:#fff;border-radius:12px;overflow:hidden;box-shadow:0 1px 3px rgba(0,0,0,.06);">
    <table style="width:100%;border-collapse:collapse;">
      <thead>
        <tr style="background:#f9fafb;">
          <th style="padding:12px 16px;text-align:left;font-size:12px;color:#6b7280;font-weight:600;border-bottom:1px solid #e5e7eb;">Username (Email)</th>
          <th style="padding:12px 16px;text-align:left;font-size:12px;color:#6b7280;font-weight:600;border-bottom:1px solid #e5e7eb;">Nama Lengkap</th>
          <th style="padding:12px 16px;text-align:left;font-size:12px;color:#6b7280;font-weight:600;border-bottom:1px solid #e5e7eb;">Role</th>
          <th style="padding:12px 16px;text-align:left;font-size:12px;color:#6b7280;font-weight:600;border-bottom:1px solid #e5e7eb;">Waktu Login Terakhir</th>
          <th style="padding:12px 16px;text-align:left;font-size:12px;color:#6b7280;font-weight:600;border-bottom:1px solid #e5e7eb;">Status</th>
          <th style="padding:12px 16px;text-align:right;font-size:12px;color:#6b7280;font-weight:600;border-bottom:1px solid #e5e7eb;">Aksi</th>
        </tr>
      </thead>
      <tbody id="usersTableBody">
        <tr><td colspan="6" style="padding:40px;text-align:center;color:#9ca3af;">Memuat...</td></tr>
      </tbody>
    </table>
  </div>

  <!-- Modal Form -->
  <div id="userFormModal" style="display:none;position:fixed;inset:0;background:rgba(0,0,0,.5);z-index:1000;align-items:center;justify-content:center;">
    <div style="background:#fff;border-radius:16px;width:480px;max-width:95vw;">
      <div style="background:#5b3df0;padding:16px 24px;display:flex;justify-content:space-between;align-items:center;border-radius:16px 16px 0 0;">
        <span style="color:#fff;font-weight:700;font-size:16px;" id="userFormTitle">Tambah Pengguna</span>
        <span onclick="closeUserForm()" style="color:#fff;cursor:pointer;font-size:22px;">✕</span>
      </div>
      <div style="padding:24px;">
        <input type="hidden" id="uFormId">
        <div class="form-group">
          <label>Email <span style="color:#dc2626;">*</span></label>
          <input type="email" id="uFormEmail" placeholder="email@raihmimpi.id" style="width:100%;padding:10px 12px;border:1px solid #d1d5db;border-radius:8px;font-size:14px;box-sizing:border-box;">
        </div>
        <div class="form-group">
          <label>Nama Lengkap <span style="color:#dc2626;">*</span></label>
          <input type="text" id="uFormNama" placeholder="Nama pengguna..." style="width:100%;padding:10px 12px;border:1px solid #d1d5db;border-radius:8px;font-size:14px;box-sizing:border-box;">
        </div>
        <div class="form-group">
          <label>Role <span style="color:#dc2626;">*</span></label>
          <select id="uFormRole" style="width:100%;padding:10px 12px;border:1px solid #d1d5db;border-radius:8px;font-size:14px;">
            <option value="ADMINISTRATOR">ADMINISTRATOR</option>
            <option value="TELEMARKETING">TELEMARKETING</option>
            <option value="FINANCE">FINANCE</option>
          </select>
        </div>
        <div class="form-group">
          <label id="uFormPwdLabel">Password <span style="color:#dc2626;">*</span></label>
          <input type="password" id="uFormPassword" placeholder="Password..." style="width:100%;padding:10px 12px;border:1px solid #d1d5db;border-radius:8px;font-size:14px;box-sizing:border-box;">
          <div style="font-size:11px;color:#9ca3af;margin-top:4px;" id="uFormPwdHint"></div>
        </div>
        <div class="form-group" id="uFormStatusGroup" style="display:none;">
          <label>Status</label>
          <select id="uFormStatus" style="width:100%;padding:10px 12px;border:1px solid #d1d5db;border-radius:8px;font-size:14px;">
            <option value="aktif">Aktif</option>
            <option value="nonaktif">Nonaktif</option>
          </select>
        </div>
      </div>
      <div style="padding:16px 24px;border-top:1px solid #f3f4f6;display:flex;gap:10px;justify-content:flex-end;">
        <button onclick="closeUserForm()" style="padding:10px 20px;border:1px solid #e5e7eb;border-radius:8px;background:#fff;color:#374151;cursor:pointer;font-weight:600;">Batal</button>
        <button onclick="saveUserForm()" style="padding:10px 24px;background:#5b3df0;color:#fff;border:none;border-radius:8px;cursor:pointer;font-weight:600;">Simpan</button>
      </div>
    </div>
  </div>

<script>
loadUsersTable();

async function loadUsersTable() {{
  const tbody = document.getElementById("usersTableBody");
  try {{
    const res = await fetch("/api/users");
    const json = await res.json();
    const users = json.users || [];
    const roleColor = {{ADMINISTRATOR:"#5b3df0", TELEMARKETING:"#0891b2", FINANCE:"#16a34a"}};
    tbody.innerHTML = users.map(u => {{
      const rc = roleColor[u.role] || "#6b7280";
      const lastLogin = u.last_login ? u.last_login.replace("T"," ").substring(0,16) : "-";
      const statusBg = u.status === "aktif" ? "#dcfce7" : "#f3f4f6";
      const statusColor = u.status === "aktif" ? "#16a34a" : "#6b7280";
      return `<tr style="border-bottom:1px solid #f3f4f6;">
        <td style="padding:12px 16px;font-size:13px;color:#5b3df0;">${{u.email}}</td>
        <td style="padding:12px 16px;font-size:13px;font-weight:600;">${{u.nama}}</td>
        <td style="padding:12px 16px;">
          <span style="padding:3px 10px;border-radius:12px;font-size:11px;font-weight:700;background:${{rc}}20;color:${{rc}};">${{u.role}}</span>
        </td>
        <td style="padding:12px 16px;font-size:12px;color:#6b7280;">${{lastLogin}}</td>
        <td style="padding:12px 16px;">
          <span style="padding:3px 10px;border-radius:12px;font-size:11px;font-weight:700;background:${{statusBg}};color:${{statusColor}};">${{u.status}}</span>
        </td>
        <td style="padding:12px 16px;text-align:right;display:flex;gap:6px;justify-content:flex-end;">
          <button onclick="editUser('${{u.id}}','${{u.email}}','${{u.nama}}','${{u.role}}','${{u.status}}')"
            style="padding:5px 12px;border:1px solid #5b3df0;color:#5b3df0;border-radius:6px;font-size:12px;cursor:pointer;background:#fff;">Edit</button>
          <button onclick="deleteUser('${{u.id}}','${{u.nama}}')"
            style="padding:5px 12px;border:1px solid #dc2626;color:#dc2626;border-radius:6px;font-size:12px;cursor:pointer;background:#fff;">Hapus</button>
        </td>
      </tr>`;
    }}).join("");
  }} catch(e) {{
    tbody.innerHTML = '<tr><td colspan="6" style="padding:20px;color:#dc2626;">Gagal memuat.</td></tr>';
  }}
}}

function showUserForm() {{
  document.getElementById("userFormTitle").textContent = "Tambah Pengguna";
  document.getElementById("uFormId").value = "";
  document.getElementById("uFormEmail").value = "";
  document.getElementById("uFormEmail").disabled = false;
  document.getElementById("uFormNama").value = "";
  document.getElementById("uFormRole").value = "TELEMARKETING";
  document.getElementById("uFormPassword").value = "";
  document.getElementById("uFormPwdHint").textContent = "";
  document.getElementById("uFormStatusGroup").style.display = "none";
  document.getElementById("userFormModal").style.display = "flex";
}}

function editUser(id, email, nama, role, status) {{
  document.getElementById("userFormTitle").textContent = "Edit Pengguna";
  document.getElementById("uFormId").value = id;
  document.getElementById("uFormEmail").value = email;
  document.getElementById("uFormEmail").disabled = true;
  document.getElementById("uFormNama").value = nama;
  document.getElementById("uFormRole").value = role;
  document.getElementById("uFormPassword").value = "";
  document.getElementById("uFormPwdHint").textContent = "Kosongkan jika tidak ingin ganti password";
  document.getElementById("uFormStatus").value = status;
  document.getElementById("uFormStatusGroup").style.display = "block";
  document.getElementById("userFormModal").style.display = "flex";
}}

function closeUserForm() {{
  document.getElementById("userFormModal").style.display = "none";
}}

async function saveUserForm() {{
  const id = document.getElementById("uFormId").value;
  const email = document.getElementById("uFormEmail").value.trim();
  const nama = document.getElementById("uFormNama").value.trim();
  const role = document.getElementById("uFormRole").value;
  const password = document.getElementById("uFormPassword").value;
  const status = document.getElementById("uFormStatus").value || "aktif";

  if (!nama) {{ alert("Nama wajib diisi"); return; }}

  if (id) {{
    // Edit
    const body = {{nama, role, status}};
    if (password) body.password = password;
    const res = await fetch("/api/users/" + id, {{method:"PUT", headers:{{"Content-Type":"application/json"}}, body:JSON.stringify(body)}});
    const json = await res.json();
    if (json.error) {{ alert("Error: " + json.error); return; }}
  }} else {{
    // Create
    if (!email || !password) {{ alert("Email dan password wajib diisi"); return; }}
    const res = await fetch("/api/users", {{method:"POST", headers:{{"Content-Type":"application/json"}}, body:JSON.stringify({{email, nama, role, password}})}});
    const json = await res.json();
    if (json.error) {{ alert("Error: " + json.error); return; }}
  }}
  closeUserForm();
  loadUsersTable();
}}

async function deleteUser(id, nama) {{
  if (!confirm("Hapus pengguna '" + nama + "'?")) return;
  const res = await fetch("/api/users/" + id, {{method:"DELETE"}});
  const json = await res.json();
  if (json.error) {{ alert("Error: " + json.error); return; }}
  loadUsersTable();
}}
</script>
"""
    return Response(render_page("pengaturan", "Daftar Pengguna", "Kelola akses pengguna sistem Raihmimpi", body), mimetype="text/html")

@app.route("/logout")
def logout():
    session.clear()
    return redirect("/login")

# ---- User Management API ----
@app.route("/api/users", methods=["GET"])
def api_users_list():
    if not session.get("user_id"):
        return jsonify({"error": "Unauthorized"}), 401
    users = load_users()
    # Jangan return password
    safe = [{k:v for k,v in u.items() if k != "password"} for u in users]
    return jsonify({"users": safe})

@app.route("/api/users", methods=["POST"])
def api_users_create():
    if session.get("user_role") != "ADMINISTRATOR":
        return jsonify({"error": "Akses ditolak"}), 403
    body = request.get_json(silent=True) or {}
    email = body.get("email", "").strip().lower()
    nama = body.get("nama", "").strip()
    role = body.get("role", "TELEMARKETING").upper()
    password = body.get("password", "")
    if not email or not nama or not password:
        return jsonify({"error": "email, nama, password wajib diisi"}), 400
    users = load_users()
    if any(u["email"].lower() == email for u in users):
        return jsonify({"error": "Email sudah terdaftar"}), 400
    new_user = {
        "id": f"user_{int(datetime.now().timestamp()*1000)}",
        "email": email, "password": hash_password(password),
        "nama": nama, "role": role, "status": "aktif",
        "created_at": datetime.now().isoformat(), "last_login": None
    }
    users.append(new_user)
    save_users(users)
    return jsonify({"status": "created", "user": {k:v for k,v in new_user.items() if k != "password"}})

@app.route("/api/users/<user_id>", methods=["PUT"])
def api_users_update(user_id):
    if session.get("user_role") != "ADMINISTRATOR":
        return jsonify({"error": "Akses ditolak"}), 403
    body = request.get_json(silent=True) or {}
    users = load_users()
    for u in users:
        if u["id"] == user_id:
            if "nama" in body: u["nama"] = body["nama"]
            if "role" in body: u["role"] = body["role"].upper()
            if "status" in body: u["status"] = body["status"]
            if "password" in body and body["password"]:
                u["password"] = hash_password(body["password"])
            save_users(users)
            return jsonify({"status": "updated"})
    return jsonify({"error": "User tidak ditemukan"}), 404

@app.route("/api/users/<user_id>", methods=["DELETE"])
def api_users_delete(user_id):
    if session.get("user_role") != "ADMINISTRATOR":
        return jsonify({"error": "Akses ditolak"}), 403
    if user_id == session.get("user_id"):
        return jsonify({"error": "Tidak bisa hapus akun sendiri"}), 400
    users = load_users()
    users = [u for u in users if u["id"] != user_id]
    save_users(users)
    return jsonify({"status": "deleted"})

@app.route("/api/laporan/chat-harian", methods=["GET"])
@login_required
def api_laporan_chat_harian():
    """API laporan chat harian - 1 baris per kontak per hari."""
    date_from = request.args.get("from", "")
    date_to = request.args.get("to", "")
    filter_tipe = request.args.get("tipe", "")
    
    inbox = load_inbox()
    contacts = inbox.get("contacts", {})
    messages_db = inbox.get("messages", {})
    
    rows = []
    for phone, contact in contacts.items():
        msgs = messages_db.get(phone, [])
        if not msgs:
            msgs = contact.get("messages", [])
        if not msgs and contact.get("last_message_at"):
            msgs = [{
                "direction": "in",
                "text": contact.get("last_message", ""),
                "timestamp": contact.get("last_message_at", ""),
                "type": "text"
            }]
        
        # Group pesan per hari
        days = {}
        for msg in msgs:
            ts = msg.get("timestamp", "")
            if not ts:
                continue
            date_part = ts[:10]
            if date_from and date_part < date_from:
                continue
            if date_to and date_part > date_to:
                continue
            if date_part not in days:
                days[date_part] = []
            days[date_part].append(msg)
        
        # 1 baris per kontak per hari
        is_ads = bool(contact.get("ctwa_clid") or contact.get("ad_headline"))
        
        for date_part, day_msgs in days.items():
            # Tentukan tipe: kalau ada pesan masuk dari user = User Initiated
            has_incoming = any(m.get("direction", "in") == "in" for m in day_msgs)
            first_msg = min(day_msgs, key=lambda m: m.get("timestamp", ""))
            first_direction = first_msg.get("direction", "in")
            
            if is_ads:
                msg_type = "Ads"
            elif has_incoming:
                msg_type = "User Initiated"
            else:
                msg_type = "Business Initiated"
            
            rows.append({
                "chat_date": date_part,
                "chat_datetime": date_part,
                "nama_user": session.get("user_nama", "-"),
                "nama": contact.get("name", phone),
                "phone": phone,
                "type": msg_type,
                "labels": contact.get("labels", []),
                "direction": first_direction
            })
    
    # Sort by date desc
    rows.sort(key=lambda x: x["chat_date"], reverse=True)
    
    # Hitung summary per kontak unik per hari
    user_initiated = sum(1 for r in rows if r["type"] == "User Initiated")
    business_initiated = sum(1 for r in rows if r["type"] == "Business Initiated")
    ads = sum(1 for r in rows if r["type"] == "Ads")
    
    # Apply filter tipe jika ada
    if filter_tipe:
        rows = [r for r in rows if r["type"] == filter_tipe]

    total_contact = len(rows)
    total_unik = len(set(r["phone"] for r in rows))

    return jsonify({
        "rows": rows,
        "total": total_contact,
        "user_initiated": user_initiated,
        "business_initiated": business_initiated,
        "ads": ads,
        "total_contact": total_contact,
        "total_unik": total_unik
    })

@app.route("/api/laporan/summary", methods=["GET"])
@login_required
def api_laporan_summary():
    """API Laporan Summary - group by Day atau Month."""
    date_from = request.args.get("from", "")
    date_to = request.args.get("to", "")
    group_by = request.args.get("group_by", "day")  # day atau month

    inbox = load_inbox()
    contacts = inbox.get("contacts", {})
    messages_db = inbox.get("messages", {})

    # Label columns yang dimonitor
    LABEL_COLS = ["Non Donatur", "No Respon", "Respon", "Donatur Rutin",
                  "Keluhan", "Kerjasama", "Laporan", "Galang dana", "Donasi", "Lainnya"]

    # Kumpulkan data per kontak per hari
    contact_days = {}  # key: (group_key, phone)

    for phone, contact in contacts.items():
        msgs = messages_db.get(phone, [])
        if not msgs:
            msgs = contact.get("messages", [])
        if not msgs and contact.get("last_message_at"):
            msgs = [{"direction": "in", "timestamp": contact.get("last_message_at", ""), "type": "text"}]

        days_seen = set()
        for msg in msgs:
            ts = msg.get("timestamp", "")
            if not ts:
                continue
            # Konversi UTC ke WIB (UTC+7)
            try:
                from datetime import datetime as dt2, timedelta
                if "T" in ts:
                    dt_utc = dt2.fromisoformat(ts[:19])
                else:
                    dt_utc = dt2.strptime(ts[:19], "%Y-%m-%d %H:%M:%S")
                dt_wib = dt_utc + timedelta(hours=7)
                date_part = dt_wib.strftime("%Y-%m-%d")
            except Exception:
                date_part = ts[:10]

            if date_from and date_part < date_from:
                continue
            if date_to and date_part > date_to:
                continue

            # Tentukan group key
            try:
                from datetime import datetime as dt2
                d = dt2.strptime(date_part, "%Y-%m-%d")
                if group_by == "month":
                    group_key = d.strftime("%Y-%m")
                    group_label = d.strftime("%B")
                elif group_by == "hour":
                    hour = dt_wib.hour
                    group_key = f"{hour:02d}"
                    group_label = f"{hour:02d}:00 - {hour:02d}:59"
                else:
                    days_id = ["Senin", "Selasa", "Rabu", "Kamis", "Jumat", "Sabtu", "Minggu"]
                    group_key = str(d.weekday())
                    group_label = days_id[d.weekday()]
            except Exception:
                continue

            day_key = (group_key, group_label, phone, date_part)
            if day_key not in days_seen:
                days_seen.add(day_key)
                if (group_key, group_label) not in contact_days:
                    contact_days[(group_key, group_label)] = {}
                if phone not in contact_days[(group_key, group_label)]:
                    contact_days[(group_key, group_label)][phone] = {
                        "labels": contact.get("labels", []),
                        "day_count": 0
                    }
                contact_days[(group_key, group_label)][phone]["day_count"] += 1

    # Untuk Day: phone_first_group = group_key pertama kali nomor muncul
    phone_first_group = {}
    for (gk, gl), phones_data in sorted(contact_days.items(), key=lambda x: x[0][0]):
        for phone in phones_data:
            if phone not in phone_first_group:
                phone_first_group[phone] = gk

    # Build rows
    rows = []
    for (group_key, group_label), phones_data in contact_days.items():
        # Jumlah Contact = total baris (tiap nomor × hari aktif dalam group)
        jumlah_contact = sum(d.get("day_count", 1) for d in phones_data.values())

        if group_by == "month":
            # Month: unik = semua nomor yang aktif di bulan ini
            unik_phones = list(phones_data.keys())
        else:
            # Day/Hour: unik = nomor yang PERTAMA KALI muncul di group ini
            unik_phones = [p for p in phones_data if phone_first_group.get(p) == group_key]

        jumlah_unik = len(unik_phones)
        sudah_label = sum(1 for p in unik_phones if phones_data[p]["labels"])
        pct_label = round(sudah_label / jumlah_unik * 100, 1) if jumlah_unik > 0 else 0

        label_counts = {}
        for col in LABEL_COLS:
            label_counts[col] = sum(1 for p in unik_phones if col in phones_data[p]["labels"])

        rows.append({
            "group_key": group_key,
            "group_label": group_label,
            "jumlah_contact": jumlah_contact,
            "jumlah_unik": jumlah_unik,
            "sudah_label": sudah_label,
            "pct_label": pct_label,
            "labels": label_counts
        })

    # Sort
    # Untuk Hour: tambahkan semua jam 00-23 yang kosong
    if group_by == "hour":
        existing_keys = {r["group_key"] for r in rows}
        for h in range(24):
            hk = "%02d" % h
            if hk not in existing_keys:
                rows.append({
                    "group_key": hk,
                    "group_label": "%02d:00 - %02d:59" % (h, h),
                    "jumlah_contact": 0,
                    "jumlah_unik": 0,
                    "sudah_label": 0,
                    "pct_label": 0,
                    "labels": dict((col, 0) for col in LABEL_COLS)
                })
    rows.sort(key=lambda x: x["group_key"])

    # total_unik_all = semua nomor unik dalam seluruh periode (untuk hitung total %)
    all_phones_with_label = set()
    all_phones = set()
    for (gk, gl), phones_data in contact_days.items():
        for phone, d in phones_data.items():
            all_phones.add(phone)
            if d["labels"]:
                all_phones_with_label.add(phone)
    
    total_unik_all = len(all_phones)
    total_sudah_label_all = len(all_phones_with_label)
    total_pct_all = round(total_sudah_label_all / total_unik_all * 100, 1) if total_unik_all > 0 else 0

    return jsonify({
        "rows": rows,
        "label_cols": LABEL_COLS,
        "total_unik_all": total_unik_all,
        "total_sudah_label_all": total_sudah_label_all,
        "total_pct_all": total_pct_all
    })

@app.route("/api/laporan/contact-label", methods=["GET"])
@login_required
def api_laporan_contact_label():
    """API laporan contact label - list label dengan total kontak unik."""
    date_from = request.args.get("from", "")
    date_to = request.args.get("to", "")
    search_label = request.args.get("label", "").strip().lower()
    kategori = request.args.get("kategori", "").strip()

    inbox = load_inbox()
    contacts = inbox.get("contacts", {})
    messages_db = inbox.get("messages", {})
    labels_cfg = load_labels()

    # Tentukan type per kontak (User Initiated / Business Initiated / Ads)
    def get_contact_type(phone, contact):
        if contact.get("ctwa_clid") or contact.get("ad_headline"):
            return "Ads"
        msgs = messages_db.get(phone, [])
        if not msgs:
            msgs = contact.get("messages", [])
        # Filter berdasarkan tanggal
        for msg in sorted(msgs, key=lambda m: m.get("timestamp", "")):
            ts = msg.get("timestamp", "")
            if not ts:
                continue
            date_part = ts[:10]
            if date_from and date_part < date_from:
                continue
            if date_to and date_part > date_to:
                continue
            return "User Initiated" if msg.get("direction", "in") == "in" else "Business Initiated"
        return "User Initiated"

    # Kumpulkan kontak per label dalam rentang tanggal
    label_data = {}  # label_name -> {phones: set, contacts: []}

    for phone, contact in contacts.items():
        # Cek apakah kontak aktif dalam rentang tanggal
        last_msg_at = contact.get("last_message_at", "")
        if not last_msg_at:
            continue
        date_part = last_msg_at[:10]

        # Cek pesan dalam rentang
        msgs = messages_db.get(phone, [])
        if not msgs:
            msgs = contact.get("messages", [])
        
        in_range = False
        if not date_from and not date_to:
            in_range = True
        else:
            for msg in msgs:
                ts = msg.get("timestamp", "")
                if not ts:
                    continue
                dp = ts[:10]
                if date_from and dp < date_from:
                    continue
                if date_to and dp > date_to:
                    continue
                in_range = True
                break

        if not in_range:
            continue

        contact_labels = contact.get("labels", [])
        for lname in contact_labels:
            if lname not in label_data:
                label_data[lname] = {"phones": set(), "contacts": []}
            if phone not in label_data[lname]["phones"]:
                label_data[lname]["phones"].add(phone)
                label_data[lname]["contacts"].append({
                    "nama": contact.get("name", phone),
                    "phone": phone,
                    "type": get_contact_type(phone, contact)
                })

    # Build result
    result = []
    for lcfg in labels_cfg:
        lname = lcfg["name"]
        lkat = lcfg.get("category", "-")

        if search_label and search_label not in lname.lower():
            continue
        if kategori and lkat != kategori:
            continue

        data = label_data.get(lname, {"phones": set(), "contacts": []})
        result.append({
            "name": lname,
            "category": lkat,
            "bg_color": lcfg.get("bg_color", "#e5e7eb"),
            "text_color": lcfg.get("text_color", "#374151"),
            "total_contact": len(data["phones"]),
            "contacts": data["contacts"]
        })

    total_label = len(result)
    return jsonify({"labels": result, "total_label": total_label})

@app.route("/api/laporan/tracking-campaign", methods=["GET"])
@login_required
def api_laporan_tracking_campaign():
    """API Laporan Tracking Campaign - dari label [bracket] + CTWA data."""
    date_from = request.args.get("from", "")
    date_to = request.args.get("to", "")
    search_campaign = request.args.get("campaign", "").strip().lower()

    inbox = load_inbox()
    contacts = inbox.get("contacts", {})
    messages_db = inbox.get("messages", {})
    transaksi_all = load_data()

    # Kumpulkan data per campaign (bracket label)
    campaigns = {}  # campaign_name -> data

    import re as _re

    for phone, contact in contacts.items():
        msgs = messages_db.get(phone, [])
        if not msgs:
            msgs = contact.get("messages", [])

        # Cari pesan dengan bracket label dan dalam rentang tanggal
        bracket_labels_found = []
        first_msg_date = None
        first_msg_text = None
        ad_id = contact.get("ad_source_id", "") or ""
        ctwa_clid = contact.get("ctwa_clid", "") or ""

        for msg in sorted(msgs, key=lambda m: m.get("timestamp", "")):
            ts = msg.get("timestamp", "")
            if not ts:
                continue
            date_part = ts[:10]
            if date_from and date_part < date_from:
                continue
            if date_to and date_part > date_to:
                continue
            text = msg.get("text", "") or ""
            found = _re.findall(r'\[([^\[\]]+)\]', text)
            if found:
                bracket_labels_found = found
                first_msg_date = date_part
                first_msg_text = text
                break

        if not bracket_labels_found:
            continue

        for bl in bracket_labels_found:
            bl = bl.strip()
            if not bl:
                continue
            if search_campaign and search_campaign not in bl.lower():
                continue

            if bl not in campaigns:
                campaigns[bl] = {
                    "campaign_name": bl,
                    "ad_id": ad_id,
                    "phones": set(),
                    "leads": set(),
                    "total_contacts_set": set(),
                    "view_content": set(),
                    "add_to_cart": set(),
                    "purchase": set(),
                    "contacts_detail": []
                }

            if phone not in campaigns[bl]["phones"]:
                campaigns[bl]["phones"].add(phone)
                campaigns[bl]["total_contacts_set"].add(phone)
                is_valid_ctwa = ctwa_clid and len(ctwa_clid) >= 20 and not ctwa_clid.startswith("clid_test") and not ctwa_clid.startswith("test_")
                if is_valid_ctwa:
                    campaigns[bl]["leads"].add(phone)
                if not campaigns[bl]["ad_id"] and ad_id:
                    campaigns[bl]["ad_id"] = ad_id
                campaigns[bl]["contacts_detail"].append({
                    "nama": contact.get("name", phone),
                    "phone": phone,
                    "first_msg": first_msg_text[:50] if first_msg_text else "",
                    "first_date": first_msg_date or "",
                    "ad_id": ad_id,
                    "ctwa_clid": ctwa_clid[:20] if ctwa_clid else ""
                })

    # Hitung Add to Cart dan Purchase dari transaksi
    # Hanya kontak dengan ctwa_clid valid (dari iklan Meta)
    for t in transaksi_all:
        phone = t.get("phone", "")
        status = t.get("status", "")
        contact_c = contacts.get(phone, {})
        ctwa_c = contact_c.get("ctwa_clid", "") or ""
        is_valid_c = ctwa_c and len(ctwa_c) >= 20 and not ctwa_c.startswith("clid_test") and not ctwa_c.startswith("test_")
        if not is_valid_c:
            continue
        for cname, cdata in campaigns.items():
            if phone in cdata["phones"]:
                cdata["add_to_cart"].add(phone)
                if status == "lunas":
                    cdata["purchase"].add(phone)


    # Build result
    result = []
    for cname, cdata in campaigns.items():
        result.append({
            "campaign_name": cname,
            "ad_id": cdata["ad_id"] or "-",
            "type": "META",
            "channel": "WHATSAPP",
            "total_contacts": len(cdata["total_contacts_set"]),
            "total_leads": len(cdata["leads"]),
            "total_view_content": len(cdata["view_content"]),
            "total_add_to_cart": len(cdata["add_to_cart"]),
            "total_purchase": len(cdata["purchase"]),
            "contacts": cdata["contacts_detail"]
        })

    result.sort(key=lambda x: x["total_leads"], reverse=True)
    return jsonify({"campaigns": result, "total": len(result)})

@app.route("/api/blast/template-download", methods=["GET"])
def api_blast_template_download():
    """Download template Excel untuk upload nomor blast."""
    import io
    try:
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "WA Blast"
        ws.append(["Nama", "Nomor HP"])
        ws.append(["Contoh Nama", "6281234567890"])
        ws.append(["Donatur 2", "6289876543210"])
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 20
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        as_attachment=True, download_name="wablast_template.xlsx")
    except ImportError:
        # Fallback: CSV
        import csv, io as sio
        output = sio.StringIO()
        writer = csv.writer(output)
        writer.writerow(["Nama", "Nomor HP"])
        writer.writerow(["Contoh Nama", "6281234567890"])
        writer.writerow(["Donatur 2", "6289876543210"])
        output.seek(0)
        return Response(output.getvalue(), mimetype="text/csv",
                       headers={"Content-Disposition": "attachment;filename=wablast_template.csv"})

@app.route("/api/blast/upload-contacts", methods=["POST"])
def api_blast_upload_contacts():
    """Upload file Excel/CSV berisi nomor HP untuk blast."""
    try:
        file = request.files.get("file")
        if not file:
            return jsonify({"error": "File tidak ditemukan"}), 400

        filename = file.filename.lower()
        contacts = []

        if filename.endswith(".csv"):
            import csv, io
            text = file.read().decode("utf-8", errors="ignore")
            reader = csv.DictReader(io.StringIO(text))
            for row in reader:
                nama = row.get("Nama", row.get("nama", "")).strip()
                phone = row.get("Nomor HP", row.get("nomor_hp", row.get("phone", ""))).strip()
                phone = "".join(filter(str.isdigit, phone))
                if phone:
                    if phone.startswith("0"):
                        phone = "62" + phone[1:]
                    elif not phone.startswith("62"):
                        phone = "62" + phone
                    contacts.append({"name": nama or phone, "phone": phone})

        elif filename.endswith(".xlsx") or filename.endswith(".xls"):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(file)
                ws = wb.active
                headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows())]
                nama_idx = next((i for i, h in enumerate(headers) if "nama" in h.lower()), 0)
                phone_idx = next((i for i, h in enumerate(headers) if "nomor" in h.lower() or "hp" in h.lower() or "phone" in h.lower()), 1)
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row or not row[phone_idx]:
                        continue
                    phone = "".join(filter(str.isdigit, str(row[phone_idx])))
                    if not phone:
                        continue
                    if phone.startswith("0"):
                        phone = "62" + phone[1:]
                    elif not phone.startswith("62"):
                        phone = "62" + phone
                    nama = str(row[nama_idx]).strip() if row[nama_idx] else phone
                    contacts.append({"name": nama, "phone": phone})
            except ImportError:
                return jsonify({"error": "openpyxl tidak terinstall. Gunakan format CSV."}), 400
        else:
            return jsonify({"error": "Format file harus .xlsx atau .csv"}), 400

        return jsonify({"contacts": contacts, "total": len(contacts)})
    except Exception as e:
        logger.error(f"api_blast_upload_contacts error: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500

@app.route("/api/blast/history", methods=["GET"])
def api_blast_history():
    """List history WA Blast."""
    history = load_blast_history()
    return jsonify({"history": history})

@app.route("/api/blast/history/<blast_id>", methods=["DELETE"])
def api_blast_history_delete(blast_id):
    """Hapus history blast."""
    history = load_blast_history()
    history = [h for h in history if h["id"] != blast_id]
    save_blast_history(history)
    return jsonify({"status": "deleted"})


@app.route("/api/donasi", methods=["GET"])
def api_donasi():
    """JSON data semua transaksi donasi, untuk konsumsi dashboard."""
    transaksi = load_data()
    transaksi_sorted = sorted(transaksi, key=lambda t: t.get("created_at", ""), reverse=True)

    total_nominal_lunas = sum(t.get("nominal", 0) for t in transaksi if t.get("status") == "lunas")
    total_donasi_lunas = sum(1 for t in transaksi if t.get("status") == "lunas")
    donatur_set = set(t.get("phone", "") for t in transaksi if t.get("status") == "lunas" and t.get("phone"))

    tipe_count = {}
    for t in transaksi:
        if t.get("status") == "lunas":
            tipe = t.get("tipe", "sekali")
            tipe_count[tipe] = tipe_count.get(tipe, 0) + 1

    kampanye_count = {}
    for t in transaksi:
        if t.get("status") == "lunas":
            k = t.get("kampanye", "Lainnya")
            kampanye_count[k] = kampanye_count.get(k, 0) + 1
    top_kampanye = sorted(kampanye_count.items(), key=lambda x: x[1], reverse=True)[:5]

    return jsonify({
        "summary": {
            "total_nominal": total_nominal_lunas,
            "total_donasi": total_donasi_lunas,
            "total_donatur": len(donatur_set),
        },
        "tipe_donasi": tipe_count,
        "top_kampanye": [{"nama": k, "jumlah": v} for k, v in top_kampanye],
        "transaksi": transaksi_sorted[:100],
    })

LAYOUT_CSS = """
  * { margin:0; padding:0; box-sizing:border-box; font-family:-apple-system,Segoe UI,Roboto,sans-serif; }
  body { background:#f3f4f8; color:#1f2330; display:flex; min-height:100vh; }
  .sidebar { width:220px; background:#5b3df0; color:#fff; padding:24px 0; flex-shrink:0; }
  .sidebar .logo { font-size:20px; font-weight:700; padding:0 24px 24px; }
  .sidebar a { display:flex; align-items:center; gap:10px; padding:12px 24px; color:#e0d9ff; text-decoration:none; font-size:14px; }
  .sidebar a.active { background:rgba(255,255,255,.15); color:#fff; font-weight:600; border-radius:0 20px 20px 0; }
  .sidebar a:hover { color:#fff; }
  .main { flex:1; padding:24px; min-width:0; }
  h1 { font-size:22px; margin-bottom:4px; }
  .subtitle { color:#6b7280; margin-bottom:20px; font-size:14px; }
  .cards { display:grid; grid-template-columns:repeat(auto-fit,minmax(220px,1fr)); gap:16px; margin-bottom:24px; }
  .card { background:#fff; border-radius:12px; padding:20px; box-shadow:0 1px 3px rgba(0,0,0,.08); }
  .card .icon { font-size:28px; }
  .card .value { font-size:28px; font-weight:700; color:#5b3df0; margin:8px 0 4px; }
  .card .label { color:#6b7280; font-size:14px; }
  .grid2 { display:grid; grid-template-columns:1fr 1fr; gap:16px; }
  @media (max-width: 800px) { .grid2 { grid-template-columns:1fr; } .sidebar { display:none; } }
  .panel { background:#fff; border-radius:12px; padding:20px; box-shadow:0 1px 3px rgba(0,0,0,.08); }
  .panel h2 { font-size:16px; margin-bottom:12px; }
  table { width:100%; border-collapse:collapse; font-size:13px; }
  th, td { text-align:left; padding:10px 8px; border-bottom:1px solid #eee; }
  th { color:#6b7280; font-weight:600; }
  tr:hover td { background:#f9fafb; }
  a.row-link { text-decoration:none; color:inherit; }
  .badge { padding:2px 8px; border-radius:6px; font-size:11px; font-weight:600; }
  .badge.lunas { background:#dcfce7; color:#16a34a; }
  .badge.pending { background:#fef3c7; color:#d97706; }
  .badge.gagal { background:#fee2e2; color:#dc2626; }
  .refresh { font-size:12px; color:#9ca3af; margin-top:16px; }
  .filters { display:flex; gap:12px; flex-wrap:wrap; margin-bottom:16px; }
  .filters label { font-size:12px; color:#6b7280; display:block; margin-bottom:4px; }
  .filters input, .filters select { padding:8px 10px; border:1px solid #ddd; border-radius:8px; font-size:13px; min-width:160px; }
  .btn { background:#5b3df0; color:#fff; border:none; padding:9px 20px; border-radius:8px; font-size:13px; font-weight:600; cursor:pointer; }
  .btn:hover { background:#4c30d9; }
  .back-link { display:inline-block; margin-bottom:16px; color:#5b3df0; text-decoration:none; font-size:14px; font-weight:600; }
  .detail-row { display:flex; justify-content:space-between; padding:12px 0; border-bottom:1px solid #eee; font-size:14px; }
  .detail-row .label { color:#6b7280; }
  .detail-row .val { font-weight:600; text-align:right; }

  /* Chat / Inbox */
  .chat-wrap { display:flex; height:calc(100vh - 52px); background:#fff; border-radius:12px; overflow:hidden; box-shadow:0 1px 3px rgba(0,0,0,.08); max-width:100%; }
  .chat-list { width:340px; flex-shrink:0; border-right:1px solid #eee; display:flex; flex-direction:column; }
  @media (max-width: 768px) {
    html { height: 100%; }
    body { height: 100%; overflow: hidden; }
    .main { padding:8px; height:100%; box-sizing:border-box; display:flex; flex-direction:column; overflow:hidden; }
    .chat-wrap { height:100%; border-radius:8px; overflow:hidden; width:100%; flex:1; min-height:0; }
    .chat-list { width:100%; border-right:none; position:fixed; top:0; left:0; right:0; bottom:0; z-index:10; background:#fff; transition:transform .25s; overflow:hidden; }
    .chat-list.hidden-mobile { transform:translateX(-100%); pointer-events:none; }
    .chat-panel { position:fixed; top:0; left:0; right:0; bottom:0; z-index:5; background:#fff; transform:translateX(100%); transition:transform .25s; display:flex; flex-direction:column; overflow:hidden; }
    .chat-panel.show-mobile { transform:translateX(0); }
    .chat-header { flex-shrink:0; }
    .chat-messages { flex:1; min-height:0; overflow-y:auto; -webkit-overflow-scrolling:touch; overscroll-behavior:contain; }
    .chat-input-bar { flex-shrink:0; width:100%; box-sizing:border-box; padding-bottom:env(safe-area-inset-bottom, 4px); }
    .shortcut-popup { bottom:60px; }
  }
  .mobile-back-btn { display:none; }
  @media (max-width: 768px) { .mobile-back-btn { display:inline-flex !important; } }

  .chat-tabs { display:flex; gap:8px; padding:12px; border-bottom:1px solid #eee; }
  .chat-tab { flex:1; text-align:center; padding:8px; border-radius:8px; font-size:13px; font-weight:600; cursor:pointer; background:#f3f4f8; color:#6b7280; }
  .chat-tab.active { background:#5b3df0; color:#fff; }
  .chat-items { flex:1; overflow-y:auto; }
  .chat-item { padding:14px 16px; border-bottom:1px solid #f3f4f6; cursor:pointer; display:flex; gap:10px; }
  .chat-item:hover { background:#f9fafb; }
  .chat-item.selected { background:#f0edff; }
  .chat-avatar { width:40px; height:40px; border-radius:50%; background:#e0d9ff; color:#5b3df0; display:flex; align-items:center; justify-content:center; font-weight:700; flex-shrink:0; font-size:15px; }
  .chat-item-body { flex:1; min-width:0; }
  .chat-item-top { display:flex; justify-content:space-between; align-items:baseline; gap:8px; }
  .chat-item-name { font-weight:600; font-size:14px; white-space:nowrap; overflow:hidden; text-overflow:ellipsis; }
  .chat-item-time { font-size:11px; color:#9ca3af; flex-shrink:0; }
  .chat-item-preview { font-size:12px; color:#6b7280; white-space:nowrap; overflow:hidden; text-overflow:ellipsis; margin-top:2px; }
  .chat-labels { display:flex; gap:4px; margin-top:6px; flex-wrap:wrap; }
  .chat-label { font-size:10px; font-weight:700; padding:2px 8px; border-radius:6px; background:#e0e7ff; color:#4338ca; text-transform:uppercase; }
  .chat-unread { background:#ef4444; color:#fff; font-size:11px; font-weight:700; border-radius:10px; min-width:20px; height:20px; display:flex; align-items:center; justify-content:center; padding:0 6px; }
  .chat-empty { flex:1; display:flex; align-items:center; justify-content:center; color:#9ca3af; font-size:14px; flex-direction:column; gap:12px; text-align:center; padding:40px; }
  .chat-panel { flex:1; display:flex; flex-direction:column; min-width:0; overflow:hidden; height:100%; max-width:100%; }
  .chat-header { padding:16px 20px; border-bottom:1px solid #eee; display:flex; align-items:center; gap:12px; }
  .chat-header .chat-avatar { width:36px; height:36px; font-size:13px; }
  .chat-header-name { font-weight:700; font-size:15px; }
  .chat-header-phone { font-size:12px; color:#9ca3af; }
  .chat-messages { flex:1; min-height:0; overflow-y:auto; padding:20px 20px 8px; display:flex; flex-direction:column; gap:10px; background:#f9fafb; -webkit-overflow-scrolling:touch; }
  .chat-bubble { max-width:60%; padding:10px 14px; border-radius:12px; font-size:14px; line-height:1.4; }
  .chat-bubble.in { background:#fff; align-self:flex-start; box-shadow:0 1px 2px rgba(0,0,0,.06); }
  .chat-bubble.out { background:#f3f4f6; color:#1f2937; align-self:flex-end; box-shadow:0 1px 2px rgba(0,0,0,.04); }
  .chat-bubble-time { font-size:10px; opacity:.6; margin-top:4px; text-align:right; }
  .chat-date-separator { align-self:center; background:#fff; border:1px solid #e5e7eb; color:#6b7280; font-size:12px; padding:6px 16px; border-radius:20px; margin:8px 0; box-shadow:0 1px 2px rgba(0,0,0,.04); }
  .chat-input-bar { display:flex; gap:8px; padding:10px 12px; border-top:1px solid #eee; align-items:center; position:sticky; bottom:0; background:#fff; z-index:10; box-sizing:border-box; width:100%; }
  .chat-input-bar input { flex:1; min-width:0; padding:10px 14px; border:1px solid #ddd; border-radius:20px; font-size:16px; box-sizing:border-box; }
  .chat-input-bar button { background:#5b3df0; color:#fff; border:none; border-radius:20px; padding:10px 18px; font-weight:600; font-size:14px; cursor:pointer; flex-shrink:0; white-space:nowrap; }
  .chat-input-bar button:hover { background:#4c30d9; }
  .shortcut-popup { position:absolute;bottom:70px;left:0;right:0;background:#fff;border-radius:12px 12px 0 0;box-shadow:0 -4px 24px rgba(0,0,0,.12);max-height:60vh;overflow-y:auto;z-index:300;display:none; }
  .shortcut-popup.open { display:block; }
  .shortcut-popup-header { background:#5b3df0;padding:14px 18px;display:flex;justify-content:space-between;align-items:center;border-radius:12px 12px 0 0; }
  .shortcut-popup-header span { color:#fff;font-weight:700;font-size:15px; }
  .shortcut-item { padding:14px 18px;border-bottom:1px solid #f3f4f6;cursor:pointer;transition:background .1s; }
  .shortcut-item:hover { background:#f9fafb; }
  .shortcut-item-key { font-size:13px;font-weight:700;color:#5b3df0;margin-bottom:4px; }
  .shortcut-item-body { font-size:13px;color:#6b7280;overflow:hidden;text-overflow:ellipsis;white-space:nowrap; }
  .attach-btn { flex-shrink:0; background:#f3f4f6; border:none; border-radius:50%; width:40px; height:40px; cursor:pointer; font-size:20px; display:flex; align-items:center; justify-content:center; flex-shrink:0; color:#5b3df0; transition:background .15s; }
  .attach-btn:hover { background:#e0e7ff; }
  .attach-menu { position:absolute; bottom:70px; left:16px; background:#fff; border-radius:12px; box-shadow:0 4px 24px rgba(0,0,0,.13); min-width:220px; overflow:hidden; z-index:200; display:none; }
  .attach-menu.open { display:block; }
  .attach-menu-item { display:flex; align-items:center; gap:12px; padding:13px 18px; cursor:pointer; font-size:14px; color:#1f2937; transition:background .1s; border-bottom:1px solid #f3f4f6; }
  .attach-menu-item:last-child { border-bottom:none; }
  .attach-menu-item:hover { background:#f9fafb; }
  .attach-menu-item span.icon { font-size:20px; width:28px; text-align:center; }

  /* Pengaturan */
  .settings-wrap { display:flex; gap:0; min-height:calc(100vh - 80px); }
  .settings-tabs { display:flex; flex-direction:column; width:220px; flex-shrink:0; background:#fff; border-right:1px solid #e5e7eb; border-radius:12px 0 0 12px; padding:16px 0; box-shadow:0 1px 3px rgba(0,0,0,.06); }
  .settings-tabs-group { padding:4px 12px 2px; font-size:11px; font-weight:700; color:#9ca3af; text-transform:uppercase; letter-spacing:.05em; margin-top:8px; }
  .settings-tab { display:flex; align-items:center; gap:10px; padding:10px 20px; font-size:14px; font-weight:500; cursor:pointer; color:#374151; border-left:3px solid transparent; transition:all .15s; }
  .settings-tab:hover { background:#f9fafb; color:#5b3df0; }
  .settings-tab.active { background:#f0edff; color:#5b3df0; font-weight:700; border-left:3px solid #5b3df0; }
  .settings-tab .tab-icon { font-size:16px; flex-shrink:0; }
  .settings-content { flex:1; padding:24px; background:#f9fafb; border-radius:0 12px 12px 0; min-width:0; }
  .settings-section { display:none; }
  .settings-section.active { display:block; }
  @media (max-width: 768px) {
    .settings-wrap { flex-direction:column; }
    .settings-tabs { width:100%; flex-direction:row; overflow-x:auto; border-right:none; border-bottom:1px solid #e5e7eb; border-radius:12px 12px 0 0; padding:8px; gap:4px; }
    .settings-tabs-group { display:none; }
    .settings-tab { border-left:none; border-bottom:3px solid transparent; border-radius:8px; padding:8px 12px; white-space:nowrap; font-size:12px; }
    .settings-tab.active { border-left:none; border-bottom:3px solid #5b3df0; }
    .settings-content { border-radius:0 0 12px 12px; padding:16px; }
  }
  .form-group { margin-bottom:16px; }
  .form-group label { display:block; font-size:13px; font-weight:600; margin-bottom:6px; color:#374151; }
  .form-group textarea, .form-group input[type=text] { width:100%; padding:10px 12px; border:1px solid #ddd; border-radius:8px; font-size:14px; font-family:inherit; }
  .form-group textarea { min-height:100px; resize:vertical; }
  .form-hint { font-size:11px; color:#9ca3af; margin-top:4px; }
  .button-row { display:flex; gap:12px; align-items:center; padding:10px 0; border-bottom:1px solid #f3f4f6; }
  .button-row:last-child { border-bottom:none; }
  .button-row input[type=text] { flex:1; padding:8px 10px; border:1px solid #ddd; border-radius:8px; font-size:13px; }
  .button-row select { padding:8px 10px; border:1px solid #ddd; border-radius:8px; font-size:13px; }
  .save-msg { color:#16a34a; font-size:13px; margin-left:12px; }
  .blast-contacts { max-height:260px; overflow-y:auto; border:1px solid #eee; border-radius:8px; padding:8px; margin-bottom:12px; }
  .blast-contact-item { display:flex; align-items:center; gap:10px; padding:8px; border-radius:6px; }
  .blast-contact-item:hover { background:#f9fafb; }
  .blast-count { font-size:12px; color:#6b7280; margin-bottom:8px; }
"""

def render_sidebar(active):
    items = [
        ("dashboard", "/dashboard", "&#x1F4CA;", "Dashboard"),
        ("pesanan", "/pesanan", "&#x1F4CB;", "Transaksi"),
        ("chat", "/chat", "&#x1F4AC;", "Chat"),
        ("kampanye", "/kampanye", "&#x1F3AF;", "Kampanye"),
        ("whatsapp", "/whatsapp", "WA", "WhatsApp"),

    ]
    html_links = ""
    for key, url, icon, label in items:
        css = "active" if key == active else ""
        if key == "whatsapp":
            icon_html = '<img src="/static/iconWA.png" style="width:20px;height:20px;object-fit:contain;filter:brightness(0) invert(1);vertical-align:middle;">'
        else:
            icon_html = f'<span>{icon}</span>'
        html_links += f'<a href="{url}" class="{css}"><span>{icon_html}</span><span>{label}</span></a>'
    
    is_admin = session.get("user_role") == "ADMINISTRATOR"
    pengaturan_open = "open" if active == "pengaturan" else ""
    
    sub_link_style = "display:flex;align-items:center;gap:8px;padding:8px 16px;color:rgba(255,255,255,.75);font-size:13px;text-decoration:none;border-radius:8px;"
    laporan_open = "open" if active in ("laporan", "laporan-summary") else ""
    laporan_block = (
        f'<details {laporan_open} style="margin-top:4px;">' +
        '<summary style="display:flex;align-items:center;gap:10px;padding:10px 20px;cursor:pointer;' +
        'color:rgba(255,255,255,.85);font-size:14px;font-weight:500;list-style:none;border-radius:8px;">' +
        '<span>&#x1F4CA;</span><span>Laporan</span>' +
        '<span style="margin-left:auto;font-size:11px;opacity:.6;">&#9660;</span></summary>' +
        '<div style="padding:4px 0 4px 16px;">' +
        f'<a href="/laporan/chat-harian" style="{sub_link_style}"><span>&#x1F4AC;</span> Chat Harian</a>' +
        f'<a href="/laporan/summary" style="{sub_link_style}"><span>&#x1F4CB;</span> Summary</a>' +
        f'<a href="/laporan/contact-label" style="{sub_link_style}"><span>&#x1F3F7;</span> Contact Label</a>' +
        f'<a href="/laporan/tracking-campaign" style="{sub_link_style}"><span>&#x1F4CA;</span> Tracking Campaign</a>' +
        '</div></details>'
    )

    if is_admin:
        pengaturan_block = (
            f'<details {pengaturan_open} style="margin-top:4px;">' +
            '<summary style="display:flex;align-items:center;gap:10px;padding:10px 20px;cursor:pointer;' +
            'color:rgba(255,255,255,.85);font-size:14px;font-weight:500;list-style:none;border-radius:8px;">' +
            '<span>&#x2699;&#xFE0F;</span><span>Pengaturan</span>' +
            '<span style="margin-left:auto;font-size:11px;opacity:.6;">&#9660;</span></summary>' +
            '<div style="padding:4px 0 4px 16px;">' +
            f'<a href="/pengaturan/pengguna" style="{sub_link_style}">' +
            '<span>&#x1F465;</span> Daftar Pengguna</a>' +
            '</div></details>'
        )
    else:
        pengaturan_block = ""
    
    logout_style = "display:flex;align-items:center;gap:10px;padding:10px 20px;color:rgba(255,255,255,.6);font-size:14px;text-decoration:none;margin-top:8px;border-radius:8px;"
    logout_block = f'<a href="/logout" style="{logout_style}"><span>&#x1F6AA;</span> Logout</a>'
    
    user_nama = session.get("user_nama", "")
    user_role = session.get("user_role", "")
    
    return (
        '<div class="sidebar">' +
        '<div class="logo" style="padding:12px 16px;">' +
        '<img src="/static/IconRM.png" style="height:36px;object-fit:contain;"></div>' +
        '<div style="padding:8px 16px 12px;border-bottom:1px solid rgba(255,255,255,.1);margin-bottom:4px;">' +
        '<div style="font-size:12px;color:rgba(255,255,255,.7);">Masuk sebagai</div>' +
        f'<div style="font-size:13px;font-weight:700;color:#fff;">{user_nama}</div>' +
        '</div>' +
        html_links +
        laporan_block +
        pengaturan_block +
        logout_block +
        '</div>'
    )


def render_page(active, title, subtitle, body_html, extra_head=""):
    return f"""<!DOCTYPE html>
<html lang="id">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0, maximum-scale=1.0, user-scalable=no, viewport-fit=cover">
<title>{title} - Raihmimpi</title>
{extra_head}
<style>{LAYOUT_CSS}</style>
</head>
<body>
{render_sidebar(active)}
<div class="main">
  {f'<h1>{title}</h1>' if title else ''}
  {f'<div class="subtitle">{subtitle}</div>' if subtitle else ''}
  {body_html}
</div>
</body>
</html>"""

@app.route("/dashboard", methods=["GET"])
@login_required
def dashboard():
    """Dashboard donasi sederhana."""
    body = """
  <div class="cards">
    <div class="card"><div class="icon">&#128176;</div><div class="value" id="total_nominal">-</div><div class="label">Total Nominal Donasi (Lunas)</div></div>
    <div class="card"><div class="icon">&#128203;</div><div class="value" id="total_donasi">-</div><div class="label">Total Donasi (Lunas)</div></div>
    <div class="card"><div class="icon">&#128101;</div><div class="value" id="total_donatur">-</div><div class="label">Total Donatur</div></div>
  </div>

  <div class="grid2">
    <div class="panel">
      <h2>Tipe Donasi</h2>
      <canvas id="chartTipe" height="220"></canvas>
    </div>
    <div class="panel">
      <h2>Kampanye Terbanyak</h2>
      <div id="topKampanye"></div>
    </div>
  </div>

  <div class="panel" style="margin-top:16px;">
    <h2>Daftar Donasi Terbaru</h2>
    <table>
      <thead><tr><th>Waktu</th><th>Donatur</th><th>Kampanye</th><th>Nominal</th><th>Tipe</th><th>Status</th></tr></thead>
      <tbody id="tabelDonasi"></tbody>
    </table>
  </div>
  <div class="refresh">Auto-refresh setiap 30 detik &middot; <span id="lastUpdate"></span></div>

<script>
function formatRupiah(n) {
  return "Rp " + Number(n).toLocaleString("id-ID");
}
function formatWaktu(iso) {
  if (!iso) return "-";
  const d = new Date(iso);
  return d.toLocaleString("id-ID", {dateStyle:"medium", timeStyle:"short"});
}
let chart;
async function loadData() {
  const res = await fetch("/api/donasi");
  const json = await res.json();

  document.getElementById("total_nominal").textContent = formatRupiah(json.summary.total_nominal);
  document.getElementById("total_donasi").textContent = json.summary.total_donasi;
  document.getElementById("total_donatur").textContent = json.summary.total_donatur;

  const tipeLabels = Object.keys(json.tipe_donasi).map(t => t === "sekali" ? "Donasi Sekali" : "Donasi Rutin");
  const tipeData = Object.values(json.tipe_donasi);
  if (chart) chart.destroy();
  chart = new Chart(document.getElementById("chartTipe"), {
    type: "pie",
    data: { labels: tipeLabels.length ? tipeLabels : ["Belum ada data"], datasets: [{ data: tipeData.length ? tipeData : [1], backgroundColor: ["#f97316","#5b3df0","#10b981","#f43f5e"] }] },
    options: { plugins: { legend: { position: "bottom" } } }
  });

  const topKampanyeEl = document.getElementById("topKampanye");
  topKampanyeEl.innerHTML = json.top_kampanye.length
    ? json.top_kampanye.map(k => `<div style="display:flex;justify-content:space-between;padding:8px 0;border-bottom:1px solid #eee;"><span>${k.nama}</span><strong>${k.jumlah} donasi</strong></div>`).join("")
    : "<div style='color:#9ca3af'>Belum ada data</div>";

  const tbody = document.getElementById("tabelDonasi");
  tbody.innerHTML = json.transaksi.map(t => `
    <tr class="row-link" onclick="window.location='/pesanan/${t.order_id}'" style="cursor:pointer">
      <td>${formatWaktu(t.created_at)}</td>
      <td>${t.donatur || "-"}</td>
      <td>${t.kampanye || "-"}</td>
      <td>${formatRupiah(t.nominal || 0)}</td>
      <td>${t.tipe === "rutin" ? "Rutin" : "Sekali"}</td>
      <td><span class="badge ${t.status}">${t.status}</span></td>
    </tr>
  `).join("") || "<tr><td colspan='6' style='text-align:center;color:#9ca3af'>Belum ada donasi</td></tr>";

  document.getElementById("lastUpdate").textContent = "Update: " + new Date().toLocaleTimeString("id-ID");
}
loadData();
setInterval(loadData, 30000);
</script>
"""
    return Response(render_page("dashboard", "Dashboard Donasi Raihmimpi", "via WhatsApp Flow &middot; +62 851-1123-4962", body,
        extra_head='<script src="https://cdn.jsdelivr.net/npm/chart.js"></script>'), mimetype="text/html")


@app.route("/pesanan", methods=["GET"])
@login_required
def pesanan():
    """Daftar pesanan donasi dengan filter."""
    body = """
  <div class="panel">
    <div class="filters">
      <div>
        <label>No. Transaksi</label>
        <input type="text" id="f_order" placeholder="RM-...">
      </div>
      <div>
        <label>Nama Donatur</label>
        <input type="text" id="f_donatur" placeholder="Nama donatur">
      </div>
      <div>
        <label>Nomor HP</label>
        <input type="text" id="f_phone" placeholder="62...">
      </div>
      <div>
        <label>Status Donasi</label>
        <select id="f_status">
          <option value="">(Semua)</option>
          <option value="pending">Pending</option>
          <option value="lunas">Lunas</option>
          <option value="gagal">Gagal</option>
        </select>
      </div>
      <div style="align-self:flex-end;display:flex;gap:8px;">
        <button class="btn" onclick="loadData()">Cari</button>
        <button class="btn" onclick="syncMidtransAll()" id="btnSyncAll" style="background:#10b981;" title="Cek manual status semua transaksi pending ke Midtrans (auto-poll tiap 5 menit di background)">Sinkron Midtrans</button>
      </div>
    </div>
    <table>
      <thead><tr>
        <th>No. Transaksi</th><th>Tanggal</th><th>Nama Donatur</th><th>Nomor HP</th>
        <th>Nama Donasi</th><th>Tipe</th><th>Total</th><th>Status</th>
      </tr></thead>
      <tbody id="tabelPesanan"></tbody>
    </table>
  </div>
  <div class="refresh">Auto-refresh setiap 30 detik &middot; <span id="lastUpdate"></span></div>

<script>
function formatRupiah(n) {
  return "Rp " + Number(n).toLocaleString("id-ID");
}
function formatWaktu(iso) {
  if (!iso) return "-";
  const d = new Date(iso);
  return d.toLocaleString("id-ID", {dateStyle:"medium", timeStyle:"short"});
}
let allData = [];
async function loadData() {
  const res = await fetch("/api/donasi");
  const json = await res.json();
  allData = json.transaksi;
  render();
  document.getElementById("lastUpdate").textContent = "Update: " + new Date().toLocaleTimeString("id-ID");
}
function render() {
  const fOrder = document.getElementById("f_order").value.toLowerCase();
  const fDonatur = document.getElementById("f_donatur").value.toLowerCase();
  const fPhone = document.getElementById("f_phone").value.toLowerCase();
  const fStatus = document.getElementById("f_status").value;

  const filtered = allData.filter(t => {
    if (fOrder && !(t.order_id || "").toLowerCase().includes(fOrder)) return false;
    if (fDonatur && !(t.donatur || "").toLowerCase().includes(fDonatur)) return false;
    if (fPhone && !(t.phone || "").toLowerCase().includes(fPhone)) return false;
    if (fStatus && t.status !== fStatus) return false;
    return true;
  });

  const tbody = document.getElementById("tabelPesanan");
  tbody.innerHTML = filtered.map(t => `
    <tr class="row-link" onclick="window.location='/pesanan/${t.order_id}'" style="cursor:pointer">
      <td>${t.order_id}</td>
      <td>${formatWaktu(t.created_at)}</td>
      <td>${t.donatur || "-"}</td>
      <td>${t.phone || "-"}</td>
      <td>${t.kampanye || "-"}</td>
      <td>${t.tipe === "rutin" ? "Rutin" : "Sekali"}</td>
      <td>${formatRupiah(t.nominal || 0)}</td>
      <td><span class="badge ${t.status}">${t.status}</span></td>
    </tr>
  `).join("") || "<tr><td colspan='8' style='text-align:center;color:#9ca3af'>Tidak ada pesanan ditemukan</td></tr>";
}
async function syncMidtransAll() {
  const btn = document.getElementById("btnSyncAll");
  const pendingCount = allData.filter(t => t.status === "pending").length;
  if (!pendingCount) {
    alert("Tidak ada transaksi pending untuk disinkronkan.");
    return;
  }
  if (!confirm("Sinkronisasi " + pendingCount + " transaksi pending dengan Midtrans? Proses memakan beberapa detik per transaksi. (Sistem juga auto-poll tiap 5 menit di background.)")) return;
  btn.disabled = true;
  btn.textContent = "Menyinkronkan...";
  try {
    const res = await fetch("/api/sync-midtrans-all", {method: "POST"});
    const json = await res.json();
    let msg = "Sinkronisasi selesai. Total pending diperiksa: " + json.checked + ", Berubah status: " + json.changed + ", Menjadi LUNAS: " + json.lunas + ", Menjadi GAGAL: " + json.gagal + (json.errors ? ", Error: " + json.errors : "");
    alert(msg);
    loadData();
  } catch (e) {
    alert("Error: " + e.message);
  } finally {
    btn.disabled = false;
    btn.textContent = "Sinkron Midtrans";
  }
}

['f_order','f_donatur','f_phone'].forEach(id => document.getElementById(id).addEventListener("keyup", render));
document.getElementById("f_status").addEventListener("change", render);
loadData();
setInterval(loadData, 30000);
</script>
"""
    return Response(render_page("pesanan", "Daftar Pesanan Donasi", "Semua transaksi donasi via WhatsApp Flow", body), mimetype="text/html")


@app.route("/pesanan/<order_id>", methods=["GET"])
@login_required
def pesanan_detail(order_id):
    """Detail satu pesanan donasi."""
    transaksi = load_data()
    t = next((x for x in transaksi if x.get("order_id") == order_id), None)

    if not t:
        body = '<div class="panel"><p>Pesanan tidak ditemukan.</p><a href="/pesanan" class="back-link">&larr; Kembali ke Daftar Pesanan</a></div>'
        return Response(render_page("pesanan", "Detail Pesanan", order_id, body), mimetype="text/html")

    def fmt_waktu(iso):
        if not iso:
            return "-"
        try:
            dt = datetime.fromisoformat(iso)
            return dt.strftime("%d %b %Y, %H:%M")
        except Exception:
            return iso

    status_badge = f'<span class="badge {t.get("status","pending")}">{t.get("status","pending")}</span>'

    rows = [
        ("No. Transaksi", t.get("order_id", "-")),
        ("Tanggal Dibuat", fmt_waktu(t.get("created_at"))),
        ("Tanggal Lunas", fmt_waktu(t.get("paid_at")) if t.get("paid_at") else "-"),
        ("Status", status_badge),
        ("Nama Donatur", t.get("donatur", "-")),
        ("Atas Nama", t.get("atas_nama", "-")),
        ("Nomor HP", t.get("phone", "-")),
        ("Kampanye", t.get("kampanye", "-")),
        ("ID Kampanye", t.get("kampanye_id", "-")),
        ("Tipe Donasi", "Donasi Rutin" if t.get("tipe") == "rutin" else "Donasi Sekali"),
        ("Nominal", "Rp " + "{:,}".format(t.get("nominal", 0)).replace(",", ".")),
    ]
    if t.get("payment_url"):
        rows.append(("Link Pembayaran", f'<a href="{t["payment_url"]}" target="_blank" style="color:#5b3df0;">{t["payment_url"][:50]}...</a>'))

    rows_html = "".join(
        f'<div class="detail-row"><span class="label">{label}</span><span class="val">{value}</span></div>'
        for label, value in rows
    )

    body = f"""
  <a href="/pesanan" class="back-link">&larr; Kembali ke Daftar Pesanan</a>
  <div class="panel">
    <h2>Informasi Donasi</h2>
    {rows_html}
  </div>
"""
    return Response(render_page("pesanan", "Detail Pesanan", t.get("order_id", order_id), body), mimetype="text/html")


@app.route("/chat", methods=["GET"])
@login_required
def chat_page():
    """Halaman Inbox/Chat - list kontak (kiri) + panel percakapan (kanan)."""
    body = """
  <div class="chat-wrap">
    <div class="chat-list">
      <!-- Toolbar ala Halosis -->
      <div style="background:#5b3df0;padding:10px 14px;display:flex;align-items:center;justify-content:space-between;gap:8px;flex-shrink:0;">
        <div style="display:flex;align-items:center;gap:6px;">
          <img src="/static/iconWA.png" style="width:18px;height:18px;filter:brightness(0) invert(1);flex-shrink:0;">
          <div style="line-height:1.2;">
            <div style="color:#fff;font-weight:700;font-size:11px;">WhatsApp</div>
            <div style="color:#e0d7ff;font-size:10px;">+62 851-1123-4962</div>
          </div>
        </div>
        <div style="display:flex;align-items:center;gap:4px;">
          <button onclick="resolveAll()" title="Resolve All" style="background:rgba(255,255,255,.15);border:none;color:#fff;padding:5px 8px;border-radius:6px;cursor:pointer;font-size:10px;font-weight:600;white-space:nowrap;">✓ Resolve</button>
          <button onclick="toggleChatSearch()" title="Cari" style="background:rgba(255,255,255,.15);border:none;color:#fff;width:28px;height:28px;border-radius:6px;cursor:pointer;font-size:14px;">🔍</button>
          <button onclick="toggleChatFilter()" title="Filter" style="background:rgba(255,255,255,.15);border:none;color:#fff;width:28px;height:28px;border-radius:6px;cursor:pointer;font-size:14px;">⚙</button>
          <button onclick="toggleChatSort()" title="Urutkan" id="sortBtn" style="background:rgba(255,255,255,.15);border:none;color:#fff;width:28px;height:28px;border-radius:6px;cursor:pointer;font-size:14px;">↕</button>
        </div>
      </div>

      <!-- Search bar (hidden by default) -->
      <div id="chatSearchBar" style="display:none;padding:8px 12px;border-bottom:1px solid #eee;background:#fff;">
        <input type="text" id="chatSearchInput" placeholder="Cari nama atau nomor HP..." oninput="onChatSearch(this.value)"
          style="width:100%;padding:8px 12px;border:1px solid #ddd;border-radius:20px;font-size:13px;box-sizing:border-box;outline:none;">
      </div>

      <!-- Filter panel (hidden by default) -->
      <div id="chatFilterPanel" style="display:none;background:#f9fafb;border-bottom:1px solid #eee;padding:12px;">
        <div style="font-weight:700;font-size:13px;margin-bottom:10px;">Filter</div>
        <div style="font-size:12px;font-weight:600;color:#6b7280;margin-bottom:6px;">Filter Tanggal Pesan</div>
        <div style="display:flex;gap:8px;align-items:center;margin-bottom:12px;">
          <input type="date" id="filterDateFrom" style="flex:1;padding:7px;border:1px solid #d1d5db;border-radius:8px;font-size:12px;" onchange="applyFilter()">
          <span style="color:#9ca3af;font-size:12px;">→</span>
          <input type="date" id="filterDateTo" style="flex:1;padding:7px;border:1px solid #d1d5db;border-radius:8px;font-size:12px;" onchange="applyFilter()">
        </div>
        <div style="font-size:12px;font-weight:600;color:#6b7280;margin-bottom:6px;">Pilih Contact Label</div>
        <div style="position:relative;">
          <input type="text" id="filterLabelInput" placeholder="Cari berdasarkan label..." oninput="filterLabelSearch(this.value)"
            style="width:100%;padding:8px 32px 8px 10px;border:1px solid #d1d5db;border-radius:8px;font-size:12px;box-sizing:border-box;">
          <span style="position:absolute;right:10px;top:50%;transform:translateY(-50%);color:#9ca3af;">🔍</span>
        </div>
        <div id="filterLabelOptions" style="margin-top:6px;display:flex;flex-wrap:wrap;gap:4px;"></div>
        <div style="margin-top:10px;display:flex;gap:8px;">
          <button onclick="clearFilter()" style="flex:1;padding:7px;border:1px solid #d1d5db;border-radius:8px;font-size:12px;background:#fff;cursor:pointer;">Reset</button>
          <button onclick="applyFilter();toggleChatFilter()" style="flex:1;padding:7px;background:#5b3df0;color:#fff;border:none;border-radius:8px;font-size:12px;cursor:pointer;font-weight:600;">Terapkan</button>
        </div>
      </div>

      <div class="chat-tabs">
        <div class="chat-tab active" data-status="perlu_dibalas" onclick="setTab(this)">Perlu Dibalas</div>
        <div class="chat-tab" data-status="otomatis" onclick="setTab(this)">Otomatis</div>
        <div class="chat-tab" data-status="selesai" onclick="setTab(this)">Selesai</div>
      </div>
      <div class="chat-items" id="chatItems"></div>
    </div>
    <div class="chat-panel" id="chatPanel">
      <div class="chat-empty">
        <div style="font-size:40px;">💬</div>
        <div>Pilih kontak di sebelah kiri<br>untuk memulai percakapan</div>
      </div>
    </div>
  </div>

<script>
let currentTab = "perlu_dibalas";
let currentPhone = null;
let allContacts = [];
let chatSearchQuery = "";
let chatSortOrder = "desc";
let chatFilterLabel = "";
let chatFilterDateFrom = "";
let chatFilterDateTo = "";

// ---- Toolbar functions ----
function toggleChatSearch() {
  const bar = document.getElementById("chatSearchBar");
  const fp = document.getElementById("chatFilterPanel");
  if (fp) fp.style.display = "none";
  if (bar) {
    bar.style.display = bar.style.display === "none" ? "block" : "none";
    if (bar.style.display === "block") setTimeout(() => document.getElementById("chatSearchInput").focus(), 100);
  }
}

function toggleChatFilter() {
  const fp = document.getElementById("chatFilterPanel");
  const bar = document.getElementById("chatSearchBar");
  if (bar) bar.style.display = "none";
  if (fp) {
    fp.style.display = fp.style.display === "none" ? "block" : "none";
    if (fp.style.display === "block") renderFilterLabels();
  }
}

function toggleChatSort() {
  chatSortOrder = chatSortOrder === "desc" ? "asc" : "desc";
  const btn = document.getElementById("sortBtn");
  if (btn) btn.title = chatSortOrder === "desc" ? "Terbaru → Terlama" : "Terlama → Terbaru";
  renderItems();
}

function onChatSearch(val) {
  chatSearchQuery = val.toLowerCase();
  renderItems();
}

function filterLabelSearch(val) {
  renderFilterLabels(val);
}

function renderFilterLabels(search) {
  const el = document.getElementById("filterLabelOptions");
  if (!el) return;
  const labels = window._labelAllLabels || [];
  const filtered = search ? labels.filter(l => l.name.toLowerCase().includes(search.toLowerCase())) : labels;
  el.innerHTML = filtered.map(l => {
    const isSelected = chatFilterLabel === l.name;
    const bg = isSelected ? (l.bg_color||"#5b3df0") : "#f3f4f6";
    const color = isSelected ? (l.text_color||"#fff") : "#374151";
    return `<span onclick="selectFilterLabel('${l.name.replace(/'/g,"\'")}',this)"
      style="padding:4px 10px;border-radius:12px;font-size:11px;font-weight:600;cursor:pointer;background:${bg};color:${color};">
      ${l.name}</span>`;
  }).join("");
}

function selectFilterLabel(name) {
  chatFilterLabel = chatFilterLabel === name ? "" : name;
  renderFilterLabels(document.getElementById("filterLabelInput").value);
  applyFilter();
}

function applyFilter() {
  chatFilterDateFrom = document.getElementById("filterDateFrom").value;
  chatFilterDateTo = document.getElementById("filterDateTo").value;
  renderItems();
}

function clearFilter() {
  chatFilterLabel = "";
  chatFilterDateFrom = "";
  chatFilterDateTo = "";
  document.getElementById("filterDateFrom").value = "";
  document.getElementById("filterDateTo").value = "";
  document.getElementById("filterLabelInput").value = "";
  renderFilterLabels();
  renderItems();
}

async function resolveAll() {
  if (!confirm("Tandai semua kontak sebagai Selesai?")) return;
  try {
    await fetch("/api/inbox/reset-menu-all", {method:"POST"});
    loadContacts();
  } catch(e) { alert("Error: " + e.message); }
}

// Load master labels untuk render warna di contact list
window._labelAllLabels = [];
(async function() {
  try {
    const res = await fetch("/api/labels");
    const json = await res.json();
    window._labelAllLabels = json.labels || [];
  } catch(e) {}
})();

function initials(name) {
  if (!name) return "?";
  const parts = name.trim().split(" ");
  return (parts[0][0] + (parts[1] ? parts[1][0] : "")).toUpperCase();
}
function formatTime(iso) {
  if (!iso) return "";
  // Asumsikan timestamp tanpa Z/offset adalah UTC (Railway server pakai UTC)
  // Tambah Z supaya JS parse sebagai UTC, lalu tampilkan dalam WIB (Asia/Jakarta)
  let isoFixed = iso;
  if (!/Z|[+-]\d{2}:?\d{2}$/.test(iso)) {
    isoFixed = iso + "Z";
  }
  const d = new Date(isoFixed);
  if (isNaN(d.getTime())) return "";
  const now = new Date();
  const tzOpts = {timeZone: "Asia/Jakarta"};
  // Bandingkan tanggal dalam WIB
  const dStr = d.toLocaleDateString("id-ID", tzOpts);
  const nowStr = now.toLocaleDateString("id-ID", tzOpts);
  if (dStr === nowStr) {
    return d.toLocaleTimeString("id-ID", {...tzOpts, hour:"2-digit", minute:"2-digit"});
  }
  return d.toLocaleDateString("id-ID", {...tzOpts, day:"numeric", month:"short"});
}

function setTab(el) {
  document.querySelectorAll(".chat-tab").forEach(t => t.classList.remove("active"));
  el.classList.add("active");
  currentTab = el.dataset.status;
  renderItems();
}

async function loadContacts() {
  const res = await fetch("/api/inbox");
  const json = await res.json();
  allContacts = json.contacts || [];
  renderItems();
}

function renderItems() {
  let filtered = allContacts.filter(c => (c.status || "perlu_dibalas") === currentTab);

  // Search
  if (chatSearchQuery) {
    filtered = filtered.filter(c =>
      (c.name||"").toLowerCase().includes(chatSearchQuery) ||
      (c.phone||"").includes(chatSearchQuery)
    );
  }
  // Label filter
  if (chatFilterLabel) {
    filtered = filtered.filter(c => (c.labels||[]).includes(chatFilterLabel));
  }
  // Date filter
  if (chatFilterDateFrom || chatFilterDateTo) {
    filtered = filtered.filter(c => {
      if (!c.last_message_at) return false;
      const d = c.last_message_at.substring(0,10);
      if (chatFilterDateFrom && d < chatFilterDateFrom) return false;
      if (chatFilterDateTo && d > chatFilterDateTo) return false;
      return true;
    });
  }
  // Sort
  filtered = filtered.slice().sort((a,b) => {
    const ta = a.last_message_at || "";
    const tb = b.last_message_at || "";
    return chatSortOrder === "desc" ? tb.localeCompare(ta) : ta.localeCompare(tb);
  });

  const el = document.getElementById("chatItems");
  if (!filtered.length) {
    el.innerHTML = `<div class="chat-empty">Tidak ada percakapan</div>`;
    return;
  }
  el.innerHTML = filtered.map(c => `
    <div class="chat-item ${c.phone === currentPhone ? 'selected' : ''}" onclick="openChat('${c.phone}')">
      <div class="chat-avatar">${initials(c.name)}</div>
      <div class="chat-item-body">
        <div class="chat-item-top">
          <div class="chat-item-name">${c.name || c.phone}</div>
          <div class="chat-item-time">${formatTime(c.last_message_at)}</div>
        </div>
        <div class="chat-item-preview">${(c.last_message || "").replace(/</g,"&lt;")}</div>
        ${(c.labels && c.labels.length) ? `<div class="chat-labels">${c.labels.map(l => {
          const lab = (window._labelAllLabels||[]).find(x => x.name === l);
          const bg = lab ? (lab.bg_color||"#e0e7ff") : "#e0e7ff";
          const color = lab ? (lab.text_color||"#4338ca") : "#4338ca";
          return `<span class="chat-label" style="background:${bg};color:${color};">${l}</span>`;
        }).join("")}</div>` : ""}
      </div>
      ${c.unread ? `<div class="chat-unread">${c.unread}</div>` : ""}
    </div>
  `).join("");
}

async function openChat(phone) {
  currentPhone = phone;
  renderItems();
  const res = await fetch(`/api/inbox/${phone}`);
  const json = await res.json();
  const contact = json.contact;
  const messages = json.messages || [];

  const panel = document.getElementById("chatPanel");
  panel.innerHTML = `
    <div class="chat-header" style="position:relative;background:#5b3df0;padding:12px 16px;display:flex;align-items:center;gap:10px;">
      <span class="mobile-back-btn" onclick="closeChatMobile()" style="color:#fff;font-size:22px;cursor:pointer;padding:0 8px 0 0;display:none;">‹</span>
      <div style="flex:1;min-width:0;">
        <div class="chat-header-name" style="color:#fff;font-size:15px;font-weight:700;">${contact.name || contact.phone}</div>
        <div class="chat-header-phone" style="color:#c4b5fd;font-size:12px;">+${contact.phone}</div>
      </div>
      <div style="position:relative;">
        <button onclick="toggleContactMenu(event)" id="contactMenuBtn" style="background:rgba(255,255,255,.15);border:none;color:#fff;width:36px;height:36px;border-radius:50%;cursor:pointer;font-size:20px;line-height:1;display:flex;align-items:center;justify-content:center;" title="Aksi kontak">⋮</button>
        <div id="contactMenuDropdown" style="display:none;position:absolute;right:0;top:42px;background:#fff;border:1px solid #e5e7eb;border-radius:8px;box-shadow:0 4px 12px rgba(0,0,0,0.08);min-width:180px;z-index:100;overflow:hidden;">
          <div onclick="resetMenuContact('${contact.phone}')" style="padding:10px 14px;cursor:pointer;font-size:14px;display:flex;align-items:center;gap:8px;" onmouseover="this.style.background='#f3f4f6'" onmouseout="this.style.background='#fff'">↻ Reset Menu</div>
          <div onclick="markResolved('${contact.phone}')" style="padding:10px 14px;cursor:pointer;font-size:14px;display:flex;align-items:center;gap:8px;" onmouseover="this.style.background='#f3f4f6'" onmouseout="this.style.background='#fff'">✓ Selesai</div>
          <div onclick="editContactLabel('${contact.phone}')" style="padding:10px 14px;cursor:pointer;font-size:14px;display:flex;align-items:center;gap:8px;" onmouseover="this.style.background='#f3f4f6'" onmouseout="this.style.background='#fff'">🏷 Contact Label</div>
          <div onclick="blokirContact('${contact.phone}')" style="padding:10px 14px;cursor:pointer;font-size:14px;color:#dc2626;display:flex;align-items:center;gap:8px;border-top:1px solid #f3f4f6;" onmouseover="this.style.background='#fef2f2'" onmouseout="this.style.background='#fff'">⊘ Blokir</div>
        </div>
      </div>
    </div>
    <div class="chat-messages" id="chatMessages"></div>
    <div class="chat-input-bar">
      <!-- Shortcut popup (trigger /) -->
      <div id="shortcutPopup" class="shortcut-popup">
        <div class="shortcut-popup-header">
          <span>⚡ List Shortcut</span>
          <span onclick="closeShortcutPopup()" style="cursor:pointer;font-size:20px;">✕</span>
        </div>
        <div id="shortcutPopupList"></div>
      </div>
      <div id="attachMenu" class="attach-menu">
        <div id="attachTemplateItems"></div>
        <div class="attach-menu-item" onclick="sendFlowAttachment()">
          <span class="icon">📋</span><span>Flow Donasi</span>
        </div>
        <div class="attach-menu-item" onclick="sendMediaAttachment('document')">
          <span class="icon">📄</span><span>Document</span>
        </div>
        <div class="attach-menu-item" onclick="sendMediaAttachment('image')">
          <span class="icon">🖼</span><span>Media Image/Video</span>
        </div>
        <div class="attach-menu-item" onclick="sendMediaAttachment('audio')">
          <span class="icon">🎵</span><span>Media Audio</span>
        </div>
      </div>
      <button class="attach-btn" onclick="toggleAttachMenu(event)" title="Lampiran">＋</button>
      <input type="text" id="chatInput" placeholder="Tulis balasan..." onkeydown="if(event.key==='Enter'){closeShortcutPopup();sendReply();}" oninput="onChatInput(this.value)">
      <button onclick="sendReply()">Kirim</button>
    </div>
  `;

  // Mobile: sembunyikan list, tampilkan panel
  if (window.innerWidth <= 768) {
    const chatList = document.querySelector(".chat-list");
    const chatPanel = document.getElementById("chatPanel");
    if (chatList) chatList.classList.add("hidden-mobile");
    if (chatPanel) chatPanel.classList.add("show-mobile");
  }

  const msgEl = document.getElementById("chatMessages");

  // Render ad referral card jika kontak datang dari CTWA ad
  let referralCardHtml = "";
  if (contact.ad_headline || contact.ad_source_url) {
    const adImg = contact.ad_image_url || contact.ad_thumbnail_url || "";
    const adHeadline = (contact.ad_headline || "Iklan Raihmimpi").replace(/</g,"&lt;");
    const adBody = (contact.ad_body || "").replace(/</g,"&lt;");
    const adUrl = contact.ad_source_url || "";
    referralCardHtml = `
      <div style="background:#f3f4f6;border-left:4px solid #5b3df0;border-radius:8px;padding:10px;margin-bottom:12px;display:flex;gap:10px;align-items:flex-start;max-width:75%;">
        ${adImg ? `<img src="${adImg}" style="width:64px;height:64px;object-fit:cover;border-radius:6px;flex-shrink:0;" onerror="this.style.display='none'">` : `<div style="width:64px;height:64px;background:#e5e7eb;border-radius:6px;flex-shrink:0;display:flex;align-items:center;justify-content:center;font-size:24px;">📢</div>`}
        <div style="flex:1;min-width:0;">
          <div style="font-weight:600;color:#5b3df0;font-size:13px;margin-bottom:2px;">${adHeadline}</div>
          ${adBody ? `<div style="font-size:12px;color:#6b7280;line-height:1.3;margin-bottom:4px;">${adBody.substring(0,100)}</div>` : ""}
          ${adUrl ? `<a href="${adUrl}" target="_blank" style="font-size:11px;color:#5b3df0;text-decoration:none;word-break:break-all;">${adUrl.substring(0,50)}${adUrl.length > 50 ? "..." : ""}</a>` : ""}
        </div>
      </div>
    `;
  }

  // Render messages - handle interactive (with buttons) dan text biasa + date separator
  function formatDateSeparator(iso) {
    if (!iso) return "";
    let isoFixed = iso;
    if (!/Z|[+-]\d{2}:?\d{2}$/.test(iso)) {
      isoFixed = iso + "Z";
    }
    const d = new Date(isoFixed);
    if (isNaN(d.getTime())) return "";
    const today = new Date();
    const todayWib = new Date(today.toLocaleString("en-US", {timeZone: "Asia/Jakarta"}));
    const dWib = new Date(d.toLocaleString("en-US", {timeZone: "Asia/Jakarta"}));
    const dayDiff = Math.floor((new Date(todayWib.getFullYear(), todayWib.getMonth(), todayWib.getDate()) - new Date(dWib.getFullYear(), dWib.getMonth(), dWib.getDate())) / 86400000);
    if (dayDiff === 0) return "Hari ini";
    if (dayDiff === 1) return "Kemarin";
    if (dayDiff > 1 && dayDiff < 7) {
      const hari = ["Minggu","Senin","Selasa","Rabu","Kamis","Jumat","Sabtu"];
      return hari[dWib.getDay()];
    }
    const bulan = ["Januari","Februari","Maret","April","Mei","Juni","Juli","Agustus","September","Oktober","November","Desember"];
    return dWib.getDate() + " " + bulan[dWib.getMonth()] + " " + dWib.getFullYear();
  }
  function dateKey(iso) {
    if (!iso) return "";
    let isoFixed = iso;
    if (!/Z|[+-]\d{2}:?\d{2}$/.test(iso)) {
      isoFixed = iso + "Z";
    }
    const d = new Date(isoFixed);
    if (isNaN(d.getTime())) return "";
    const dWib = new Date(d.toLocaleString("en-US", {timeZone: "Asia/Jakarta"}));
    return dWib.getFullYear() + "-" + (dWib.getMonth()+1) + "-" + dWib.getDate();
  }
  let lastDateKey = "";
  const messagesHtml = messages.map(m => {
    const text = (m.text || "").replace(/</g,"&lt;").split(String.fromCharCode(10)).join("<br>");
    const buttons = (m.buttons && m.buttons.length) ? `
      <div style="margin-top:8px;display:flex;flex-direction:column;gap:6px;">
        ${m.buttons.map(b => `<div style="padding:8px 12px;border:1px solid #5b3df0;border-radius:20px;text-align:center;color:#5b3df0;font-size:13px;font-weight:500;">${b.replace(/</g,"&lt;")}</div>`).join("")}
      </div>
    ` : "";
    const currentKey = dateKey(m.timestamp);
    let separator = "";
    if (currentKey && currentKey !== lastDateKey) {
      separator = `<div class="chat-date-separator">${formatDateSeparator(m.timestamp)}</div>`;
      lastDateKey = currentKey;
    }
    return separator + `
      <div class="chat-bubble ${m.direction}">
        <div>${text}</div>
        ${buttons}
        <div class="chat-bubble-time">${formatTime(m.timestamp)}</div>
      </div>
    `;
  }).join("");

  msgEl.innerHTML = referralCardHtml + (messagesHtml || `<div class="chat-empty">Belum ada pesan</div>`);
  msgEl.scrollTop = msgEl.scrollHeight;

  // refresh list (unread sudah ke-reset di server)
  loadContacts();
}

function closeChatMobile() {
  const chatList = document.querySelector(".chat-list");
  const chatPanel = document.getElementById("chatPanel");
  if (chatList) chatList.classList.remove("hidden-mobile");
  if (chatPanel) chatPanel.classList.remove("show-mobile");
  currentPhone = null;
}

function toggleContactMenu(e) {
  e.stopPropagation();
  const dd = document.getElementById("contactMenuDropdown");
  dd.style.display = dd.style.display === "block" ? "none" : "block";
}

document.addEventListener("click", function(e) {
  const dd = document.getElementById("contactMenuDropdown");
  if (dd && !e.target.closest("#contactMenuBtn") && !e.target.closest("#contactMenuDropdown")) {
    dd.style.display = "none";
  }
});

async function resetMenuContact(phone) {
  document.getElementById("contactMenuDropdown").style.display = "none";
  if (!confirm("Reset Menu Otomatis untuk kontak ini? Menu Utama akan dikirim ulang saat pesan masuk berikutnya.")) return;
  try {
    const res = await fetch(`/api/inbox/${phone}/reset-menu`, {method:"POST"});
    const json = await res.json();
    if (json.status === "ok") {
      alert("✓ Menu Otomatis sudah direset untuk kontak ini.");
    } else {
      alert("Gagal: " + (json.error || "unknown"));
    }
  } catch (e) {
    alert("Error: " + e.message);
  }
}

async function markResolved(phone) {
  document.getElementById("contactMenuDropdown").style.display = "none";
  if (!confirm("Tandai percakapan ini sebagai Selesai?")) return;
  try {
    await fetch(`/api/inbox/${phone}/label`, {
      method: "POST",
      headers: {"Content-Type": "application/json"},
      body: JSON.stringify({status: "selesai"})
    });
    alert("✓ Percakapan ditandai Selesai.");
    loadContacts();
    currentPhone = null;
    document.getElementById("chatPanel").innerHTML = `<div class="chat-empty"><div style="font-size:40px;">💬</div><div>Pilih kontak di sebelah kiri<br>untuk memulai percakapan</div></div>`;
  } catch (e) {
    alert("Error: " + e.message);
  }
}

async function editContactLabel(phone) {
  document.getElementById("contactMenuDropdown").style.display = "none";
  const current = (allContacts.find(c => c.phone === phone) || {}).labels || [];

  // Fetch daftar label dari /api/labels
  let allLabels = [];
  try {
    const res = await fetch("/api/labels");
    const json = await res.json();
    allLabels = json.labels || [];
  } catch(e) {}

  // Buat selected set
  let selected = new Set(current);

  // Render modal
  const overlay = document.createElement("div");
  overlay.id = "labelModal";
  overlay.style.cssText = "position:fixed;inset:0;background:rgba(0,0,0,0.5);z-index:1000;display:flex;align-items:center;justify-content:center;";

  let searchQuery = "";

  function renderModal() {
    const selectedChipsHtml = [...selected].map(s => {
      const lab = allLabels.find(l => l.name === s);
      const bg = lab ? (lab.bg_color || "#5b3df0") : "#5b3df0";
      const color = lab ? (lab.text_color || "#fff") : "#fff";
      return `<span data-chip="${s.replace(/"/g,'&quot;')}" style="display:inline-flex;align-items:center;gap:4px;padding:4px 10px;border-radius:20px;background:${bg};color:${color};font-size:12px;font-weight:600;">
        ${s.replace(/</g,"&lt;")}
        <span onclick="removeLabelChip('${s.replace(/'/g,"\'")}'" style="cursor:pointer;font-size:14px;line-height:1;margin-left:2px;">×</span>
      </span>`;
    }).join("");

    const filtered = allLabels.filter(l =>
      !searchQuery || l.name.toLowerCase().includes(searchQuery.toLowerCase())
    );
    const listHtml = filtered.map(l => {
      const isSelected = selected.has(l.name);
      const bg = l.bg_color || "#5b3df0";
      const color = l.text_color || "#fff";
      return `<div onclick="toggleLabelChip('${l.name.replace(/'/g,"\'")}','${bg}','${color}')"
        style="display:flex;align-items:center;justify-content:space-between;padding:10px 14px;cursor:pointer;border-bottom:1px solid #f3f4f6;font-size:14px;transition:background .1s;"
        onmouseover="this.style.background='#f9fafb'" onmouseout="this.style.background='#fff'">
        <div style="display:flex;align-items:center;gap:10px;">
          <span style="width:12px;height:12px;border-radius:50%;background:${bg};flex-shrink:0;"></span>
          <span>${l.name.replace(/</g,"&lt;")}</span>
        </div>
        ${isSelected ? `<span style="color:#5b3df0;font-weight:700;font-size:16px;">✓</span>` : ""}
      </div>`;
    }).join("") || `<div style="padding:16px;text-align:center;color:#9ca3af;font-size:13px;">Tidak ada label. Buat di WhatsApp → Label Kontak.</div>`;

    overlay.innerHTML = `
      <div style="background:#fff;border-radius:16px;width:480px;max-width:90vw;max-height:90vh;display:flex;flex-direction:column;overflow:hidden;box-shadow:0 8px 32px rgba(0,0,0,.18);">
        <div style="background:#5b3df0;padding:16px 20px;display:flex;align-items:center;justify-content:space-between;flex-shrink:0;">
          <span style="color:#fff;font-weight:700;font-size:16px;">🏷 Contact Label</span>
          <span onclick="closeLabelModal()" style="color:#fff;cursor:pointer;font-size:22px;line-height:1;">✕</span>
        </div>
        <div style="padding:14px 16px;border-bottom:1px solid #f3f4f6;flex-shrink:0;">
          <div style="border:1.5px solid #5b3df0;border-radius:8px;padding:8px 10px;display:flex;flex-wrap:wrap;gap:6px;align-items:center;min-height:42px;cursor:text;" onclick="document.getElementById('labelSearch').focus()">
            ${selectedChipsHtml}
            <input id="labelSearch" placeholder="${selected.size ? '' : 'Pilih Contact Label'}" value="${searchQuery}"
              oninput="window._labelSearch(this.value)"
              style="border:none;outline:none;font-size:13px;flex:1;min-width:80px;background:transparent;color:#374151;">
          </div>
        </div>
        <div style="overflow-y:auto;height:250px;border-top:1px solid #f3f4f6;">
          ${listHtml}
        </div>
        <div style="padding:14px 16px;border-top:1px solid #f3f4f6;flex-shrink:0;">
          <button onclick="saveLabelModal('${phone}')" style="width:100%;padding:12px;background:#5b3df0;color:#fff;border:none;border-radius:10px;cursor:pointer;font-size:15px;font-weight:600;">Simpan</button>
        </div>
      </div>
    `;
    const inp = document.getElementById("labelSearch");
    if (inp) inp.focus();
  }

  window._labelSelected = selected;
  window._labelAllLabels = allLabels;
  window._labelSearch = function(val) {
    searchQuery = val;
    renderModal();
  };
  window.toggleLabelChip = function(name, bg, color) {
    if (window._labelSelected.has(name)) {
      window._labelSelected.delete(name);
    } else {
      window._labelSelected.add(name);
    }
    searchQuery = "";
    renderModal();
  };
  window.removeLabelChip = function(name) {
    window._labelSelected.delete(name);
    renderModal();
  };
  window.closeLabelModal = function() {
    document.body.removeChild(overlay);
  };
  window.saveLabelModal = async function(ph) {
    const labels = [...window._labelSelected];
    try {
      await fetch("/api/inbox/" + ph + "/label", {
        method: "POST",
        headers: {"Content-Type": "application/json"},
        body: JSON.stringify({labels})
      });
      document.body.removeChild(overlay);
      loadContacts();
    } catch(e) {
      alert("Error: " + e.message);
    }
  };

  document.body.appendChild(overlay);
  renderModal();
}

async function blokirContact(phone) {
  document.getElementById("contactMenuDropdown").style.display = "none";
  if (!confirm("Blokir kontak ini? Pesan dari kontak ini akan diabaikan oleh sistem.")) return;
  try {
    await fetch(`/api/inbox/${phone}/label`, {
      method: "POST",
      headers: {"Content-Type": "application/json"},
      body: JSON.stringify({status: "blokir"})
    });
    alert("✓ Kontak diblokir.");
    loadContacts();
    currentPhone = null;
    document.getElementById("chatPanel").innerHTML = `<div class="chat-empty"><div style="font-size:40px;">💬</div><div>Pilih kontak di sebelah kiri<br>untuk memulai percakapan</div></div>`;
  } catch (e) {
    alert("Error: " + e.message);
  }
}

// ---- Shortcut Popup (trigger /) ----
let _allShortcuts = [];
(async function() {
  try {
    const res = await fetch("/api/shortcuts");
    const json = await res.json();
    _allShortcuts = json.shortcuts || [];
  } catch(e) {}
})();

function onChatInput(val) {
  const popup = document.getElementById("shortcutPopup");
  const list = document.getElementById("shortcutPopupList");
  if (!popup || !list) return;

  if (!val.startsWith("/")) {
    popup.classList.remove("open");
    return;
  }

  const query = val.toLowerCase();
  const filtered = _allShortcuts.filter(s =>
    s.shortcut.toLowerCase().startsWith(query) ||
    s.title.toLowerCase().includes(query.slice(1))
  );

  if (!filtered.length) {
    popup.classList.remove("open");
    return;
  }

  const typeIcon = {TEXT:"📝", IMAGE:"🖼", DOCUMENT:"📄", VIDEO:"🎥"};
  list.innerHTML = filtered.map(s => `
    <div class="shortcut-item" onclick="selectShortcut('${s.id}')">
      <div class="shortcut-item-key">${s.shortcut} &nbsp;<span style="font-weight:400;color:#9ca3af;">${typeIcon[s.content_type]||""} ${s.content_type}</span></div>
      <div class="shortcut-item-body">${s.file_name || s.isi || ""}</div>
    </div>
  `).join("");
  popup.classList.add("open");
}

function closeShortcutPopup() {
  const popup = document.getElementById("shortcutPopup");
  if (popup) popup.classList.remove("open");
}

async function selectShortcut(scId) {
  closeShortcutPopup();
  if (!currentPhone) return;

  const sc = _allShortcuts.find(s => s.id === scId);
  if (!sc) return;

  // Clear input
  const input = document.getElementById("chatInput");
  if (input) input.value = "";

  // Loading indicator
  const loadingEl = document.createElement("div");
  loadingEl.style.cssText = "position:fixed;bottom:80px;left:50%;transform:translateX(-50%);background:#5b3df0;color:#fff;padding:10px 20px;border-radius:20px;font-size:13px;z-index:999;";
  loadingEl.textContent = "Mengirim " + sc.title + "...";
  document.body.appendChild(loadingEl);

  try {
    const res = await fetch("/api/shortcuts/" + scId + "/send/" + currentPhone, {method:"POST"});
    const json = await res.json();
    document.body.removeChild(loadingEl);
    if (json.status === "sent") {
      openChat(currentPhone);
    } else {
      alert("Gagal kirim shortcut: " + (json.error || "unknown"));
    }
  } catch(e) {
    document.body.removeChild(loadingEl);
    alert("Error: " + e.message);
  }
}

// Tutup shortcut popup kalau klik di luar
document.addEventListener("click", function(e) {
  const popup = document.getElementById("shortcutPopup");
  if (popup && !popup.contains(e.target)) {
    popup.classList.remove("open");
  }
});

// ---- Template Chat Attachment ----
let _chatTemplates = [];
let _allWaTemplates = [];

async function reloadChatTemplates() {
  try {
    const [settRes, tmplRes] = await Promise.all([
      fetch("/api/settings"),
      fetch("/api/wa-templates")
    ]);
    const sett = await settRes.json();
    const tmpl = await tmplRes.json();
    const chatTemplateNames = sett.wa_chat_templates || [];
    _allWaTemplates = tmpl.templates || (tmpl.raw && tmpl.raw.data) || [];
    _chatTemplates = _allWaTemplates.filter(t => chatTemplateNames.includes(t.name) && t.status === "APPROVED");
    renderAttachTemplates();
  } catch(e) {}
}

function renderAttachTemplates() {
  const el = document.getElementById("attachTemplateItems");
  if (!el) return;
  if (!_chatTemplates.length) { el.innerHTML = ""; return; }

  const subItems = _chatTemplates.map(t => `
    <div class="attach-menu-item" onclick="sendTemplateAttachment('${t.name}','${t.language||"id"}');closeTemplateSub();"
      style="padding:10px 14px 10px 36px;font-size:13px;">
      <span style="margin-right:8px;">📨</span><span>${t.name.replace(/_/g," ")}</span>
    </div>
  `).join("");

  el.innerHTML = `
    <div class="attach-menu-item" onclick="toggleTemplateSub(event)" id="templateSubToggle" style="position:relative;">
      <span class="icon">📢</span><span>Template</span>
      <span style="margin-left:auto;font-size:12px;color:#9ca3af;" id="templateSubArrow">▶</span>
    </div>
    <div id="templateSubMenu" style="display:none;background:#f9fafb;border-bottom:1px solid #f3f4f6;">
      ${subItems}
    </div>
    <div style="border-top:1px solid #f3f4f6;margin:0;"></div>
  `;
}

function toggleTemplateSub(e) {
  e.stopPropagation();
  const sub = document.getElementById("templateSubMenu");
  const arrow = document.getElementById("templateSubArrow");
  if (!sub) return;
  const isOpen = sub.style.display !== "none";
  sub.style.display = isOpen ? "none" : "block";
  if (arrow) arrow.textContent = isOpen ? "▶" : "▼";
}

function closeTemplateSub() {
  const sub = document.getElementById("templateSubMenu");
  if (sub) sub.style.display = "none";
  const arrow = document.getElementById("templateSubArrow");
  if (arrow) arrow.textContent = "▶";
}

async function sendTemplateAttachment(templateName, language) {
  document.getElementById("attachMenu").classList.remove("open");
  if (!currentPhone) return;

  const loadingEl = document.createElement("div");
  loadingEl.style.cssText = "position:fixed;bottom:80px;left:50%;transform:translateX(-50%);background:#5b3df0;color:#fff;padding:10px 20px;border-radius:20px;font-size:13px;z-index:999;";
  loadingEl.textContent = "Mengirim template " + templateName + "...";
  document.body.appendChild(loadingEl);

  try {
    const res = await fetch("/api/wa-templates/send/" + currentPhone, {
      method: "POST",
      headers: {"Content-Type":"application/json"},
      body: JSON.stringify({name: templateName, language})
    });
    const json = await res.json();
    document.body.removeChild(loadingEl);
    if (json.status === "sent") {
      openChat(currentPhone);
    } else {
      alert("Gagal kirim template: " + JSON.stringify(json.wa_body || json.error));
    }
  } catch(e) {
    document.body.removeChild(loadingEl);
    alert("Error: " + e.message);
  }
}

// Load saat halaman chat dibuka
reloadChatTemplates();

// ---- iOS/Chrome mobile keyboard handler ----
// Saat keyboard muncul, visual viewport mengecil
// Kita set height chat-panel dan scroll messages ke bawah
if (window.visualViewport) {
  function onViewportResize() {
    if (window.innerWidth > 768) return;
    const panel = document.getElementById("chatPanel");
    const msgs = document.getElementById("chatMessages");
    if (!panel) return;
    const vh = window.visualViewport.height;
    panel.style.height = vh + "px";
    panel.style.top = window.visualViewport.offsetTop + "px";
    if (msgs) {
      setTimeout(() => { msgs.scrollTop = msgs.scrollHeight; }, 50);
    }
  }
  window.visualViewport.addEventListener("resize", onViewportResize);
  window.visualViewport.addEventListener("scroll", onViewportResize);
}

function toggleAttachMenu(e) {
  e.stopPropagation();
  const menu = document.getElementById("attachMenu");
  renderAttachTemplates();
  menu.classList.toggle("open");
}
document.addEventListener("click", function() {
  const menu = document.getElementById("attachMenu");
  if (menu) menu.classList.remove("open");
});

async function sendFlowAttachment() {
  document.getElementById("attachMenu").classList.remove("open");
  if (!currentPhone) return;
  try {
    const res = await fetch("/api/send-flow/" + currentPhone, {method:"POST"});
    const json = await res.json();
    if (json.status === "sent") {
      // Refresh pesan
      openChat(currentPhone);
    } else {
      alert("Gagal kirim Flow: " + (json.error || "unknown"));
    }
  } catch(e) {
    alert("Error: " + e.message);
  }
}

async function sendMediaAttachment(mediaType) {
  document.getElementById("attachMenu").classList.remove("open");
  if (!currentPhone) return;

  // Label & accept per jenis
  const cfg = {
    document: {label:"Pilih Dokumen", accept:".pdf,.doc,.docx,.xls,.xlsx,.ppt,.pptx,.txt,.zip"},
    image:    {label:"Pilih Gambar/Video", accept:"image/*,video/*"},
    audio:    {label:"Pilih Audio", accept:"audio/*,.mp3,.ogg,.m4a,.aac"},
  };
  const {label, accept} = cfg[mediaType] || cfg.document;

  // Buat file input sementara
  const fileInput = document.createElement("input");
  fileInput.type = "file";
  fileInput.accept = accept;
  fileInput.style.display = "none";
  document.body.appendChild(fileInput);

  fileInput.onchange = async function() {
    const file = fileInput.files[0];
    document.body.removeChild(fileInput);
    if (!file) return;

    // Upload ke /api/send-media/<phone>
    const formData = new FormData();
    formData.append("file", file);
    formData.append("media_type", mediaType);

    try {
      const loadingEl = document.createElement("div");
      loadingEl.style.cssText = "position:fixed;bottom:80px;left:50%;transform:translateX(-50%);background:#5b3df0;color:#fff;padding:10px 20px;border-radius:20px;font-size:13px;z-index:999;";
      loadingEl.textContent = "Mengirim " + file.name + "...";
      document.body.appendChild(loadingEl);

      const res = await fetch("/api/send-media/" + currentPhone, {method:"POST", body:formData});
      const json = await res.json();
      document.body.removeChild(loadingEl);
      if (json.status === "sent") {
        openChat(currentPhone);
      } else {
        alert("Gagal kirim: " + (json.error || "unknown"));
      }
    } catch(e) {
      alert("Error upload: " + e.message);
    }
  };
  fileInput.click();
}

async function sendReply() {
  const input = document.getElementById("chatInput");
  const text = input.value.trim();
  if (!text || !currentPhone) return;
  input.value = "";
  input.disabled = true;
  try {
    await fetch(`/api/inbox/${currentPhone}/reply`, {
      method: "POST",
      headers: {"Content-Type": "application/json"},
      body: JSON.stringify({text})
    });
    openChat(currentPhone);
  } finally {
    input.disabled = false;
    input.focus();
  }
}

loadContacts();
setInterval(() => {
  loadContacts();
  if (currentPhone) openChat(currentPhone);
}, 15000);
</script>
"""
    return Response(render_page("chat", "", "", body), mimetype="text/html")


@app.route("/whatsapp", methods=["GET"])
@login_required
def whatsapp_page():
    """Halaman Pengaturan: Menu Utama (auto-reply) dan WA Blast - sistem Raihmimpi sendiri."""
    body = """
  <div class="settings-wrap">
  <div class="settings-tabs">
    <div class="settings-tabs-group">Pengaturan Pesan</div>
    <div class="settings-tab active" data-tab="menu-utama" onclick="setSettingsTab(this)">
      <span class="tab-icon">💬</span> Menu Utama
    </div>

    <div class="settings-tabs-group">WhatsApp Blast</div>
    <div class="settings-tab" data-tab="wa-blast" onclick="setSettingsTab(this)">
      <span class="tab-icon">📢</span> WA Blast
    </div>
    <div class="settings-tab" data-tab="wa-template" onclick="setSettingsTab(this)">
      <span class="tab-icon">📋</span> Template Blast
    </div>

    <div class="settings-tabs-group">Kontak & Label</div>
    <div class="settings-tab" data-tab="label-kontak" onclick="setSettingsTab(this)">
      <span class="tab-icon">🏷</span> Label Kontak
    </div>

    <div class="settings-tabs-group">Pesan Cepat</div>
    <div class="settings-tab" data-tab="shortcuts" onclick="setSettingsTab(this)">
      <span class="tab-icon">⚡</span> Shortcuts Pesan
    </div>
  </div>
  <div class="settings-content">

  <div class="settings-section active" id="section-menu-utama">
    <div class="panel">
      <h2>Menu Utama</h2>
      <p class="form-hint" style="margin-bottom:16px;">Pesan sapaan otomatis beserta tombol aksi yang dikirim ke kontak baru.</p>

      <div class="form-group">
        <label>Isi Pesan</label>
        <textarea id="menuMessage" maxlength="1024" placeholder="Halo! kak, ..."></textarea>
        <div class="form-hint"><span id="menuMessageCount">0</span>/1024</div>
      </div>

      <div class="form-group">
        <label>Tombol</label>
        <div id="menuButtons"></div>
      </div>

      <button class="btn" onclick="saveMenuUtama()">Simpan</button>
      <span class="save-msg" id="menuSaveMsg"></span>
    </div>
  </div>

  <div class="settings-section" id="section-wa-blast">
    <!-- VIEW: Daftar Blast -->
    <div id="blastListView">
      <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:20px;">
        <h2 style="margin:0;">Daftar WhatsApp Blast</h2>
        <button class="btn" onclick="showBlastForm()">+ Buat Blast</button>
      </div>

      <!-- Filter -->
      <div style="background:#fff;border-radius:12px;padding:16px 20px;margin-bottom:16px;box-shadow:0 1px 3px rgba(0,0,0,.06);">
        <div style="display:flex;gap:12px;align-items:flex-end;flex-wrap:wrap;">
          <div>
            <div style="font-size:12px;font-weight:600;color:#6b7280;margin-bottom:6px;">Filter Tanggal</div>
            <div style="display:flex;gap:8px;align-items:center;">
              <input type="date" id="blastFilterFrom" style="padding:8px 12px;border:1px solid #d1d5db;border-radius:8px;font-size:13px;">
              <span style="color:#9ca3af;">→</span>
              <input type="date" id="blastFilterTo" style="padding:8px 12px;border:1px solid #d1d5db;border-radius:8px;font-size:13px;">
            </div>
          </div>
          <button onclick="loadBlastHistory()" style="padding:8px 20px;background:#5b3df0;color:#fff;border:none;border-radius:8px;font-size:13px;font-weight:600;cursor:pointer;">Cari</button>
        </div>
      </div>

      <!-- Tabel -->
      <div style="background:#fff;border-radius:12px;overflow:hidden;box-shadow:0 1px 3px rgba(0,0,0,.06);">
        <table style="width:100%;border-collapse:collapse;">
          <thead>
            <tr style="background:#f9fafb;">
              <th style="padding:12px 16px;text-align:left;font-size:12px;color:#6b7280;font-weight:600;border-bottom:1px solid #e5e7eb;white-space:nowrap;">Judul Campaign</th>
              <th style="padding:12px 16px;text-align:left;font-size:12px;color:#6b7280;font-weight:600;border-bottom:1px solid #e5e7eb;white-space:nowrap;">Tanggal Blast</th>
              <th style="padding:12px 16px;text-align:left;font-size:12px;color:#6b7280;font-weight:600;border-bottom:1px solid #e5e7eb;white-space:nowrap;">Nama Template</th>
              <th style="padding:12px 16px;text-align:left;font-size:12px;color:#6b7280;font-weight:600;border-bottom:1px solid #e5e7eb;white-space:nowrap;">Kategori</th>
              <th style="padding:12px 16px;text-align:left;font-size:12px;color:#6b7280;font-weight:600;border-bottom:1px solid #e5e7eb;white-space:nowrap;">Status</th>
              <th style="padding:12px 16px;text-align:center;font-size:12px;color:#6b7280;font-weight:600;border-bottom:1px solid #e5e7eb;">Total</th>
              <th style="padding:12px 16px;text-align:center;font-size:12px;color:#6b7280;font-weight:600;border-bottom:1px solid #e5e7eb;">Sent</th>
              <th style="padding:12px 16px;text-align:center;font-size:12px;color:#6b7280;font-weight:600;border-bottom:1px solid #e5e7eb;">Delivered</th>
              <th style="padding:12px 16px;text-align:center;font-size:12px;color:#6b7280;font-weight:600;border-bottom:1px solid #e5e7eb;">Read</th>
              <th style="padding:12px 16px;text-align:center;font-size:12px;color:#6b7280;font-weight:600;border-bottom:1px solid #e5e7eb;">Failed</th>
              <th style="padding:12px 16px;text-align:right;font-size:12px;color:#6b7280;font-weight:600;border-bottom:1px solid #e5e7eb;">Aksi</th>
            </tr>
          </thead>
          <tbody id="blastHistoryBody">
            <tr><td colspan="11" style="padding:40px;text-align:center;color:#9ca3af;">Memuat...</td></tr>
          </tbody>
        </table>
      </div>
    </div>

    <!-- VIEW: Form Buat Blast -->
    <div id="blastFormView" style="display:none;">
      <div style="display:flex;align-items:center;gap:12px;margin-bottom:20px;">
        <button onclick="showBlastList()" style="background:none;border:none;cursor:pointer;color:#5b3df0;font-size:20px;">←</button>
        <h2 style="margin:0;">Buat Blast</h2>
      </div>

      <div style="background:#fff;border-radius:12px;padding:24px;box-shadow:0 1px 3px rgba(0,0,0,.06);">

        <div style="display:grid;grid-template-columns:1fr 1fr;gap:16px;">

          <div class="form-group">
            <label>Kategori <span style="color:#dc2626;">*</span></label>
            <select id="blastKategori" style="width:100%;padding:10px 12px;border:1px solid #d1d5db;border-radius:8px;font-size:14px;">
              <option value="MARKETING">MARKETING</option>
              <option value="UTILITY">UTILITY</option>
              <option value="AUTHENTICATION">AUTHENTICATION</option>
            </select>
          </div>

          <div class="form-group">
            <label>Nama Template Blast <span style="color:#dc2626;">*</span></label>
            <select id="blastTemplate" onchange="onBlastTemplateChange()" style="width:100%;padding:10px 12px;border:1px solid #d1d5db;border-radius:8px;font-size:14px;">
              <option value="">-- Pilih Template --</option>
            </select>
          </div>

          <div class="form-group">
            <label>Judul <span style="color:#dc2626;">*</span></label>
            <input type="text" id="blastJudul" placeholder="Judul blast..." style="width:100%;padding:10px 12px;border:1px solid #d1d5db;border-radius:8px;font-size:14px;box-sizing:border-box;">
          </div>

          <div class="form-group">
            <label>Judul Campaign</label>
            <input type="text" id="blastJudulCampaign" placeholder="Nama campaign..." style="width:100%;padding:10px 12px;border:1px solid #d1d5db;border-radius:8px;font-size:14px;box-sizing:border-box;">
          </div>

          <div class="form-group">
            <label>Header Type</label>
            <select id="blastHeaderType" style="width:100%;padding:10px 12px;border:1px solid #d1d5db;border-radius:8px;font-size:14px;">
              <option value="">-- Pilih Header Type --</option>
              <option value="TEXT">TEXT</option>
              <option value="IMAGE">IMAGE</option>
              <option value="VIDEO">VIDEO</option>
              <option value="DOCUMENT">DOCUMENT</option>
              <option value="CAROUSEL">CAROUSEL</option>
              <option value="CAROUSEL_PRODUCT">CAROUSEL_PRODUCT</option>
            </select>
          </div>

          <div class="form-group">
            <label>Label Contact</label>
            <select id="blastLabelFilter" style="width:100%;padding:10px 12px;border:1px solid #d1d5db;border-radius:8px;font-size:14px;">
              <option value="">-- Semua Kontak --</option>
            </select>
          </div>

        </div>

        <!-- Preview template body -->
        <div id="blastTemplatePreview" style="display:none;background:#f0edff;border-radius:8px;padding:14px;margin-bottom:16px;font-size:13px;color:#374151;border-left:3px solid #5b3df0;">
          <div style="font-size:11px;font-weight:700;color:#5b3df0;margin-bottom:6px;">PREVIEW TEMPLATE</div>
          <div id="blastTemplatePreviewBody"></div>
        </div>

        <!-- Parameter body jika ada {{1}} -->
        <div id="blastParamsGroup" style="display:none;margin-bottom:16px;">
          <label style="display:block;font-size:13px;font-weight:600;margin-bottom:8px;">Parameter Body</label>
          <div id="blastParamsList"></div>
        </div>

        <div style="display:grid;grid-template-columns:1fr 1fr;gap:16px;margin-bottom:16px;">
          <div class="form-group" style="margin:0;">
            <label>Atur Jadwal Blast <span style="color:#dc2626;">*</span></label>
            <div style="display:flex;gap:16px;margin-top:8px;">
              <label style="display:flex;align-items:center;gap:6px;cursor:pointer;font-size:14px;">
                <input type="radio" name="blastSchedule" value="now" checked onchange="onScheduleChange(this)"> Kirim Sekarang
              </label>
              <label style="display:flex;align-items:center;gap:6px;cursor:pointer;font-size:14px;">
                <input type="radio" name="blastSchedule" value="schedule" onchange="onScheduleChange(this)"> Atur Jadwal
              </label>
            </div>
          </div>
          <div id="blastScheduleInput" style="display:none;">
            <label style="display:block;font-size:13px;font-weight:600;margin-bottom:6px;">Waktu Blast</label>
            <input type="datetime-local" id="blastScheduleTime" style="width:100%;padding:10px 12px;border:1px solid #d1d5db;border-radius:8px;font-size:14px;box-sizing:border-box;">
          </div>
        </div>

        <!-- Pilih Kontak -->
        <div style="margin-bottom:16px;">
          <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:8px;">
            <label style="font-size:13px;font-weight:600;">Pilih Kontak <span style="color:#dc2626;">*</span></label>
            <div style="display:flex;gap:8px;">
              <button onclick="toggleAllBlast()" style="padding:5px 12px;border:1px solid #d1d5db;border-radius:6px;font-size:12px;cursor:pointer;background:#fff;">Pilih Semua</button>
              <span id="blastCount" style="font-size:12px;color:#6b7280;align-self:center;">0 dipilih</span>
            </div>
          </div>
          <div style="margin-bottom:8px;">
            <input type="text" id="blastContactSearch" placeholder="Cari nama atau nomor..." oninput="filterBlastContacts(this.value)"
              style="width:100%;padding:8px 12px;border:1px solid #d1d5db;border-radius:8px;font-size:13px;box-sizing:border-box;">
          </div>
          <div class="blast-contacts" id="blastContacts" style="max-height:200px;overflow-y:auto;border:1px solid #e5e7eb;border-radius:8px;padding:8px;">
            Memuat kontak...
          </div>
        </div>

        <!-- Upload Kontak dari Excel -->
        <div style="border:1px solid #e5e7eb;border-radius:8px;padding:16px;margin-bottom:16px;background:#f9fafb;">
          <div style="font-size:13px;font-weight:700;margin-bottom:10px;color:#374151;">📥 Upload Nomor dari File Excel</div>
          <div style="display:flex;align-items:center;gap:12px;flex-wrap:wrap;">
            <button onclick="downloadBlastTemplate()" style="padding:8px 16px;border:1px solid #5b3df0;border-radius:8px;color:#5b3df0;background:#fff;cursor:pointer;font-size:13px;font-weight:600;">
              ⬇ Download Template
            </button>
            <span style="font-size:12px;color:#9ca3af;">wablast_template.xlsx &nbsp;*Lakukan download template terlebih dahulu</span>
          </div>
          <div style="margin-top:10px;">
            <label style="display:block;font-size:12px;font-weight:600;color:#6b7280;margin-bottom:6px;">File Upload Blast (.xlsx)</label>
            <input type="file" id="blastFileUpload" accept=".xlsx,.xls,.csv" onchange="onBlastFileUpload(this)"
              style="width:100%;padding:8px;border:1px solid #d1d5db;border-radius:8px;font-size:13px;background:#fff;">
            <div id="blastFileMsg" style="font-size:12px;color:#16a34a;margin-top:4px;"></div>
          </div>
        </div>

        <div style="background:#fffbeb;border:1px solid #fcd34d;border-radius:8px;padding:12px;font-size:12px;color:#92400e;margin-bottom:20px;">
          ⚠️ Jumlah maksimal karakter body adalah 1024 karakter, sudah termasuk parameter yang anda input.
        </div>

        <div style="display:flex;justify-content:flex-end;gap:10px;">
          <button onclick="showBlastList()" style="padding:10px 20px;border:1px solid #e5e7eb;border-radius:8px;background:#fff;color:#374151;cursor:pointer;font-weight:600;">Batal</button>
          <button onclick="sendBlastNew()" style="padding:10px 28px;background:#5b3df0;color:#fff;border:none;border-radius:8px;cursor:pointer;font-weight:700;font-size:14px;">Kirim WhatsApp</button>
        </div>
        <div id="blastSendMsg" style="margin-top:12px;font-size:13px;text-align:right;"></div>
      </div>
    </div>
  </div>

  <div class="settings-section" id="section-wa-template">
    <div class="panel">
      <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:20px;">
        <div>
          <h2 style="margin:0;">Template Blast</h2>
          <p style="margin:4px 0 0;font-size:13px;color:#6b7280;">Template pesan untuk WA Blast. Perlu approval Meta (1-24 jam).</p>
        </div>
        <button class="btn" onclick="showTemplateForm()">+ Tambah Template</button>
      </div>

      <!-- Table -->
      <table style="width:100%;border-collapse:collapse;background:#fff;border-radius:8px;overflow:hidden;">
        <thead>
          <tr style="background:#f9fafb;">
            <th style="padding:12px 16px;text-align:left;font-size:13px;color:#6b7280;font-weight:600;border-bottom:1px solid #e5e7eb;">Nama Template</th>
            <th style="padding:12px 16px;text-align:left;font-size:13px;color:#6b7280;font-weight:600;border-bottom:1px solid #e5e7eb;">Kategori</th>
            <th style="padding:12px 16px;text-align:left;font-size:13px;color:#6b7280;font-weight:600;border-bottom:1px solid #e5e7eb;">Bahasa</th>
            <th style="padding:12px 16px;text-align:left;font-size:13px;color:#6b7280;font-weight:600;border-bottom:1px solid #e5e7eb;">Status</th>
            <th style="padding:12px 16px;text-align:left;font-size:13px;color:#6b7280;font-weight:600;border-bottom:1px solid #e5e7eb;">Isi</th>
            <th style="padding:12px 16px;text-align:right;font-size:13px;color:#6b7280;font-weight:600;border-bottom:1px solid #e5e7eb;">Aksi</th>
          </tr>
        </thead>
        <tbody id="templateTableBody">
          <tr><td colspan="6" style="padding:40px;text-align:center;color:#9ca3af;">Memuat...</td></tr>
        </tbody>
      </table>
    </div>

    <!-- Modal Form Template -->
    <div id="templateFormModal" style="display:none;position:fixed;inset:0;background:rgba(0,0,0,.5);z-index:1000;align-items:flex-start;justify-content:center;overflow-y:auto;padding:40px 0;">
      <div style="background:#fff;border-radius:16px;width:680px;max-width:95vw;margin:auto;">
        <div style="background:#5b3df0;padding:16px 24px;display:flex;justify-content:space-between;align-items:center;border-radius:16px 16px 0 0;">
          <span style="color:#fff;font-weight:700;font-size:16px;">Broadcast Template</span>
          <span onclick="closeTemplateForm()" style="color:#fff;cursor:pointer;font-size:22px;">✕</span>
        </div>
        <div style="display:flex;gap:0;">
          <!-- Form -->
          <div style="flex:1;padding:24px;border-right:1px solid #f3f4f6;">
            <div style="margin-bottom:16px;">
              <label style="display:block;font-size:13px;font-weight:600;margin-bottom:6px;">Kategori</label>
              <select id="tmplCategory" style="width:100%;padding:10px 12px;border:1px solid #d1d5db;border-radius:8px;font-size:14px;" onchange="updateTemplatePreview()">
                <option value="MARKETING">MARKETING</option>
                <option value="UTILITY">UTILITY</option>
                <option value="AUTHENTICATION">AUTHENTICATION</option>
              </select>
            </div>
            <div style="margin-bottom:16px;">
              <label style="display:block;font-size:13px;font-weight:600;margin-bottom:6px;">Nama Template <span style="color:#dc2626;">*</span></label>
              <input type="text" id="tmplName" placeholder="contoh: donasi_via_wa (huruf kecil, underscore)" style="width:100%;padding:10px 12px;border:1px solid #d1d5db;border-radius:8px;font-size:14px;box-sizing:border-box;" oninput="this.value=this.value.toLowerCase().replace(/[^a-z0-9_]/g,'')">
              <div style="font-size:11px;color:#9ca3af;margin-top:4px;"><span id="tmplNameCount">0</span>/60</div>
            </div>
            <div style="margin-bottom:16px;">
              <label style="display:block;font-size:13px;font-weight:600;margin-bottom:6px;">Format Judul</label>
              <select id="tmplHeaderType" onchange="onHeaderTypeChange()" style="width:100%;padding:10px 12px;border:1px solid #d1d5db;border-radius:8px;font-size:14px;">
                <option value="">-- Pilih Tipe Judul --</option>
                <option value="TEXT">TEXT</option>
                <option value="IMAGE">IMAGE</option>
                <option value="VIDEO">VIDEO</option>
                <option value="DOCUMENT">DOCUMENT</option>
              </select>
            </div>
            <div id="tmplHeaderTextGroup" style="display:none;margin-bottom:16px;">
              <label style="display:block;font-size:13px;font-weight:600;margin-bottom:6px;" id="tmplHeaderLabel">Teks Judul</label>
              <input type="text" id="tmplHeaderText" placeholder="Judul template..." style="width:100%;padding:10px 12px;border:1px solid #d1d5db;border-radius:8px;font-size:14px;box-sizing:border-box;" oninput="updateTemplatePreview()">
              <div id="tmplHeaderFileGroup" style="display:none;margin-top:8px;">
                <input type="file" id="tmplHeaderFile" style="width:100%;padding:8px;border:1px solid #d1d5db;border-radius:8px;font-size:13px;">
                <div style="font-size:11px;color:#9ca3af;margin-top:4px;">File akan diupload saat template digunakan untuk blast</div>
              </div>
            </div>
            <div style="margin-bottom:16px;">
              <label style="display:block;font-size:13px;font-weight:600;margin-bottom:6px;">Isi <span style="color:#dc2626;">*</span></label>
              <div style="display:flex;gap:6px;margin-bottom:6px;flex-wrap:wrap;">
                <button type="button" onclick="insertParam()" style="padding:4px 10px;border:1px solid #d1d5db;border-radius:6px;font-size:12px;cursor:pointer;background:#f9fafb;">+ Parameter</button>
                <button type="button" onclick="wrapText('*')" style="padding:4px 10px;border:1px solid #d1d5db;border-radius:6px;font-size:12px;cursor:pointer;background:#f9fafb;font-weight:700;">Bold</button>
                <button type="button" onclick="wrapText('_')" style="padding:4px 10px;border:1px solid #d1d5db;border-radius:6px;font-size:12px;cursor:pointer;background:#f9fafb;font-style:italic;">Italic</button>
                <button type="button" onclick="wrapText('~')" style="padding:4px 10px;border:1px solid #d1d5db;border-radius:6px;font-size:12px;cursor:pointer;background:#f9fafb;text-decoration:line-through;">Strike</button>
                <div style="position:relative;display:inline-block;">
                  <button type="button" onclick="toggleEmojiPicker()" style="padding:4px 10px;border:1px solid #d1d5db;border-radius:6px;font-size:14px;cursor:pointer;background:#f9fafb;" title="Emoji">😀</button>
                  <div id="emojiPicker" style="display:none;position:absolute;top:32px;left:0;background:#fff;border:1px solid #e5e7eb;border-radius:10px;box-shadow:0 4px 16px rgba(0,0,0,.12);padding:10px;z-index:100;width:280px;max-height:200px;overflow-y:auto;">
                    <div style="display:flex;flex-wrap:wrap;gap:4px;" id="emojiGrid"></div>
                  </div>
                </div>
              </div>
              <textarea id="tmplBody" rows="5" placeholder="Isi pesan template..." maxlength="1024" style="width:100%;padding:10px 12px;border:1px solid #d1d5db;border-radius:8px;font-size:14px;box-sizing:border-box;resize:vertical;" oninput="updateTemplatePreview();document.getElementById('tmplBodyCount').textContent=this.value.length"></textarea>
              <div style="font-size:11px;color:#9ca3af;margin-top:4px;"><span id="tmplBodyCount">0</span>/1024</div>
            </div>
            <div style="margin-bottom:16px;">
              <label style="display:block;font-size:13px;font-weight:600;margin-bottom:6px;">Catatan Kaki</label>
              <textarea id="tmplFooter" rows="2" placeholder="Catatan kaki (opsional)..." maxlength="60" style="width:100%;padding:10px 12px;border:1px solid #d1d5db;border-radius:8px;font-size:14px;box-sizing:border-box;resize:vertical;" oninput="updateTemplatePreview();document.getElementById('tmplFooterCount').textContent=this.value.length"></textarea>
              <div style="font-size:11px;color:#9ca3af;margin-top:4px;"><span id="tmplFooterCount">0</span>/60</div>
            </div>
            <div style="margin-bottom:16px;">
              <label style="display:block;font-size:13px;font-weight:600;margin-bottom:6px;">Bahasa <span style="color:#dc2626;">*</span></label>
              <select id="tmplLanguage" style="width:100%;padding:10px 12px;border:1px solid #d1d5db;border-radius:8px;font-size:14px;">
                <option value="">-- Pilih Bahasa --</option>
                <option value="id">Indonesia (id)</option>
                <option value="en_US">English US (en_US)</option>
                <option value="en">English (en)</option>
              </select>
            </div>
            <div style="margin-bottom:16px;">
              <label style="display:block;font-size:13px;font-weight:600;margin-bottom:6px;">Tipe Tombol</label>
              <select id="tmplBtnType" onchange="onBtnTypeChange()" style="width:100%;padding:10px 12px;border:1px solid #d1d5db;border-radius:8px;font-size:14px;">
                <option value="">Tidak Ada</option>
                <option value="QUICK_REPLY">QUICK_REPLY</option>
                <option value="CALL_TO_ACTION">CALL_TO_ACTION</option>
              </select>
            </div>
            <div id="tmplBtnGroup" style="display:none;margin-bottom:16px;">
              <div id="tmplBtnList"></div>
              <button type="button" onclick="addTemplateButton()" style="margin-top:8px;padding:8px 14px;border:1px dashed #5b3df0;border-radius:8px;color:#5b3df0;background:#f5f3ff;cursor:pointer;font-size:13px;width:100%;">+ Tambah Tombol</button>
            </div>
            <div style="background:#fff3cd;border:1px solid #ffc107;border-radius:8px;padding:12px;font-size:12px;color:#856404;margin-bottom:16px;">
              ⚠️ Jumlah maksimal karakter body adalah 1024 karakter, sudah termasuk parameter yang anda input.
            </div>
          </div>
          <!-- Preview -->
          <div style="width:260px;padding:16px;background:#f0ede8;flex-shrink:0;background-image:url('data:image/svg+xml,<svg xmlns=%22http://www.w3.org/2000/svg%22 width=%2260%22 height=%2260%22><rect width=%2260%22 height=%2260%22 fill=%22%23e8e0d8%22/><circle cx=%2230%22 cy=%2230%22 r=%2218%22 fill=%22none%22 stroke=%22%23d4ccc4%22 stroke-width=%221%22/></svg>');">
            <div style="font-size:13px;font-weight:700;color:#374151;margin-bottom:12px;">Pratinjau</div>
            <div style="max-width:200px;">
              <div id="tmplPreview" style="background:#fff;border-radius:0 10px 10px 10px;padding:10px 12px;font-size:13px;line-height:1.5;word-break:break-word;box-shadow:0 1px 3px rgba(0,0,0,.15);position:relative;">
                <div id="tmplPreviewHeaderImg" style="display:none;background:#ccc;border-radius:6px;height:100px;margin-bottom:8px;display:none;align-items:center;justify-content:center;color:#666;font-size:12px;">📎 Media</div>
                <div id="tmplPreviewHeader" style="font-weight:700;margin-bottom:6px;display:none;font-size:14px;"></div>
                <div id="tmplPreviewBody" style="color:#111;white-space:pre-wrap;"></div>
                <div id="tmplPreviewFooter" style="color:#999;font-size:11px;margin-top:6px;display:none;border-top:1px solid #f0f0f0;padding-top:4px;"></div>
                <div style="text-align:right;font-size:10px;color:#999;margin-top:4px;">12:00 ✓✓</div>
              </div>
              <div id="tmplPreviewBtns" style="margin-top:4px;display:flex;flex-direction:column;gap:4px;"></div>
            </div>
          </div>
        </div>
        <div style="padding:16px 24px;border-top:1px solid #f3f4f6;display:flex;gap:10px;justify-content:flex-end;">
          <button onclick="previewTemplate()" style="padding:10px 20px;background:#22c55e;color:#fff;border:none;border-radius:8px;cursor:pointer;font-weight:600;">Pratinjau</button>
          <button onclick="saveTemplateForm()" style="padding:10px 24px;background:#5b3df0;color:#fff;border:none;border-radius:8px;cursor:pointer;font-weight:600;">Simpan</button>
        </div>
      </div>
    </div>
  </div>

  <div class="settings-section" id="section-shortcuts">
    <div class="panel">
      <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:20px;">
        <div>
          <h2 style="margin:0;">Shortcuts Pesan</h2>
          <p style="margin:4px 0 0;font-size:13px;color:#6b7280;">Ketik / di chat untuk akses shortcut cepat.</p>
        </div>
        <button class="btn" onclick="showShortcutForm()">+ Tambah Shortcut</button>
      </div>

      <!-- Table -->
      <table style="width:100%;border-collapse:collapse;background:#fff;border-radius:8px;overflow:hidden;">
        <thead>
          <tr style="background:#f9fafb;">
            <th style="padding:12px 16px;text-align:left;font-size:13px;color:#6b7280;font-weight:600;border-bottom:1px solid #e5e7eb;">Judul</th>
            <th style="padding:12px 16px;text-align:left;font-size:13px;color:#6b7280;font-weight:600;border-bottom:1px solid #e5e7eb;">Content Type</th>
            <th style="padding:12px 16px;text-align:left;font-size:13px;color:#6b7280;font-weight:600;border-bottom:1px solid #e5e7eb;">Shortcut</th>
            <th style="padding:12px 16px;text-align:left;font-size:13px;color:#6b7280;font-weight:600;border-bottom:1px solid #e5e7eb;">Isi</th>
            <th style="padding:12px 16px;text-align:right;font-size:13px;color:#6b7280;font-weight:600;border-bottom:1px solid #e5e7eb;">Aksi</th>
          </tr>
        </thead>
        <tbody id="shortcutsTableBody">
          <tr><td colspan="5" style="padding:40px;text-align:center;color:#9ca3af;">Memuat...</td></tr>
        </tbody>
      </table>
    </div>

    <!-- Modal Form Shortcut -->
    <div id="shortcutFormModal" style="display:none;position:fixed;inset:0;background:rgba(0,0,0,.5);z-index:1000;align-items:center;justify-content:center;">
      <div style="background:#fff;border-radius:16px;width:560px;max-width:95vw;max-height:90vh;overflow-y:auto;">
        <div style="background:#5b3df0;padding:16px 20px;display:flex;justify-content:space-between;align-items:center;border-radius:16px 16px 0 0;">
          <span style="color:#fff;font-weight:700;font-size:16px;" id="shortcutFormTitle">Buat Shortcuts Pesan</span>
          <span onclick="closeShortcutForm()" style="color:#fff;cursor:pointer;font-size:22px;">✕</span>
        </div>
        <div style="padding:24px;">
          <input type="hidden" id="scFormId">
          <div style="margin-bottom:16px;">
            <label style="display:block;font-size:13px;font-weight:600;margin-bottom:6px;">Channel</label>
            <select id="scFormChannel" style="width:100%;padding:10px 12px;border:1px solid #d1d5db;border-radius:8px;font-size:14px;">
              <option value="WHATSAPP">WHATSAPP</option>
            </select>
          </div>
          <div style="margin-bottom:16px;">
            <label style="display:block;font-size:13px;font-weight:600;margin-bottom:6px;">Content Type</label>
            <select id="scFormType" onchange="onScTypeChange()" style="width:100%;padding:10px 12px;border:1px solid #d1d5db;border-radius:8px;font-size:14px;">
              <option value="">-- Pilih Tipe --</option>
              <option value="TEXT">TEXT</option>
              <option value="IMAGE">IMAGE</option>
              <option value="DOCUMENT">DOCUMENT</option>
              <option value="VIDEO">VIDEO</option>
            </select>
          </div>
          <div style="margin-bottom:16px;">
            <label style="display:block;font-size:13px;font-weight:600;margin-bottom:6px;">Shortcut <span style="color:#dc2626;">*</span></label>
            <input type="text" id="scFormShortcut" placeholder="/contoh-shortcut" style="width:100%;padding:10px 12px;border:1px solid #d1d5db;border-radius:8px;font-size:14px;box-sizing:border-box;">
          </div>
          <div style="margin-bottom:16px;">
            <label style="display:block;font-size:13px;font-weight:600;margin-bottom:6px;">Judul <span style="color:#dc2626;">*</span></label>
            <input type="text" id="scFormTitle" placeholder="Nama shortcut..." style="width:100%;padding:10px 12px;border:1px solid #d1d5db;border-radius:8px;font-size:14px;box-sizing:border-box;">
          </div>
          <div style="margin-bottom:16px;" id="scIsiGroup">
            <label style="display:block;font-size:13px;font-weight:600;margin-bottom:6px;">Isi</label>
            <textarea id="scFormIsi" rows="5" placeholder="Tulis isi pesan..." style="width:100%;padding:10px 12px;border:1px solid #d1d5db;border-radius:8px;font-size:14px;box-sizing:border-box;resize:vertical;"></textarea>
            <div style="font-size:12px;color:#9ca3af;margin-top:4px;"><span id="scIsiCount">0</span>/3400</div>
          </div>
          <div style="margin-bottom:16px;display:none;" id="scFileGroup">
            <label style="display:block;font-size:13px;font-weight:600;margin-bottom:6px;">Upload File</label>
            <input type="file" id="scFormFile" style="width:100%;padding:8px;border:1px solid #d1d5db;border-radius:8px;font-size:14px;">
            <div id="scFilePreview" style="margin-top:8px;font-size:13px;color:#6b7280;"></div>
          </div>
        </div>
        <div style="padding:16px 24px;border-top:1px solid #f3f4f6;display:flex;gap:10px;justify-content:flex-end;">
          <button onclick="closeShortcutForm()" style="padding:10px 20px;border:1px solid #e5e7eb;border-radius:8px;background:#fff;color:#374151;cursor:pointer;font-weight:600;">Batal</button>
          <button onclick="saveShortcutForm()" style="padding:10px 24px;background:#5b3df0;color:#fff;border:none;border-radius:8px;cursor:pointer;font-weight:600;">Simpan</button>
        </div>
      </div>
    </div>
  </div>

  <div class="settings-section" id="section-label-kontak">
    <div class="panel">
      <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:20px;">
        <h2 style="margin:0;">Label Kontak</h2>
        <button class="btn" onclick="showLabelForm()">+ Tambah Label</button>
      </div>
      <table class="labels-table" style="width:100%;border-collapse:collapse;">
        <thead>
          <tr>
            <th style="padding:12px 16px;text-align:left;font-size:13px;color:#6b7280;font-weight:600;border-bottom:1px solid #e5e7eb;background:#f9fafb;">Nama Label</th>
            <th style="padding:12px 16px;text-align:left;font-size:13px;color:#6b7280;font-weight:600;border-bottom:1px solid #e5e7eb;background:#f9fafb;">Kategori</th>
            <th style="padding:12px 16px;text-align:left;font-size:13px;color:#6b7280;font-weight:600;border-bottom:1px solid #e5e7eb;background:#f9fafb;">Warna Latar</th>
            <th style="padding:12px 16px;text-align:left;font-size:13px;color:#6b7280;font-weight:600;border-bottom:1px solid #e5e7eb;background:#f9fafb;">Warna Teks</th>
            <th style="padding:12px 16px;text-align:right;font-size:13px;color:#6b7280;font-weight:600;border-bottom:1px solid #e5e7eb;background:#f9fafb;">Aksi</th>
          </tr>
        </thead>
        <tbody id="labelsTableBody">
          <tr><td colspan="5" style="padding:40px;text-align:center;color:#9ca3af;">Memuat...</td></tr>
        </tbody>
      </table>
    </div>

    <!-- Modal form label -->
    <div id="labelFormModal" style="display:none;position:fixed;inset:0;background:rgba(0,0,0,.5);z-index:1000;align-items:center;justify-content:center;">
      <div style="background:#fff;border-radius:12px;padding:24px;width:480px;max-width:90vw;">
        <h3 id="labelFormTitle" style="margin:0 0 20px 0;font-size:18px;font-weight:700;">Tambah Label</h3>
        <input type="hidden" id="labelFormId">
        <div style="margin-bottom:14px;">
          <label style="display:block;font-size:13px;font-weight:600;margin-bottom:6px;">Nama Label</label>
          <input type="text" id="labelFormName" placeholder="contoh: VIP, Donatur Rutin..." style="width:100%;padding:10px 12px;border:1px solid #d1d5db;border-radius:8px;font-size:14px;box-sizing:border-box;">
        </div>
        <div style="margin-bottom:14px;">
          <label style="display:block;font-size:13px;font-weight:600;margin-bottom:6px;">Kategori</label>
          <input type="text" id="labelFormCategory" placeholder="contoh: Donatur, Prospek..." style="width:100%;padding:10px 12px;border:1px solid #d1d5db;border-radius:8px;font-size:14px;box-sizing:border-box;">
        </div>
        <div style="margin-bottom:14px;display:flex;gap:16px;">
          <div style="flex:1;">
            <label style="display:block;font-size:13px;font-weight:600;margin-bottom:6px;">Warna Latar</label>
            <div style="display:flex;gap:8px;align-items:center;">
              <input type="color" id="labelFormBg" value="#eaf97b" style="width:50px;height:38px;border:1px solid #d1d5db;border-radius:6px;cursor:pointer;padding:2px;">
              <input type="text" id="labelFormBgHex" value="#eaf97b" style="width:90px;padding:8px;border:1px solid #d1d5db;border-radius:6px;font-size:13px;font-family:monospace;" oninput="syncColor('labelFormBg','labelFormBgHex');updateLabelPreview()">
            </div>
          </div>
          <div style="flex:1;">
            <label style="display:block;font-size:13px;font-weight:600;margin-bottom:6px;">Warna Teks</label>
            <div style="display:flex;gap:8px;align-items:center;">
              <input type="color" id="labelFormText" value="#000000" style="width:50px;height:38px;border:1px solid #d1d5db;border-radius:6px;cursor:pointer;padding:2px;">
              <input type="text" id="labelFormTextHex" value="#000000" style="width:90px;padding:8px;border:1px solid #d1d5db;border-radius:6px;font-size:13px;font-family:monospace;" oninput="syncColor('labelFormText','labelFormTextHex');updateLabelPreview()">
            </div>
          </div>
        </div>
        <div style="background:#f9fafb;padding:14px;border-radius:8px;margin-bottom:16px;text-align:center;">
          <div style="font-size:12px;color:#6b7280;margin-bottom:8px;">Preview</div>
          <span id="labelPreview" style="display:inline-block;padding:6px 16px;border-radius:12px;font-size:13px;font-weight:600;background:#eaf97b;color:#000000;">Nama Label</span>
        </div>
        <div style="display:flex;gap:10px;justify-content:flex-end;">
          <button onclick="closeLabelFormModal()" style="background:#f3f4f6;color:#374151;border:none;padding:10px 20px;border-radius:8px;font-weight:600;cursor:pointer;">Batal</button>
          <button onclick="saveLabelForm()" style="background:#5b3df0;color:#fff;border:none;padding:10px 20px;border-radius:8px;font-weight:600;cursor:pointer;">Simpan</button>
        </div>
      </div>
    </div>
  </div>

<script>
function setSettingsTab(el) {
  if (el.dataset.tab === "label-kontak") loadLabelsTable();
  if (el.dataset.tab === "wa-blast") loadBlastHistory();
  if (el.dataset.tab === "shortcuts") loadShortcutsTable();
  if (el.dataset.tab === "wa-template") loadTemplateTable();
  document.querySelectorAll(".settings-tab").forEach(t => t.classList.remove("active"));
  document.querySelectorAll(".settings-section").forEach(s => s.classList.remove("active"));
  el.classList.add("active");
  document.getElementById("section-" + el.dataset.tab).classList.add("active");
  if (el.dataset.tab === "wa-blast") loadBlastContacts();
}

// ---- Template Blast ----
async function loadTemplateTable() {
  const tbody = document.getElementById("templateTableBody");
  if (!tbody) return;
  tbody.innerHTML = '<tr><td colspan="6" style="padding:20px;text-align:center;color:#9ca3af;">Memuat dari Meta...</td></tr>';
  try {
    const [res, settRes] = await Promise.all([
      fetch("/api/wa-templates"),
      fetch("/api/settings")
    ]);
    const json = await res.json();
    const sett = await settRes.json();
    const chatTemplates = (sett.wa_chat_templates || []);
    let list = json.templates || [];
    list = list.map(t => ({...t, is_chat_template: chatTemplates.includes(t.name)}));
    if (!list.length) {
      tbody.innerHTML = '<tr><td colspan="6" style="padding:40px;text-align:center;color:#9ca3af;">Belum ada template. Klik "+ Tambah Template".</td></tr>';
      return;
    }
    const statusColor = {APPROVED:"#16a34a",PENDING:"#d97706",REJECTED:"#dc2626",PAUSED:"#6b7280"};
    tbody.innerHTML = list.map(t => {
      const bodyComp = (t.components||[]).find(c => c.type === "BODY");
      const bodyText = bodyComp ? (bodyComp.text||"").substring(0,80) : "-";
      const sc = statusColor[t.status] || "#6b7280";
      return `<tr style="border-bottom:1px solid #f3f4f6;">
        <td style="padding:12px 16px;font-size:14px;font-weight:600;">${t.name}</td>
        <td style="padding:12px 16px;font-size:13px;">${t.category||"-"}</td>
        <td style="padding:12px 16px;font-size:13px;">${t.language||"-"}</td>
        <td style="padding:12px 16px;">
          <span style="display:inline-block;padding:3px 10px;border-radius:12px;font-size:12px;font-weight:600;background:${sc}20;color:${sc};">${t.status}</span>
        </td>
        <td style="padding:12px 16px;font-size:13px;color:#6b7280;max-width:250px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;">${bodyText}</td>
        <td style="padding:12px 16px;text-align:right;">
          <div style="display:flex;align-items:center;gap:4px;justify-content:flex-end;">
            <label style="display:flex;align-items:center;gap:4px;cursor:pointer;font-size:12px;color:#374151;margin-right:4px;" title="Set sebagai Template Chat">
              <input type="checkbox" onchange="setTemplateChat('${t.id}','${t.name}',this.checked)" ${t.is_chat_template ? 'checked' : ''} style="cursor:pointer;">
              <span>Template Chat</span>
            </label>
            <button onclick="copyTemplate('${t.id}','${t.name}')" style="background:none;border:none;cursor:pointer;color:#5b3df0;font-size:14px;padding:4px 8px;" title="Copy">📋</button>
            <button onclick="deleteTemplate('${t.id}','${t.name}')" style="background:none;border:none;cursor:pointer;color:#dc2626;font-size:14px;padding:4px 8px;" title="Hapus">🗑</button>
          </div>
        </td>
      </tr>`;
    }).join("");
  } catch(e) {
    tbody.innerHTML = '<tr><td colspan="6" style="padding:20px;color:#dc2626;">Gagal memuat template.</td></tr>';
  }
}

function showTemplateForm() {
  document.getElementById("tmplCategory").value = "MARKETING";
  document.getElementById("tmplName").value = "";
  document.getElementById("tmplNameCount").textContent = "0";
  document.getElementById("tmplHeaderType").value = "";
  document.getElementById("tmplHeaderTextGroup").style.display = "none";
  document.getElementById("tmplHeaderText").value = "";
  document.getElementById("tmplBody").value = "";
  document.getElementById("tmplBodyCount").textContent = "0";
  document.getElementById("tmplFooter").value = "";
  document.getElementById("tmplFooterCount").textContent = "0";
  document.getElementById("tmplLanguage").value = "";
  document.getElementById("tmplBtnType").value = "";
  document.getElementById("tmplBtnGroup").style.display = "none";
  document.getElementById("tmplBtnList").innerHTML = "";
  updateTemplatePreview();
  document.getElementById("templateFormModal").style.display = "flex";
}

function closeTemplateForm() {
  document.getElementById("templateFormModal").style.display = "none";
}

function onHeaderTypeChange() {
  const t = document.getElementById("tmplHeaderType").value;
  const group = document.getElementById("tmplHeaderTextGroup");
  const textInput = document.getElementById("tmplHeaderText");
  const fileGroup = document.getElementById("tmplHeaderFileGroup");
  const label = document.getElementById("tmplHeaderLabel");

  if (!t) {
    group.style.display = "none";
  } else if (t === "TEXT") {
    group.style.display = "block";
    textInput.style.display = "block";
    fileGroup.style.display = "none";
    label.textContent = "Teks Judul";
    // Accept filter
    const af = {"IMAGE":"image/*","VIDEO":"video/*","DOCUMENT":".pdf,.doc,.docx"};
    document.getElementById("tmplHeaderFile").accept = af[t] || "*";
  } else {
    group.style.display = "block";
    textInput.style.display = "none";
    fileGroup.style.display = "block";
    label.textContent = t + " File";
    const af = {"IMAGE":"image/*","VIDEO":"video/*","DOCUMENT":".pdf,.doc,.docx"};
    document.getElementById("tmplHeaderFile").accept = af[t] || "*";
  }
  updateTemplatePreview();
}

// Emoji picker
const EMOJIS = ["😀","😊","🙏","❤️","✅","🎉","👍","🔥","💪","😍","🤲","💝","🌙","⭐","🕌","📿","💰","🎁","📢","📣","✨","🙌","💯","🤝","👋","😇","🥰","😢","😭","🤗","💬","📱","🔔","⚡","🌟","💫","🎊","🎀","🌸","🌺","🍀","🌈","☀️","🌙","⏰","📅","📌","🔑","💡","📝","✍️","📊","💼","🏠","🚀","🌍","🤩","😎","🥳","🎯","💎","🏆","🎖️","🌻","🦋"];

function initEmojiPicker() {
  const grid = document.getElementById("emojiGrid");
  if (!grid || grid.children.length > 0) return;
  grid.innerHTML = EMOJIS.map(e =>
    `<span onclick="insertEmoji('${e}')" style="cursor:pointer;font-size:20px;padding:4px;border-radius:4px;display:inline-block;" onmouseover="this.style.background='#f3f4f6'" onmouseout="this.style.background=''">${e}</span>`
  ).join("");
}

function toggleEmojiPicker() {
  const picker = document.getElementById("emojiPicker");
  if (picker.style.display === "none") {
    initEmojiPicker();
    picker.style.display = "block";
  } else {
    picker.style.display = "none";
  }
}

function insertEmoji(emoji) {
  const ta = document.getElementById("tmplBody");
  const pos = ta.selectionStart || ta.value.length;
  ta.value = ta.value.slice(0, pos) + emoji + ta.value.slice(pos);
  document.getElementById("emojiPicker").style.display = "none";
  updateTemplatePreview();
  document.getElementById("tmplBodyCount").textContent = ta.value.length;
}

document.addEventListener("click", function(e) {
  const picker = document.getElementById("emojiPicker");
  if (picker && !picker.contains(e.target) && !e.target.closest("[onclick*=toggleEmojiPicker]")) {
    picker.style.display = "none";
  }
});

function onBtnTypeChange() {
  const t = document.getElementById("tmplBtnType").value;
  document.getElementById("tmplBtnGroup").style.display = t ? "block" : "none";
  document.getElementById("tmplBtnList").innerHTML = "";
  if (t) addTemplateButton();
}

let _btnCount = 0;
function addTemplateButton() {
  const btnType = document.getElementById("tmplBtnType").value;
  const id = "btn_" + (++_btnCount);
  const list = document.getElementById("tmplBtnList");
  const div = document.createElement("div");
  div.id = id;
  div.style.cssText = "background:#f9fafb;border:1px solid #e5e7eb;border-radius:8px;padding:12px;margin-bottom:8px;";
  if (btnType === "QUICK_REPLY") {
    div.innerHTML = `
      <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:8px;">
        <span style="font-size:13px;font-weight:600;">Quick Reply</span>
        <span onclick="document.getElementById('${id}').remove()" style="cursor:pointer;color:#dc2626;font-size:16px;">✕</span>
      </div>
      <input type="text" placeholder="Teks tombol..." data-btn-text style="width:100%;padding:8px;border:1px solid #d1d5db;border-radius:6px;font-size:13px;box-sizing:border-box;">
      <input type="hidden" data-btn-type value="QUICK_REPLY">
    `;
  } else {
    div.innerHTML = `
      <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:8px;">
        <span style="font-size:13px;font-weight:600;">Call to Action</span>
        <span onclick="document.getElementById('${id}').remove()" style="cursor:pointer;color:#dc2626;font-size:16px;">✕</span>
      </div>
      <select data-btn-subtype style="width:100%;padding:8px;border:1px solid #d1d5db;border-radius:6px;font-size:13px;margin-bottom:6px;">
        <option value="URL">URL</option>
        <option value="PHONE_NUMBER">Phone Number</option>
      </select>
      <input type="text" placeholder="Teks tombol..." data-btn-text style="width:100%;padding:8px;border:1px solid #d1d5db;border-radius:6px;font-size:13px;box-sizing:border-box;margin-bottom:6px;">
      <input type="text" placeholder="URL atau nomor telepon..." data-btn-value style="width:100%;padding:8px;border:1px solid #d1d5db;border-radius:6px;font-size:13px;box-sizing:border-box;">
      <input type="hidden" data-btn-type value="CALL_TO_ACTION">
    `;
  }
  list.appendChild(div);
}

function insertParam() {
  const ta = document.getElementById("tmplBody");
  const existing = (ta.value.match(/\{\{(\d+)\}\}/g) || []);
  const nextNum = existing.length + 1;
  const pos = ta.selectionStart;
  ta.value = ta.value.slice(0, pos) + "{{" + nextNum + "}}" + ta.value.slice(pos);
  updateTemplatePreview();
}

function wrapText(marker) {
  const ta = document.getElementById("tmplBody");
  const start = ta.selectionStart, end = ta.selectionEnd;
  const selected = ta.value.slice(start, end) || "teks";
  ta.value = ta.value.slice(0, start) + marker + selected + marker + ta.value.slice(end);
  updateTemplatePreview();
}

function updateTemplatePreview() {
  const headerType = document.getElementById("tmplHeaderType").value;
  const header = document.getElementById("tmplHeaderText").value;
  const body = document.getElementById("tmplBody").value;
  const footer = document.getElementById("tmplFooter").value;
  const hEl = document.getElementById("tmplPreviewHeader");
  const hImgEl = document.getElementById("tmplPreviewHeaderImg");
  const bEl = document.getElementById("tmplPreviewBody");
  const fEl = document.getElementById("tmplPreviewFooter");
  const btnsEl = document.getElementById("tmplPreviewBtns");

  // Header
  if (headerType === "TEXT" && header) {
    hEl.textContent = header;
    hEl.style.display = "block";
    hImgEl.style.display = "none";
  } else if (["IMAGE","VIDEO","DOCUMENT"].includes(headerType)) {
    const icons = {IMAGE:"🖼 Image", VIDEO:"🎥 Video", DOCUMENT:"📄 Document"};
    hImgEl.textContent = icons[headerType] || "📎 Media";
    hImgEl.style.display = "flex";
    hEl.style.display = "none";
  } else {
    hEl.style.display = "none";
    hImgEl.style.display = "none";
  }

  // Body - render bold (*text*) dan italic (_text_)
  let bodyHtml = (body || "").replace(/</g,"&lt;")
    .replace(/\*(.*?)\*/g,"<strong>$1</strong>")
    .replace(/_(.*?)_/g,"<em>$1</em>")
    .replace(/~(.*?)~/g,"<s>$1</s>")
    .split(String.fromCharCode(10)).join("<br>");
  bEl.innerHTML = bodyHtml;

  // Footer
  if (footer) { fEl.textContent = footer; fEl.style.display = "block"; } else { fEl.style.display = "none"; }

  // Buttons preview
  const btnType = document.getElementById("tmplBtnType").value;
  if (btnsEl) {
    const btnEls = [];
    document.querySelectorAll("#tmplBtnList [data-btn-text]").forEach(el => {
      if (el.value) btnEls.push(`<div style="background:#fff;border:1px solid #e5e7eb;border-radius:8px;padding:8px;text-align:center;font-size:13px;color:#5b3df0;font-weight:600;">${el.value}</div>`);
    });
    btnsEl.innerHTML = btnEls.join("");
  }

  document.getElementById("tmplNameCount").textContent = document.getElementById("tmplName").value.length;
}

function previewTemplate() {
  updateTemplatePreview();
  document.getElementById("tmplPreview").scrollIntoView({behavior:"smooth"});
}

async function saveTemplateForm() {
  const name = document.getElementById("tmplName").value.trim();
  const category = document.getElementById("tmplCategory").value;
  const language = document.getElementById("tmplLanguage").value;
  const body_text = document.getElementById("tmplBody").value.trim();
  const header_type = document.getElementById("tmplHeaderType").value;
  const header_text = document.getElementById("tmplHeaderText").value.trim();
  const footer_text = document.getElementById("tmplFooter").value.trim();

  if (!name || !body_text || !language) {
    alert("Nama template, Isi, dan Bahasa wajib diisi");
    return;
  }

  // Collect buttons
  const buttons = [];
  document.querySelectorAll("#tmplBtnList > div").forEach(div => {
    const btnType = (div.querySelector("[data-btn-type]") || {}).value || "";
    const btnText = (div.querySelector("[data-btn-text]") || {}).value || "";
    const btnSubtype = (div.querySelector("[data-btn-subtype]") || {}).value || "URL";
    const btnValue = (div.querySelector("[data-btn-value]") || {}).value || "";
    if (btnText) buttons.push({type: btnType === "QUICK_REPLY" ? "QUICK_REPLY" : btnSubtype, text: btnText, value: btnValue});
  });

  const payload = {name, category, language, body_text, header_type, header_text, footer_text, buttons};

  try {
    const res = await fetch("/api/wa-templates", {
      method: "POST",
      headers: {"Content-Type": "application/json"},
      body: JSON.stringify(payload)
    });
    const json = await res.json();
    if (json.error) { alert("Error: " + json.error); return; }
    closeTemplateForm();
    alert("Template berhasil disubmit ke Meta. Status: PENDING (menunggu approval 1-24 jam).");
    loadTemplateTable();
  } catch(e) {
    alert("Error: " + e.message);
  }
}

async function copyTemplate(id, name) {
  const newName = prompt("Nama template baru (huruf kecil, underscore):", name + "_copy");
  if (!newName) return;
  // Fetch template detail
  try {
    const res = await fetch("/api/wa-templates");
    const json = await res.json();
    const t = (json.templates || []).find(x => x.id === id);
    if (!t) { alert("Template tidak ditemukan"); return; }
    const bodyComp = (t.components||[]).find(c => c.type === "BODY");
    const headerComp = (t.components||[]).find(c => c.type === "HEADER");
    const footerComp = (t.components||[]).find(c => c.type === "FOOTER");
    const payload = {
      name: newName.toLowerCase().replace(/[^a-z0-9_]/g,""),
      category: t.category,
      language: t.language,
      body_text: bodyComp ? bodyComp.text : "",
      header_type: headerComp ? headerComp.format : "",
      header_text: (headerComp && headerComp.format === "TEXT") ? headerComp.text : "",
      footer_text: footerComp ? footerComp.text : "",
      buttons: []
    };
    const saveRes = await fetch("/api/wa-templates", {
      method: "POST",
      headers: {"Content-Type":"application/json"},
      body: JSON.stringify(payload)
    });
    const saveJson = await saveRes.json();
    if (saveJson.error) { alert("Error: " + saveJson.error); return; }
    alert("Template berhasil di-copy sebagai '" + payload.name + "'. Status: PENDING.");
    loadTemplateTable();
  } catch(e) { alert("Error: " + e.message); }
}

async function setTemplateChat(id, name, checked) {
  try {
    await fetch("/api/wa-templates/chat-setting", {
      method: "POST",
      headers: {"Content-Type":"application/json"},
      body: JSON.stringify({name, enabled: checked})
    });
    // reloadChatTemplates hanya ada di halaman /chat
    if (typeof reloadChatTemplates === "function") {
      await reloadChatTemplates();
    }
  } catch(e) { alert("Error: " + e.message); }
}

async function deleteTemplate(id, name) {
  if (!confirm("Hapus template '" + name + "'? Tindakan ini tidak bisa dibatalkan.")) return;
  try {
    const res = await fetch("/api/wa-templates/" + id + "?name=" + encodeURIComponent(name), {method:"DELETE"});
    const json = await res.json();
    loadTemplateTable();
  } catch(e) { alert("Error: " + e.message); }
}

// ---- Shortcuts Pesan ----
async function loadShortcutsTable() {
  const tbody = document.getElementById("shortcutsTableBody");
  if (!tbody) return;
  try {
    const res = await fetch("/api/shortcuts");
    const json = await res.json();
    const list = json.shortcuts || [];
    if (!list.length) {
      tbody.innerHTML = '<tr><td colspan="5" style="padding:40px;text-align:center;color:#9ca3af;">Belum ada shortcut. Klik "+ Tambah Shortcut".</td></tr>';
      return;
    }
    const typeIcon = {TEXT:"📝", IMAGE:"🖼", DOCUMENT:"📄", VIDEO:"🎥"};
    tbody.innerHTML = list.map(s => `
      <tr style="border-bottom:1px solid #f3f4f6;">
        <td style="padding:12px 16px;font-size:14px;font-weight:600;">${s.title}</td>
        <td style="padding:12px 16px;font-size:13px;">${typeIcon[s.content_type]||""} ${s.content_type}</td>
        <td style="padding:12px 16px;"><code style="background:#f3f4f6;padding:2px 8px;border-radius:4px;font-size:13px;">${s.shortcut}</code></td>
        <td style="padding:12px 16px;font-size:13px;color:#6b7280;max-width:200px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;">${s.file_name ? s.file_name : (s.isi||"").substring(0,60)}</td>
        <td style="padding:12px 16px;text-align:right;">
          <button onclick="deleteShortcut('${s.id}')" style="background:none;border:none;cursor:pointer;color:#dc2626;font-size:14px;padding:4px 8px;" title="Hapus">🗑</button>
        </td>
      </tr>
    `).join("");
  } catch(e) {
    tbody.innerHTML = '<tr><td colspan="5" style="padding:20px;color:#dc2626;">Gagal memuat shortcuts.</td></tr>';
  }
}

function showShortcutForm() {
  document.getElementById("shortcutFormTitle").textContent = "Buat Shortcuts Pesan";
  document.getElementById("scFormId").value = "";
  document.getElementById("scFormShortcut").value = "";
  document.getElementById("scFormTitle").value = "";
  document.getElementById("scFormType").value = "";
  document.getElementById("scFormIsi").value = "";
  document.getElementById("scIsiCount").textContent = "0";
  document.getElementById("scIsiGroup").style.display = "block";
  document.getElementById("scFileGroup").style.display = "none";
  document.getElementById("scFilePreview").textContent = "";
  document.getElementById("shortcutFormModal").style.display = "flex";
}

function closeShortcutForm() {
  document.getElementById("shortcutFormModal").style.display = "none";
}

function onScTypeChange() {
  const t = document.getElementById("scFormType").value;
  const isMedia = ["IMAGE","DOCUMENT","VIDEO"].includes(t);
  document.getElementById("scIsiGroup").style.display = isMedia ? "none" : "block";
  document.getElementById("scFileGroup").style.display = isMedia ? "block" : "none";
  // Set accept filter
  const fileInput = document.getElementById("scFormFile");
  if (t === "IMAGE") fileInput.accept = "image/*";
  else if (t === "VIDEO") fileInput.accept = "video/*";
  else if (t === "DOCUMENT") fileInput.accept = ".pdf,.doc,.docx,.xls,.xlsx,.ppt,.pptx,.txt,.zip";
  // Update shortcut placeholder
  const scInput = document.getElementById("scFormShortcut");
  if (!scInput.value) scInput.value = "/";
}

document.addEventListener("DOMContentLoaded", function() {
  const isiEl = document.getElementById("scFormIsi");
  if (isiEl) isiEl.addEventListener("input", function() {
    document.getElementById("scIsiCount").textContent = this.value.length;
  });
  const fileEl = document.getElementById("scFormFile");
  if (fileEl) fileEl.addEventListener("change", function() {
    const f = this.files[0];
    document.getElementById("scFilePreview").textContent = f ? f.name + " (" + (f.size/1024).toFixed(1) + " KB)" : "";
  });
});

async function saveShortcutForm() {
  const shortcut = document.getElementById("scFormShortcut").value.trim();
  const title = document.getElementById("scFormTitle").value.trim();
  const content_type = document.getElementById("scFormType").value;
  const isi = document.getElementById("scFormIsi").value.trim();
  const file = document.getElementById("scFormFile").files[0];

  if (!shortcut || !title || !content_type) {
    alert("Shortcut, Judul, dan Content Type wajib diisi");
    return;
  }
  if (["IMAGE","DOCUMENT","VIDEO"].includes(content_type) && !file) {
    alert("File wajib diupload untuk tipe " + content_type);
    return;
  }

  const formData = new FormData();
  formData.append("shortcut", shortcut);
  formData.append("title", title);
  formData.append("content_type", content_type);
  formData.append("isi", isi);
  if (file) formData.append("file", file);

  try {
    const res = await fetch("/api/shortcuts", {method:"POST", body:formData});
    const json = await res.json();
    if (json.error) { alert("Error: " + json.error); return; }
    closeShortcutForm();
    loadShortcutsTable();
  } catch(e) {
    alert("Error: " + e.message);
  }
}

async function deleteShortcut(id) {
  if (!confirm("Hapus shortcut ini?")) return;
  try {
    await fetch("/api/shortcuts/" + id, {method:"DELETE"});
    loadShortcutsTable();
  } catch(e) { alert("Error: " + e.message); }
}

// ---- WA Blast ----
let _blastAllContacts = [];
let _blastTemplates = [];

function showBlastForm() {
  document.getElementById("blastListView").style.display = "none";
  document.getElementById("blastFormView").style.display = "block";
  loadBlastTemplates();
  loadBlastContactList();
  loadBlastLabels();
}

function showBlastList() {
  document.getElementById("blastFormView").style.display = "none";
  document.getElementById("blastListView").style.display = "block";
  loadBlastHistory();
}

async function loadBlastHistory() {
  const tbody = document.getElementById("blastHistoryBody");
  if (!tbody) return;
  tbody.innerHTML = '<tr><td colspan="11" style="padding:20px;text-align:center;color:#9ca3af;">Memuat...</td></tr>';
  try {
    const res = await fetch("/api/blast/history");
    const json = await res.json();
    const list = json.history || [];
    if (!list.length) {
      tbody.innerHTML = '<tr><td colspan="11" style="padding:40px;text-align:center;color:#9ca3af;">Belum ada blast. Klik "+ Buat Blast".</td></tr>';
      return;
    }
    const statusColor = {DONE:"#16a34a", PENDING:"#d97706", FAILED:"#dc2626", SENDING:"#2563eb"};
    tbody.innerHTML = list.map(b => {
      const sc = statusColor[b.status] || "#6b7280";
      const dt = b.created_at ? b.created_at.replace("T"," ").substring(0,16) : "-";
      return `<tr style="border-bottom:1px solid #f3f4f6;">
        <td style="padding:12px 16px;font-size:13px;font-weight:600;color:#5b3df0;cursor:pointer;" onclick="showBlastDetail('${b.id}')">${b.judul_campaign || b.judul || "-"}</td>
        <td style="padding:12px 16px;font-size:12px;color:#6b7280;white-space:nowrap;">${dt}</td>
        <td style="padding:12px 16px;font-size:13px;">${b.template_name||"-"}</td>
        <td style="padding:12px 16px;font-size:12px;">${b.kategori||"-"}</td>
        <td style="padding:12px 16px;">
          <span style="display:inline-block;padding:3px 10px;border-radius:12px;font-size:11px;font-weight:700;background:${sc}20;color:${sc};">${b.status}</span>
        </td>
        <td style="padding:12px 16px;text-align:center;font-size:13px;">${b.total||0}</td>
        <td style="padding:12px 16px;text-align:center;font-size:13px;color:#2563eb;">${b.sent||0}</td>
        <td style="padding:12px 16px;text-align:center;font-size:13px;color:#16a34a;">${b.delivered||0}</td>
        <td style="padding:12px 16px;text-align:center;font-size:13px;color:#7c3aed;">${b.read||0}</td>
        <td style="padding:12px 16px;text-align:center;font-size:13px;color:#dc2626;">${b.failed||0}</td>
        <td style="padding:12px 16px;text-align:right;">
          <button onclick="deleteBlast('${b.id}')" style="background:none;border:none;cursor:pointer;color:#dc2626;font-size:14px;" title="Hapus">🗑</button>
        </td>
      </tr>`;
    }).join("");
  } catch(e) {
    tbody.innerHTML = '<tr><td colspan="11" style="padding:20px;color:#dc2626;">Gagal memuat.</td></tr>';
  }
}

async function deleteBlast(id) {
  if (!confirm("Hapus history blast ini?")) return;
  await fetch("/api/blast/history/" + id, {method:"DELETE"});
  loadBlastHistory();
}

function showBlastDetail(id) {
  alert("Detail blast " + id + " - Coming soon");
}

async function loadBlastTemplates() {
  const sel = document.getElementById("blastTemplate");
  if (!sel) return;
  try {
    const res = await fetch("/api/wa-templates");
    const json = await res.json();
    _blastTemplates = (json.templates || json.raw?.data || []).filter(t => t.status === "APPROVED");
    sel.innerHTML = '<option value="">-- Pilih Template --</option>' +
      _blastTemplates.map(t => `<option value="${t.name}" data-lang="${t.language||"id"}">${t.name.replace(/_/g," ")} (${t.language||"id"})</option>`).join("");
  } catch(e) {}
}

function onBlastTemplateChange() {
  const sel = document.getElementById("blastTemplate");
  const name = sel.value;
  const preview = document.getElementById("blastTemplatePreview");
  const previewBody = document.getElementById("blastTemplatePreviewBody");
  const paramsGroup = document.getElementById("blastParamsGroup");
  const paramsList = document.getElementById("blastParamsList");

  if (!name) {
    preview.style.display = "none";
    paramsGroup.style.display = "none";
    return;
  }

  const tmpl = _blastTemplates.find(t => t.name === name);
  if (!tmpl) return;

  const bodyComp = (tmpl.components||[]).find(c => c.type === "BODY");
  const bodyText = bodyComp ? bodyComp.text : "";
  previewBody.textContent = bodyText;
  preview.style.display = "block";

  // Detect parameters {{1}}, {{2}}, etc
  const params = bodyText.match(/\{\{(\d+)\}\}/g) || [];
  if (params.length) {
    paramsList.innerHTML = params.map((p, i) => `
      <div style="margin-bottom:8px;">
        <label style="font-size:12px;color:#6b7280;display:block;margin-bottom:4px;">Parameter ${p}</label>
        <input type="text" data-param="${i}" placeholder="Isi parameter ${p}..."
          style="width:100%;padding:8px 12px;border:1px solid #d1d5db;border-radius:8px;font-size:13px;box-sizing:border-box;">
      </div>
    `).join("");
    paramsGroup.style.display = "block";
  } else {
    paramsGroup.style.display = "none";
  }
}

async function loadBlastContactList() {
  const el = document.getElementById("blastContacts");
  if (!el) return;
  try {
    const res = await fetch("/api/inbox");
    const json = await res.json();
    _blastAllContacts = json.contacts || [];
    renderBlastContacts(_blastAllContacts);
  } catch(e) {
    el.innerHTML = '<div style="color:#dc2626;padding:8px;">Gagal memuat kontak.</div>';
  }
}

function renderBlastContacts(contacts) {
  const el = document.getElementById("blastContacts");
  const countEl = document.getElementById("blastCount");
  if (!el) return;
  if (!contacts.length) {
    el.innerHTML = '<div style="padding:12px;color:#9ca3af;text-align:center;">Tidak ada kontak.</div>';
    return;
  }
  el.innerHTML = contacts.map(c => `
    <div class="blast-contact-item">
      <input type="checkbox" class="blast-chk" value="${c.phone}" id="bchk_${c.phone}">
      <label for="bchk_${c.phone}" style="cursor:pointer;flex:1;">
        <div style="font-size:13px;font-weight:600;">${c.name||c.phone}</div>
        <div style="font-size:11px;color:#9ca3af;">+${c.phone}</div>
      </label>
    </div>
  `).join("");
  document.querySelectorAll(".blast-chk").forEach(chk => {
    chk.addEventListener("change", updateBlastCount);
  });
  updateBlastCount();
}

function filterBlastContacts(val) {
  const q = val.toLowerCase();
  const filtered = q ? _blastAllContacts.filter(c =>
    (c.name||"").toLowerCase().includes(q) || (c.phone||"").includes(q)
  ) : _blastAllContacts;
  renderBlastContacts(filtered);
}

function updateBlastCount() {
  const checked = document.querySelectorAll(".blast-chk:checked").length;
  const el = document.getElementById("blastCount");
  if (el) el.textContent = checked + " dipilih";
}

function toggleAllBlast() {
  const chks = document.querySelectorAll(".blast-chk");
  const allChecked = Array.from(chks).every(c => c.checked);
  chks.forEach(c => { c.checked = !allChecked; });
  updateBlastCount();
}

async function loadBlastLabels() {
  const sel = document.getElementById("blastLabelFilter");
  if (!sel) return;
  try {
    const res = await fetch("/api/labels");
    const json = await res.json();
    const labels = json.labels || [];
    sel.innerHTML = '<option value="">-- Semua Kontak --</option>' +
      labels.map(l => `<option value="${l.name}">${l.name}</option>`).join("");
  } catch(e) {}
}

function downloadBlastTemplate() {
  window.open("/api/blast/template-download", "_blank");
}

async function onBlastFileUpload(input) {
  const file = input.files[0];
  if (!file) return;
  const msgEl = document.getElementById("blastFileMsg");
  msgEl.style.color = "#6b7280";
  msgEl.textContent = "Memproses file...";

  const formData = new FormData();
  formData.append("file", file);

  try {
    const res = await fetch("/api/blast/upload-contacts", {method:"POST", body:formData});
    const json = await res.json();
    if (json.error) { msgEl.style.color="#dc2626"; msgEl.textContent="Error: "+json.error; return; }

    const uploaded = json.contacts || [];
    if (!uploaded.length) { msgEl.style.color="#dc2626"; msgEl.textContent="Tidak ada kontak valid di file."; return; }

    // Gabungkan dengan kontak yang ada, hindari duplikat
    const existingPhones = new Set(_blastAllContacts.map(c => c.phone));
    const newContacts = uploaded.filter(c => !existingPhones.has(c.phone));

    // Tambah kontak dari file ke daftar (mark sebagai uploaded)
    const uploadedMarked = uploaded.map(c => ({...c, fromFile: true}));

    // Render kontak dari file (ganti daftar kontak dengan yang dari file)
    const el = document.getElementById("blastContacts");
    const countEl = document.getElementById("blastCount");

    el.innerHTML = uploadedMarked.map(c => `
      <div class="blast-contact-item">
        <input type="checkbox" class="blast-chk" value="${c.phone}" id="bchk_${c.phone}" checked>
        <label for="bchk_${c.phone}" style="cursor:pointer;flex:1;">
          <div style="font-size:13px;font-weight:600;">${c.name||c.phone}</div>
          <div style="font-size:11px;color:#9ca3af;">+${c.phone}</div>
        </label>
        <span style="font-size:10px;background:#dcfce7;color:#16a34a;padding:2px 6px;border-radius:4px;">dari file</span>
      </div>
    `).join("");

    document.querySelectorAll(".blast-chk").forEach(chk => {
      chk.addEventListener("change", updateBlastCount);
    });
    updateBlastCount();

    msgEl.style.color = "#16a34a";
    msgEl.textContent = `✓ ${uploaded.length} kontak berhasil dimuat dari file. Semua otomatis dipilih.`;
  } catch(e) {
    msgEl.style.color = "#dc2626";
    msgEl.textContent = "Error: " + e.message;
  }
}

function downloadBlastTemplate() {
  window.open("/api/blast/template-download", "_blank");
}

async function onBlastFileUpload(input) {
  const file = input.files[0];
  if (!file) return;
  const msgEl = document.getElementById("blastFileMsg");
  msgEl.style.color = "#6b7280";
  msgEl.textContent = "Memproses file...";

  const formData = new FormData();
  formData.append("file", file);

  try {
    const res = await fetch("/api/blast/upload-contacts", {method:"POST", body:formData});
    const json = await res.json();
    if (json.error) { msgEl.style.color="#dc2626"; msgEl.textContent="Error: "+json.error; return; }

    const uploaded = json.contacts || [];
    if (!uploaded.length) { msgEl.style.color="#dc2626"; msgEl.textContent="Tidak ada kontak valid di file."; return; }

    // Gabungkan dengan kontak yang ada, hindari duplikat
    const existingPhones = new Set(_blastAllContacts.map(c => c.phone));
    const newContacts = uploaded.filter(c => !existingPhones.has(c.phone));

    // Tambah kontak dari file ke daftar (mark sebagai uploaded)
    const uploadedMarked = uploaded.map(c => ({...c, fromFile: true}));

    // Render kontak dari file (ganti daftar kontak dengan yang dari file)
    const el = document.getElementById("blastContacts");
    const countEl = document.getElementById("blastCount");

    el.innerHTML = uploadedMarked.map(c => `
      <div class="blast-contact-item">
        <input type="checkbox" class="blast-chk" value="${c.phone}" id="bchk_${c.phone}" checked>
        <label for="bchk_${c.phone}" style="cursor:pointer;flex:1;">
          <div style="font-size:13px;font-weight:600;">${c.name||c.phone}</div>
          <div style="font-size:11px;color:#9ca3af;">+${c.phone}</div>
        </label>
        <span style="font-size:10px;background:#dcfce7;color:#16a34a;padding:2px 6px;border-radius:4px;">dari file</span>
      </div>
    `).join("");

    document.querySelectorAll(".blast-chk").forEach(chk => {
      chk.addEventListener("change", updateBlastCount);
    });
    updateBlastCount();

    msgEl.style.color = "#16a34a";
    msgEl.textContent = `✓ ${uploaded.length} kontak berhasil dimuat dari file. Semua otomatis dipilih.`;
  } catch(e) {
    msgEl.style.color = "#dc2626";
    msgEl.textContent = "Error: " + e.message;
  }
}

function onScheduleChange(el) {
  document.getElementById("blastScheduleInput").style.display =
    el.value === "schedule" ? "block" : "none";
}

async function sendBlastNew() {
  const templateName = document.getElementById("blastTemplate").value;
  const judul = document.getElementById("blastJudul").value.trim();
  const judulCampaign = document.getElementById("blastJudulCampaign").value.trim();
  const kategori = document.getElementById("blastKategori").value;
  const headerType = document.getElementById("blastHeaderType").value;

  if (!templateName) { alert("Pilih template terlebih dahulu"); return; }
  if (!judul) { alert("Judul wajib diisi"); return; }

  const phones = Array.from(document.querySelectorAll(".blast-chk:checked")).map(c => c.value);
  if (!phones.length) { alert("Pilih minimal 1 kontak"); return; }

  // Collect params
  const body_params = Array.from(document.querySelectorAll("#blastParamsList input")).map(i => i.value);

  const sel = document.getElementById("blastTemplate");
  const opt = sel.options[sel.selectedIndex];
  const lang = opt ? (opt.dataset.lang || "id") : "id";

  const msgEl = document.getElementById("blastSendMsg");
  msgEl.style.color = "#6b7280";
  msgEl.textContent = "Mengirim blast ke " + phones.length + " kontak...";

  try {
    const res = await fetch("/api/blast", {
      method: "POST",
      headers: {"Content-Type":"application/json"},
      body: JSON.stringify({
        phones, template_name: templateName, template_language: lang,
        judul, judul_campaign: judulCampaign, kategori, header_type: headerType, body_params
      })
    });
    const json = await res.json();
    if (json.error) { msgEl.style.color="#dc2626"; msgEl.textContent="Error: "+json.error; return; }
    msgEl.style.color = "#16a34a";
    msgEl.textContent = "Berhasil! " + json.sent + "/" + json.total + " pesan terkirim.";
    setTimeout(() => { showBlastList(); }, 2000);
  } catch(e) {
    msgEl.style.color = "#dc2626";
    msgEl.textContent = "Error: " + e.message;
  }
}

// ---- Label Kontak ----
async function loadLabelsTable() {
  const tbody = document.getElementById("labelsTableBody");
  if (!tbody) return;
  try {
    const res = await fetch("/api/labels");
    const json = await res.json();
    const labels = json.labels || [];
    if (!labels.length) {
      tbody.innerHTML = '<tr><td colspan="5" style="padding:40px;text-align:center;color:#9ca3af;">Belum ada label. Klik "+ Tambah Label" untuk membuat.</td></tr>';
      return;
    }
    tbody.innerHTML = labels.map(l => `
      <tr>
        <td style="padding:14px 16px;border-bottom:1px solid #f3f4f6;">
          <span style="display:inline-block;padding:4px 12px;border-radius:12px;font-size:12px;font-weight:600;background:${l.bg_color||"#e0e7ff"};color:${l.text_color||"#4338ca"};">${l.name}</span>
        </td>
        <td style="padding:14px 16px;border-bottom:1px solid #f3f4f6;font-size:14px;color:#6b7280;">${l.category||"-"}</td>
        <td style="padding:14px 16px;border-bottom:1px solid #f3f4f6;">
          <div style="display:flex;align-items:center;gap:8px;">
            <div style="width:20px;height:20px;border-radius:4px;background:${l.bg_color||"#e0e7ff"};border:1px solid #e5e7eb;"></div>
            <span style="font-family:monospace;font-size:12px;">${l.bg_color||""}</span>
          </div>
        </td>
        <td style="padding:14px 16px;border-bottom:1px solid #f3f4f6;">
          <div style="display:flex;align-items:center;gap:8px;">
            <div style="width:20px;height:20px;border-radius:4px;background:${l.text_color||"#000"};border:1px solid #e5e7eb;"></div>
            <span style="font-family:monospace;font-size:12px;">${l.text_color||""}</span>
          </div>
        </td>
        <td style="padding:14px 16px;border-bottom:1px solid #f3f4f6;text-align:right;">
          <button onclick="editLabelForm('${l.id}','${l.name.replace(/'/g,"\'")}','${l.category||""}','${l.bg_color||"#eaf97b"}','${l.text_color||"#000000"}')" style="background:none;border:none;cursor:pointer;color:#5b3df0;font-size:14px;padding:4px 8px;" title="Edit">✏️</button>
          <button onclick="deleteLabelRow('${l.id}')" style="background:none;border:none;cursor:pointer;color:#dc2626;font-size:14px;padding:4px 8px;" title="Hapus">🗑</button>
        </td>
      </tr>
    `).join("");
  } catch(e) {
    tbody.innerHTML = '<tr><td colspan="5" style="padding:20px;color:#dc2626;">Gagal memuat label.</td></tr>';
  }
}

function showLabelForm() {
  document.getElementById("labelFormTitle").textContent = "Tambah Label";
  document.getElementById("labelFormId").value = "";
  document.getElementById("labelFormName").value = "";
  document.getElementById("labelFormCategory").value = "";
  document.getElementById("labelFormBg").value = "#eaf97b";
  document.getElementById("labelFormBgHex").value = "#eaf97b";
  document.getElementById("labelFormText").value = "#000000";
  document.getElementById("labelFormTextHex").value = "#000000";
  updateLabelPreview();
  document.getElementById("labelFormModal").style.display = "flex";
}

function editLabelForm(id, name, category, bg, text) {
  document.getElementById("labelFormTitle").textContent = "Edit Label";
  document.getElementById("labelFormId").value = id;
  document.getElementById("labelFormName").value = name;
  document.getElementById("labelFormCategory").value = category;
  document.getElementById("labelFormBg").value = bg;
  document.getElementById("labelFormBgHex").value = bg;
  document.getElementById("labelFormText").value = text;
  document.getElementById("labelFormTextHex").value = text;
  updateLabelPreview();
  document.getElementById("labelFormModal").style.display = "flex";
}

function closeLabelFormModal() {
  document.getElementById("labelFormModal").style.display = "none";
}

function syncColor(colorId, hexId) {
  const hex = document.getElementById(hexId).value;
  if (/^#[0-9a-fA-F]{6}$/.test(hex)) {
    document.getElementById(colorId).value = hex;
  }
}

document.addEventListener("DOMContentLoaded", function() {
  ["labelFormBg","labelFormText"].forEach(id => {
    const el = document.getElementById(id);
    if (el) el.addEventListener("input", function() {
      document.getElementById(id+"Hex").value = this.value;
      updateLabelPreview();
    });
  });
});

function updateLabelPreview() {
  const name = document.getElementById("labelFormName").value || "Nama Label";
  const bg = document.getElementById("labelFormBg").value || "#eaf97b";
  const text = document.getElementById("labelFormText").value || "#000000";
  const prev = document.getElementById("labelPreview");
  if (prev) { prev.style.background = bg; prev.style.color = text; prev.textContent = name; }
}

async function saveLabelForm() {
  const id = document.getElementById("labelFormId").value;
  const name = document.getElementById("labelFormName").value.trim();
  const category = document.getElementById("labelFormCategory").value.trim();
  const bg_color = document.getElementById("labelFormBg").value;
  const text_color = document.getElementById("labelFormText").value;
  if (!name) { alert("Nama label wajib diisi"); return; }
  const payload = {name, category, bg_color, text_color};
  try {
    const url = id ? `/api/labels/${id}` : "/api/labels";
    const method = id ? "PUT" : "POST";
    const res = await fetch(url, {method, headers:{"Content-Type":"application/json"}, body:JSON.stringify(payload)});
    if (!res.ok) throw new Error(await res.text());
    closeLabelFormModal();
    loadLabelsTable();
  } catch(e) { alert("Error: " + e.message); }
}

async function deleteLabelRow(id) {
  if (!confirm("Hapus label ini?")) return;
  try {
    await fetch(`/api/labels/${id}`, {method:"DELETE"});
    loadLabelsTable();
  } catch(e) { alert("Error: " + e.message); }
}

// ---- Menu Utama ----
const ACTION_OPTIONS = [
  {value: "donasi", label: "Mulai Donasi"},
  {value: "admin", label: "Hubungi Admin"},
  {value: "kampanye", label: "Pilih Kampanye"},
];

function renderButtonRow(btn, idx) {
  const optsHtml = ACTION_OPTIONS.map(o => `<option value="${o.value}" ${o.value === btn.action ? "selected" : ""}>${o.label}</option>`).join("");
  return `
    <div class="button-row">
      <input type="checkbox" ${btn.enabled ? "checked" : ""} onchange="updateButton(${idx}, 'enabled', this.checked)">
      <input type="text" value="${(btn.label||"").replace(/"/g,'&quot;')}" placeholder="Label tombol" maxlength="20" oninput="updateButton(${idx}, 'label', this.value)">
      <select onchange="updateButton(${idx}, 'action', this.value)">${optsHtml}</select>
    </div>
  `;
}

let menuButtons = [];
function renderMenuButtons() {
  document.getElementById("menuButtons").innerHTML = menuButtons.map((b, i) => renderButtonRow(b, i)).join("");
}
function updateButton(idx, key, value) {
  menuButtons[idx][key] = value;
}

async function loadMenuUtama() {
  const res = await fetch("/api/settings/menu-utama");
  const data = await res.json();
  document.getElementById("menuMessage").value = data.message || "";
  document.getElementById("menuMessageCount").textContent = (data.message || "").length;
  menuButtons = data.buttons || [];
  renderMenuButtons();
}
document.getElementById("menuMessage").addEventListener("input", function() {
  document.getElementById("menuMessageCount").textContent = this.value.length;
});

async function saveMenuUtama() {
  const message = document.getElementById("menuMessage").value;
  await fetch("/api/settings/menu-utama", {
    method: "POST",
    headers: {"Content-Type": "application/json"},
    body: JSON.stringify({message, buttons: menuButtons})
  });
  const msg = document.getElementById("menuSaveMsg");
  msg.textContent = "Tersimpan!";
  setTimeout(() => msg.textContent = "", 2000);
}

// ---- WA Blast ----
let blastContactsData = [];
async function loadBlastContacts() {
  const res = await fetch("/api/inbox");
  const json = await res.json();
  blastContactsData = json.contacts || [];
  const el = document.getElementById("blastContacts");
  if (!blastContactsData.length) {
    el.innerHTML = "<div style='color:#9ca3af;padding:8px;'>Belum ada kontak</div>";
    return;
  }
  el.innerHTML = blastContactsData.map(c => `
    <label class="blast-contact-item">
      <input type="checkbox" value="${c.phone}" onchange="updateBlastCount()">
      <span>${c.name || c.phone} <span style="color:#9ca3af">(${c.phone})</span></span>
    </label>
  `).join("");
  updateBlastCount();
}
function updateBlastCount() {
  const checked = document.querySelectorAll("#blastContacts input:checked").length;
  document.getElementById("blastCount").textContent = checked + " kontak dipilih";
}
function toggleAllBlast() {
  const boxes = document.querySelectorAll("#blastContacts input");
  const allChecked = Array.from(boxes).every(b => b.checked);
  boxes.forEach(b => b.checked = !allChecked);
  updateBlastCount();
}
async function sendBlast() {
  const phones = Array.from(document.querySelectorAll("#blastContacts input:checked")).map(b => b.value);
  const message = document.getElementById("blastMessage").value.trim();
  const msgEl = document.getElementById("blastSendMsg");
  if (!phones.length) { msgEl.textContent = "Pilih minimal 1 kontak"; msgEl.style.color = "#dc2626"; return; }
  if (!message) { msgEl.textContent = "Isi pesan dulu"; msgEl.style.color = "#dc2626"; return; }
  msgEl.textContent = "Mengirim...";
  msgEl.style.color = "#6b7280";
  const res = await fetch("/api/blast", {
    method: "POST",
    headers: {"Content-Type": "application/json"},
    body: JSON.stringify({phones, message})
  });
  const json = await res.json();
  msgEl.style.color = "#16a34a";
  msgEl.textContent = `Terkirim ke ${json.sent}/${json.total} kontak`;
}

loadMenuUtama();
</script>
"""
    return Response(render_page("whatsapp", "WhatsApp", "Menu Utama dan WhatsApp Blast - sistem Raihmimpi", body), mimetype="text/html")


@app.route("/api/kampanye-list", methods=["GET"])
def api_kampanye_list():
    """Return semua kampanye dari API Raihmimpi + flag aktif."""
    campaigns = get_campaigns(full=True)
    settings = load_settings()
    aktif_ids = [str(x) for x in settings.get("kampanye_aktif", [])]
    result = []
    for c in campaigns:
        cid = str(c.get("ID_CAMPAIGN", ""))
        result.append({
            "id": cid,
            "name": c.get("CAMPAIGN_NAME", ""),
            "image": c.get("IMG_MOBILE") or c.get("IMG_CAMPAIGNER") or c.get("IMG_BIG", ""),
            "total_donasi": c.get("TOTAL_DONASI", 0),
            "target": c.get("TARGET_DONASI_UANG", 0),
            "aktif": (cid in aktif_ids) if aktif_ids else False,
        })
    return jsonify({"campaigns": result, "filter_active": bool(aktif_ids)})

@app.route("/api/settings/kampanye-aktif", methods=["GET", "POST"])
def api_settings_kampanye_aktif():
    settings = load_settings()
    if request.method == "GET":
        return jsonify({"kampanye_aktif": settings.get("kampanye_aktif", [])})
    body = request.get_json(silent=True) or {}
    ids = body.get("kampanye_aktif", [])
    if not isinstance(ids, list):
        return jsonify({"error": "kampanye_aktif harus list"}), 400
    settings["kampanye_aktif"] = [str(x) for x in ids]
    save_settings(settings)
    logger.info(f"Kampanye aktif disimpan: {len(ids)} kampanye")
    return jsonify({"status": "ok", "kampanye_aktif": settings["kampanye_aktif"]})

@app.route("/kampanye", methods=["GET"])
@login_required
def kampanye_page():
    """Halaman kelola kampanye yang ditampilkan di Flow donasi."""
    body = """
  <div class="panel">
    <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:16px;flex-wrap:wrap;gap:12px;">
      <div>
        <h3 style="margin:0 0 4px;">Pilih Kampanye Aktif</h3>
        <p style="color:#6b7280;margin:0;font-size:14px;">Kampanye yang dicentang akan muncul di Flow donasi WhatsApp (max 5 yang ditampilkan dengan gambar).</p>
      </div>
      <div style="display:flex;gap:8px;">
        <button class="btn" onclick="clearAll()" style="background:#f3f4f6;color:#374151;">Kosongkan</button>
        <button class="btn" onclick="saveSelection()" id="btnSimpan">Simpan</button>
      </div>
    </div>
    <div id="filterStatus" style="margin-bottom:16px;padding:10px 14px;background:#fef3c7;border:1px solid #f59e0b;border-radius:8px;color:#92400e;font-size:14px;display:none;"></div>

    <h4 style="margin:20px 0 10px;color:#111827;font-size:15px;">Kampanye Aktif <span id="aktifCount" style="color:#6b7280;font-weight:normal;font-size:13px;"></span></h4>
    <div id="kampanyeAktifGrid" style="display:grid;grid-template-columns:repeat(auto-fill,minmax(280px,1fr));gap:16px;margin-bottom:24px;">
      <div style="color:#9ca3af;grid-column:1/-1;text-align:center;padding:20px;font-size:14px;">Belum ada kampanye aktif</div>
    </div>

    <h4 style="margin:24px 0 10px;color:#111827;font-size:15px;">Daftar Kampanye Lainnya <span id="totalCount" style="color:#6b7280;font-weight:normal;font-size:13px;"></span></h4>
    <input type="text" id="searchKampanye" placeholder="Cari berdasarkan ID atau nama kampanye..." onkeyup="renderList()" style="width:100%;padding:10px 14px;border:1px solid #e5e7eb;border-radius:8px;font-size:14px;margin-bottom:12px;box-sizing:border-box;">
    <div id="kampanyeList" style="border:1px solid #e5e7eb;border-radius:8px;max-height:500px;overflow-y:auto;">
      <div style="color:#9ca3af;text-align:center;padding:20px;">Memuat...</div>
    </div>
  </div>

<script>
let allCampaigns = [];

function formatRupiah(n) {
  return "Rp " + Number(n || 0).toLocaleString("id-ID");
}

async function loadKampanye() {
  try {
    const res = await fetch("/api/kampanye-list");
    const json = await res.json();
    allCampaigns = json.campaigns || [];
    const filterStatus = document.getElementById("filterStatus");
    if (json.filter_active) {
      const aktifCount = allCampaigns.filter(c => c.aktif).length;
      filterStatus.style.display = "block";
      filterStatus.textContent = `Filter aktif: ${aktifCount} kampanye dipilih ditampilkan di Flow. Kalau kosong, semua kampanye akan ditampilkan.`;
    } else {
      filterStatus.style.display = "block";
      filterStatus.textContent = "Belum ada kampanye dipilih — saat ini SEMUA kampanye ditampilkan di Flow (default).";
    }
    render();
  } catch (e) {
    document.getElementById("kampanyeAktifGrid").innerHTML = `<div style="color:#dc2626;grid-column:1/-1;">Error: ${e.message}</div>`;
  }
}

function render() {
  renderAktif();
  renderList();
}

function renderAktif() {
  const aktif = allCampaigns.filter(c => c.aktif);
  const grid = document.getElementById("kampanyeAktifGrid");
  document.getElementById("aktifCount").textContent = `(${aktif.length})`;

  if (!aktif.length) {
    grid.innerHTML = `<div style="color:#9ca3af;grid-column:1/-1;text-align:center;padding:20px;font-size:14px;">Belum ada kampanye aktif — centang dari daftar di bawah</div>`;
    return;
  }
  grid.innerHTML = aktif.map(c => {
    const pct = c.target > 0 ? Math.min(100, Math.round(c.total_donasi / c.target * 100)) : 0;
    return `
      <label style="border:2px solid #5b3df0;border-radius:10px;padding:12px;cursor:pointer;background:#fff;display:block;">
        <div style="display:flex;gap:12px;align-items:flex-start;">
          <input type="checkbox" checked onchange="toggleAktif('${c.id}', this.checked)" style="width:18px;height:18px;margin-top:2px;cursor:pointer;">
          ${c.image ? `<img src="${c.image}" style="width:60px;height:60px;object-fit:cover;border-radius:6px;flex-shrink:0;" onerror="this.style.display='none'">` : ''}
          <div style="flex:1;min-width:0;">
            <div style="font-weight:600;font-size:14px;color:#111827;margin-bottom:4px;line-height:1.3;">${(c.name || '').replace(/</g,'&lt;')}</div>
            <div style="font-size:12px;color:#6b7280;">ID: ${c.id}</div>
            <div style="font-size:12px;color:#374151;margin-top:4px;">${formatRupiah(c.total_donasi)} <span style="color:#9ca3af;">/ ${formatRupiah(c.target)} (${pct}%)</span></div>
            <div style="height:4px;background:#e5e7eb;border-radius:2px;margin-top:4px;overflow:hidden;"><div style="height:100%;background:#5b3df0;width:${pct}%;"></div></div>
          </div>
        </div>
      </label>
    `;
  }).join("");
}

function renderList() {
  const nonAktif = allCampaigns.filter(c => !c.aktif);
  const q = (document.getElementById("searchKampanye").value || "").toLowerCase().trim();
  const filtered = q ? nonAktif.filter(c =>
    (c.name || "").toLowerCase().includes(q) || (c.id || "").toLowerCase().includes(q)
  ) : nonAktif;

  document.getElementById("totalCount").textContent = `(${filtered.length}${q ? ` dari ${nonAktif.length}` : ''})`;

  const listEl = document.getElementById("kampanyeList");
  if (!filtered.length) {
    listEl.innerHTML = `<div style="color:#9ca3af;text-align:center;padding:30px;font-size:14px;">${q ? 'Tidak ada hasil untuk "' + q + '"' : 'Semua kampanye sudah aktif'}</div>`;
    return;
  }
  listEl.innerHTML = filtered.map((c, i) => `
    <label style="display:flex;align-items:center;gap:12px;padding:10px 14px;cursor:pointer;${i < filtered.length-1 ? 'border-bottom:1px solid #f3f4f6;' : ''}" onmouseover="this.style.background='#f9fafb'" onmouseout="this.style.background='#fff'">
      <input type="checkbox" onchange="toggleAktif('${c.id}', this.checked)" style="width:16px;height:16px;cursor:pointer;flex-shrink:0;">
      <span style="font-size:13px;color:#6b7280;font-family:monospace;min-width:60px;">${c.id}</span>
      <span style="font-size:14px;color:#111827;flex:1;">${(c.name || '').replace(/</g,'&lt;')}</span>
    </label>
  `).join("");
  document.getElementById("totalCount").textContent = `(${filtered.length}${q ? ` dari ${nonAktif.length}` : ''})`;
}

function toggleAktif(id, checked) {
  const c = allCampaigns.find(x => x.id === id);
  if (c) c.aktif = checked;
  render();
}

function selectAll() {
  allCampaigns.forEach(c => c.aktif = true);
  render();
}

function clearAll() {
  allCampaigns.forEach(c => c.aktif = false);
  render();
}

async function saveSelection() {
  const btn = document.getElementById("btnSimpan");
  btn.disabled = true;
  btn.textContent = "Menyimpan...";
  try {
    const ids = allCampaigns.filter(c => c.aktif).map(c => c.id);
    const res = await fetch("/api/settings/kampanye-aktif", {
      method: "POST",
      headers: {"Content-Type": "application/json"},
      body: JSON.stringify({kampanye_aktif: ids})
    });
    const json = await res.json();
    if (json.status === "ok") {
      alert(`✓ Tersimpan: ${ids.length} kampanye akan ditampilkan di Flow.`);
      loadKampanye();
    } else {
      alert("Gagal: " + (json.error || "unknown"));
    }
  } catch (e) {
    alert("Error: " + e.message);
  } finally {
    btn.disabled = false;
    btn.textContent = "Simpan";
  }
}

loadKampanye();
</script>
"""
    return Response(render_page("kampanye", "Kelola Kampanye", "Pilih kampanye yang muncul di Flow donasi", body), mimetype="text/html")


# ============================================================
# Label Kontak (Contact Labels)
# ============================================================
def load_labels():
    """Load master labels dari settings.json."""
    settings = load_settings()
    return settings.get("labels", [])

def save_labels(labels):
    """Simpan master labels ke settings.json."""
    settings = load_settings()
    settings["labels"] = labels
    save_settings(settings)

@app.route("/api/labels", methods=["GET"])
def api_labels_list():
    """List semua master label."""
    return jsonify({"labels": load_labels()})

@app.route("/api/labels", methods=["POST"])
def api_labels_create():
    """Create label baru."""
    body = request.get_json(silent=True) or {}
    name = (body.get("name") or "").strip()
    if not name:
        return jsonify({"error": "Nama label wajib diisi"}), 400
    labels = load_labels()
    # Cek duplikat nama
    if any(l.get("name", "").lower() == name.lower() for l in labels):
        return jsonify({"error": f"Label dengan nama '{name}' sudah ada"}), 400
    new_label = {
        "id": str(uuid.uuid4()),
        "name": name,
        "bg_color": body.get("bg_color", "#5b3df0"),
        "text_color": body.get("text_color", "#ffffff"),
        "category": body.get("category", ""),
    }
    labels.append(new_label)
    save_labels(labels)
    return jsonify(new_label)

@app.route("/api/labels/<label_id>", methods=["PUT"])
def api_labels_update(label_id):
    """Update label by id."""
    body = request.get_json(silent=True) or {}
    labels = load_labels()
    for label in labels:
        if label.get("id") == label_id:
            if "name" in body:
                name = (body["name"] or "").strip()
                if not name:
                    return jsonify({"error": "Nama tidak boleh kosong"}), 400
                # Cek duplikat (kecuali diri sendiri)
                if any(l.get("name", "").lower() == name.lower() and l.get("id") != label_id for l in labels):
                    return jsonify({"error": f"Label dengan nama '{name}' sudah ada"}), 400
                label["name"] = name
            if "bg_color" in body:
                label["bg_color"] = body["bg_color"]
            if "text_color" in body:
                label["text_color"] = body["text_color"]
            if "category" in body:
                label["category"] = body["category"]
            save_labels(labels)
            return jsonify(label)
    return jsonify({"error": "Label tidak ditemukan"}), 404

@app.route("/api/labels/<label_id>", methods=["DELETE"])
def api_labels_delete(label_id):
    """Hapus label + bersihkan dari semua kontak."""
    labels = load_labels()
    new_labels = [l for l in labels if l.get("id") != label_id]
    if len(new_labels) == len(labels):
        return jsonify({"error": "Label tidak ditemukan"}), 404
    save_labels(new_labels)
    # Bersihkan label dari semua contact
    inbox = load_inbox()
    cleaned = 0
    for phone, contact in inbox.get("contacts", {}).items():
        if label_id in contact.get("labels", []):
            contact["labels"] = [lid for lid in contact["labels"] if lid != label_id]
            cleaned += 1
    if cleaned > 0:
        save_inbox(inbox)
    return jsonify({"status": "ok", "cleaned_contacts": cleaned})

# ============================================================
# SHORTCUTS API
# ============================================================

@app.route("/api/shortcuts", methods=["GET"])
def api_shortcuts_list():
    """List semua shortcuts."""
    shortcuts = load_shortcuts()
    return jsonify({"shortcuts": shortcuts})

@app.route("/api/shortcuts", methods=["POST"])
def api_shortcuts_create():
    """Buat shortcut baru (text) atau dengan file attachment."""
    try:
        shortcut_key = request.form.get("shortcut", "").strip()
        title = request.form.get("title", "").strip()
        content_type = request.form.get("content_type", "TEXT").upper()
        isi = request.form.get("isi", "").strip()

        if not shortcut_key or not title:
            return jsonify({"error": "shortcut dan judul wajib diisi"}), 400
        if not shortcut_key.startswith("/"):
            shortcut_key = "/" + shortcut_key

        shortcuts = load_shortcuts()
        # Cek duplikat
        if any(s["shortcut"] == shortcut_key for s in shortcuts):
            return jsonify({"error": f"Shortcut '{shortcut_key}' sudah ada"}), 400

        new_id = f"sc_{int(datetime.now().timestamp()*1000)}"
        sc = {
            "id": new_id,
            "shortcut": shortcut_key,
            "title": title,
            "content_type": content_type,
            "isi": isi,
            "file_path": None,
            "file_name": None,
            "media_id": None,
            "created_at": datetime.now().isoformat(),
        }

        # Handle file upload
        file = request.files.get("file")
        if file and content_type in ("IMAGE", "DOCUMENT", "VIDEO"):
            os.makedirs(SHORTCUTS_MEDIA_DIR, exist_ok=True)
            filename = f"{new_id}_{file.filename}"
            filepath = os.path.join(SHORTCUTS_MEDIA_DIR, filename)
            file.save(filepath)
            sc["file_path"] = filepath
            sc["file_name"] = file.filename

            # Upload ke Meta untuk dapat media_id (re-usable)
            try:
                mime = file.content_type or "application/octet-stream"
                with open(filepath, "rb") as f_read:
                    upload_resp = requests.post(
                        f"https://graph.facebook.com/v22.0/{WA_PHONE_NUMBER_ID}/media",
                        headers={"Authorization": f"Bearer {WA_TOKEN}"},
                        files={"file": (file.filename, f_read, mime)},
                        data={"messaging_product": "whatsapp", "type": mime},
                        timeout=30)
                if upload_resp.ok:
                    sc["media_id"] = upload_resp.json().get("id")
                    logger.info(f"Shortcut media uploaded: {sc['media_id']}")
            except Exception as ue:
                logger.error(f"Upload shortcut media gagal: {ue}")

        shortcuts.append(sc)
        save_shortcuts(shortcuts)
        return jsonify({"status": "created", "shortcut": sc})
    except Exception as e:
        logger.error(f"api_shortcuts_create error: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500

@app.route("/api/shortcuts/<sc_id>", methods=["PUT"])
def api_shortcuts_update(sc_id):
    """Update shortcut (text fields saja, file diupdate via DELETE + POST baru)."""
    try:
        body = request.get_json(silent=True) or {}
        shortcuts = load_shortcuts()
        for sc in shortcuts:
            if sc["id"] == sc_id:
                sc["title"] = body.get("title", sc["title"])
                sc["isi"] = body.get("isi", sc["isi"])
                sc["shortcut"] = body.get("shortcut", sc["shortcut"])
                save_shortcuts(shortcuts)
                return jsonify({"status": "updated", "shortcut": sc})
        return jsonify({"error": "Shortcut tidak ditemukan"}), 404
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/shortcuts/<sc_id>", methods=["DELETE"])
def api_shortcuts_delete(sc_id):
    """Hapus shortcut + file attachment-nya."""
    try:
        shortcuts = load_shortcuts()
        sc = next((s for s in shortcuts if s["id"] == sc_id), None)
        if not sc:
            return jsonify({"error": "Shortcut tidak ditemukan"}), 404
        # Hapus file lokal
        if sc.get("file_path") and os.path.exists(sc["file_path"]):
            os.remove(sc["file_path"])
        shortcuts = [s for s in shortcuts if s["id"] != sc_id]
        save_shortcuts(shortcuts)
        return jsonify({"status": "deleted"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/shortcuts/<sc_id>/send/<phone>", methods=["POST"])
def api_shortcuts_send(sc_id, phone):
    """Kirim shortcut ke kontak (text atau media)."""
    try:
        shortcuts = load_shortcuts()
        sc = next((s for s in shortcuts if s["id"] == sc_id), None)
        if not sc:
            return jsonify({"error": "Shortcut tidak ditemukan"}), 404

        content_type = sc.get("content_type", "TEXT")

        if content_type == "TEXT":
            resp = send_wa_message(phone, sc["isi"])
            record_outgoing_message(phone, sc["isi"], msg_type="text")
            return jsonify({"status": "sent", "wa_status": resp.status_code})

        # Media: pakai media_id yang sudah diupload, atau re-upload
        media_id = sc.get("media_id")
        if not media_id and sc.get("file_path") and os.path.exists(sc["file_path"]):
            # Re-upload
            mime = "application/octet-stream"
            with open(sc["file_path"], "rb") as f_read:
                upload_resp = requests.post(
                    f"https://graph.facebook.com/v22.0/{WA_PHONE_NUMBER_ID}/media",
                    headers={"Authorization": f"Bearer {WA_TOKEN}"},
                    files={"file": (sc["file_name"], f_read, mime)},
                    data={"messaging_product": "whatsapp", "type": mime},
                    timeout=30)
            if upload_resp.ok:
                media_id = upload_resp.json().get("id")
                # Update cached media_id
                sc["media_id"] = media_id
                save_shortcuts(shortcuts)

        if not media_id:
            return jsonify({"error": "File tidak ditemukan atau upload gagal"}), 500

        type_map = {"IMAGE": "image", "DOCUMENT": "document", "VIDEO": "video"}
        wa_type = type_map.get(content_type, "document")
        if wa_type == "document":
            media_payload = {"id": media_id, "filename": sc.get("file_name", "file")}
        else:
            media_payload = {"id": media_id}

        send_url = f"https://graph.facebook.com/v22.0/{WA_PHONE_NUMBER_ID}/messages"
        resp = requests.post(send_url,
            headers={"Authorization": f"Bearer {WA_TOKEN}", "Content-Type": "application/json"},
            json={"messaging_product": "whatsapp", "to": phone, "recipient_type": "individual",
                  "type": wa_type, wa_type: media_payload},
            timeout=15)
        label = f"{'🖼' if wa_type=='image' else '🎥' if wa_type=='video' else '📄'} {sc['title']}"
        record_outgoing_message(phone, label, msg_type=wa_type)
        return jsonify({"status": "sent", "wa_status": resp.status_code})
    except Exception as e:
        logger.error(f"api_shortcuts_send error: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500

# ============================================================
# TEMPLATE WA BLAST API
# ============================================================

@app.route("/api/wa-templates", methods=["GET"])
def api_wa_templates_list():
    """Fetch list template dari Meta API."""
    try:
        resp = requests.get(
            f"https://graph.facebook.com/v22.0/{WABA_ID}/message_templates",
            params={
                "limit": 50,
                "fields": "name,status,category,language,components,rejected_reason"
            },
            headers={"Authorization": f"Bearer {WA_ACCESS_TOKEN}"},
            timeout=15
        )
        data = resp.json()
        return jsonify({"status": resp.status_code, "templates": data.get("data", []), "raw": data})
    except Exception as e:
        logger.error(f"api_wa_templates_list error: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500

@app.route("/api/wa-templates", methods=["POST"])
def api_wa_templates_create():
    """Buat template baru dan submit ke Meta untuk approval."""
    try:
        body = request.get_json(silent=True) or {}
        name = body.get("name", "").strip().lower().replace(" ", "_")
        category = body.get("category", "MARKETING").upper()
        language = body.get("language", "id")
        body_text = body.get("body_text", "").strip()
        header_type = body.get("header_type", "")
        header_text = body.get("header_text", "").strip()
        footer_text = body.get("footer_text", "").strip()
        buttons = body.get("buttons", [])

        if not name or not body_text:
            return jsonify({"error": "name dan body_text wajib diisi"}), 400

        components = []

        # Header (optional)
        if header_type == "TEXT" and header_text:
            components.append({"type": "HEADER", "format": "TEXT", "text": header_text})

        # Body (wajib)
        components.append({"type": "BODY", "text": body_text})

        # Footer (optional)
        if footer_text:
            components.append({"type": "FOOTER", "text": footer_text})

        # Buttons (optional)
        if buttons:
            btn_components = []
            for btn in buttons:
                btn_type = btn.get("type", "QUICK_REPLY")
                if btn_type == "QUICK_REPLY":
                    btn_components.append({"type": "QUICK_REPLY", "text": btn.get("text", "")})
                elif btn_type == "PHONE_NUMBER":
                    btn_components.append({"type": "PHONE_NUMBER", "text": btn.get("text", ""), "phone_number": btn.get("value", "")})
                elif btn_type == "URL":
                    btn_components.append({"type": "URL", "text": btn.get("text", ""), "url": btn.get("value", "")})
            if btn_components:
                components.append({"type": "BUTTONS", "buttons": btn_components})

        payload = {
            "name": name,
            "category": category,
            "language": language,
            "components": components
        }

        resp = requests.post(
            f"https://graph.facebook.com/v22.0/{WABA_ID}/message_templates",
            headers={"Authorization": f"Bearer {WA_ACCESS_TOKEN}", "Content-Type": "application/json"},
            json=payload,
            timeout=15
        )
        logger.info(f"api_wa_templates_create: {resp.status_code} {resp.text[:300]}")
        result = resp.json()
        if resp.ok:
            return jsonify({"status": "submitted", "id": result.get("id"), "name": name})
        else:
            return jsonify({"error": result.get("error", {}).get("message", "Unknown error"), "detail": result}), 400
    except Exception as e:
        logger.error(f"api_wa_templates_create error: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500

@app.route("/api/wa-templates/<template_id>", methods=["DELETE"])
def api_wa_templates_delete(template_id):
    """Hapus template dari Meta."""
    try:
        name = request.args.get("name", "")
        resp = requests.delete(
            f"https://graph.facebook.com/v22.0/{WABA_ID}/message_templates",
            params={"name": name},
            headers={"Authorization": f"Bearer {WA_ACCESS_TOKEN}"},
            timeout=15
        )
        return jsonify({"status": resp.status_code, "result": resp.json()})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/wa-templates/chat-setting", methods=["POST"])
def api_wa_templates_chat_setting():
    """Set/unset template sebagai Template Chat (simpan ke settings)."""
    try:
        body = request.get_json(silent=True) or {}
        name = body.get("name", "")
        enabled = body.get("enabled", False)
        settings = load_settings()
        chat_templates = settings.get("wa_chat_templates", [])
        if enabled and name not in chat_templates:
            chat_templates.append(name)
        elif not enabled and name in chat_templates:
            chat_templates.remove(name)
        settings["wa_chat_templates"] = chat_templates
        save_settings(settings)
        return jsonify({"status": "ok", "wa_chat_templates": chat_templates})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/wa-templates/send/<phone>", methods=["POST"])
def api_wa_templates_send(phone):
    """Kirim template message ke kontak."""
    try:
        body = request.get_json(silent=True) or {}
        template_name = body.get("name", "")
        language = body.get("language", "id")
        components = body.get("components", [])

        payload = {
            "messaging_product": "whatsapp",
            "to": phone,
            "type": "template",
            "template": {
                "name": template_name,
                "language": {"code": language},
            }
        }
        if components:
            payload["template"]["components"] = components

        resp = requests.post(
            f"https://graph.facebook.com/v22.0/{WA_PHONE_NUMBER_ID}/messages",
            headers={"Authorization": f"Bearer {WA_ACCESS_TOKEN}", "Content-Type": "application/json"},
            json=payload, timeout=15
        )
        logger.info(f"api_wa_templates_send to={phone} template={template_name} status={resp.status_code}")
        record_outgoing_message(phone, f"📢 Template: {template_name}", msg_type="text")
        return jsonify({"status": "sent", "wa_status": resp.status_code, "wa_body": resp.json()})
    except Exception as e:
        logger.error(f"api_wa_templates_send error: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500

@app.route("/api/settings", methods=["GET"])
def api_settings_get():
    """Get settings (untuk frontend baca wa_chat_templates dll)."""
    settings = load_settings()
    return jsonify(settings)

@app.route("/api/test-templates", methods=["GET"])
def api_test_templates():
    """Debug: test fetch template dari Meta API."""
    try:
        resp = requests.get(
            f"https://graph.facebook.com/v22.0/{WABA_ID}/message_templates",
            params={"limit": 5, "fields": "name,status,category,components"},
            headers={"Authorization": f"Bearer {WA_TOKEN}"},
            timeout=15
        )
        return jsonify({"status": resp.status_code, "data": resp.json()})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/labels", methods=["GET"])
def labels_page():
    """Halaman manajemen master label."""
    body = """
<style>
  .labels-wrap { padding:20px; }
  .labels-header { display:flex; justify-content:space-between; align-items:center; margin-bottom:20px; }
  .labels-header h2 { margin:0; font-size:20px; font-weight:700; }
  .btn-add { background:#5b3df0; color:#fff; border:none; padding:10px 20px; border-radius:8px; font-weight:600; cursor:pointer; font-size:14px; }
  .btn-add:hover { background:#4c30d9; }
  .labels-table { width:100%; background:#fff; border-radius:8px; overflow:hidden; box-shadow:0 1px 3px rgba(0,0,0,.06); }
  .labels-table th { background:#f9fafb; padding:12px 16px; text-align:left; font-size:13px; color:#6b7280; font-weight:600; border-bottom:1px solid #e5e7eb; }
  .labels-table td { padding:14px 16px; border-bottom:1px solid #f3f4f6; font-size:14px; }
  .labels-table tr:hover { background:#fafafa; }
  .label-preview { display:inline-block; padding:4px 12px; border-radius:12px; font-size:12px; font-weight:600; }
  .color-swatch { display:inline-flex; align-items:center; gap:6px; font-family:monospace; font-size:12px; }
  .color-box { width:20px; height:20px; border-radius:4px; border:1px solid #e5e7eb; }
  .btn-icon { background:none; border:none; cursor:pointer; padding:6px 8px; border-radius:6px; font-size:14px; }
  .btn-icon:hover { background:#f3f4f6; }
  .btn-edit { color:#5b3df0; }
  .btn-delete { color:#dc2626; }
  .modal-overlay { display:none; position:fixed; inset:0; background:rgba(0,0,0,.5); align-items:center; justify-content:center; z-index:1000; }
  .modal-overlay.show { display:flex; }
  .modal-card { background:#fff; border-radius:12px; padding:24px; width:480px; max-width:90vw; }
  .modal-card h3 { margin:0 0 20px 0; font-size:18px; font-weight:700; }
  .form-group { margin-bottom:16px; }
  .form-group label { display:block; font-size:13px; font-weight:600; margin-bottom:6px; color:#374151; }
  .form-group input[type=text] { width:100%; padding:10px 12px; border:1px solid #d1d5db; border-radius:8px; font-size:14px; box-sizing:border-box; }
  .form-group input[type=color] { width:60px; height:40px; border:1px solid #d1d5db; border-radius:8px; cursor:pointer; padding:2px; }
  .color-row { display:flex; gap:16px; align-items:flex-end; }
  .color-input-wrap { display:flex; gap:8px; align-items:center; }
  .color-input-wrap input[type=text] { width:100px; font-family:monospace; }
  .preview-section { background:#f9fafb; padding:14px; border-radius:8px; margin:16px 0; text-align:center; }
  .preview-section .pv-label { font-size:12px; color:#6b7280; margin-bottom:8px; }
  .modal-actions { display:flex; gap:10px; justify-content:flex-end; margin-top:20px; }
  .btn-cancel { background:#f3f4f6; color:#374151; border:none; padding:10px 20px; border-radius:8px; font-weight:600; cursor:pointer; }
  .btn-save { background:#5b3df0; color:#fff; border:none; padding:10px 20px; border-radius:8px; font-weight:600; cursor:pointer; }
  .empty-state { padding:60px 20px; text-align:center; color:#9ca3af; background:#fff; border-radius:8px; }
  .empty-state .icon { font-size:48px; margin-bottom:12px; }
</style>
<div class="labels-wrap">
  <div class="labels-header">
    <h2>Daftar Label Kontak</h2>
    <button class="btn-add" onclick="openLabelModal()">+ Tambah Label</button>
  </div>
  <div id="labelsContainer"></div>
</div>

<div class="modal-overlay" id="labelModal">
  <div class="modal-card">
    <h3 id="modalTitle">Tambah Label</h3>
    <input type="hidden" id="labelId">
    <div class="form-group">
      <label>Nama Label *</label>
      <input type="text" id="labelName" placeholder="Contoh: HOT, Donatur Rutin" maxlength="50">
    </div>
    <div class="form-group">
      <label>Kategori (opsional)</label>
      <input type="text" id="labelCategory" placeholder="Contoh: Donatur, Sumber, Status" maxlength="30">
    </div>
    <div class="color-row">
      <div class="form-group" style="margin:0;">
        <label>Warna Latar</label>
        <div class="color-input-wrap">
          <input type="color" id="labelBgColor" value="#5b3df0" onchange="syncColor('bg')">
          <input type="text" id="labelBgHex" value="#5b3df0" onchange="syncHex('bg')">
        </div>
      </div>
      <div class="form-group" style="margin:0;">
        <label>Warna Teks</label>
        <div class="color-input-wrap">
          <input type="color" id="labelTextColor" value="#ffffff" onchange="syncColor('text')">
          <input type="text" id="labelTextHex" value="#ffffff" onchange="syncHex('text')">
        </div>
      </div>
    </div>
    <div class="preview-section">
      <div class="pv-label">Preview:</div>
      <span class="label-preview" id="labelPreview" style="background:#5b3df0;color:#fff;">Label</span>
    </div>
    <div class="modal-actions">
      <button class="btn-cancel" onclick="closeLabelModal()">Batal</button>
      <button class="btn-save" id="btnSaveLabel" onclick="saveLabel()">Simpan</button>
    </div>
  </div>
</div>

<script>
let allLabels = [];

async function loadLabels() {
  try {
    const res = await fetch("/api/labels");
    const data = await res.json();
    allLabels = data.labels || [];
    renderLabels();
  } catch (e) {
    document.getElementById("labelsContainer").innerHTML = `<div class="empty-state">Error memuat label: ${e.message}</div>`;
  }
}

function renderLabels() {
  const container = document.getElementById("labelsContainer");
  if (allLabels.length === 0) {
    container.innerHTML = `<div class="empty-state"><div class="icon">🏷️</div><div>Belum ada label.</div><div style="margin-top:8px;font-size:13px;">Klik <b>+ Tambah Label</b> untuk membuat label pertama.</div></div>`;
    return;
  }
  const rows = allLabels.map(l => `
    <tr>
      <td>
        <span class="label-preview" style="background:${l.bg_color};color:${l.text_color};">${escapeHtml(l.name)}</span>
      </td>
      <td>${l.category ? escapeHtml(l.category) : '<span style="color:#9ca3af;">-</span>'}</td>
      <td><span class="color-swatch"><span class="color-box" style="background:${l.bg_color};"></span>${l.bg_color}</span></td>
      <td><span class="color-swatch"><span class="color-box" style="background:${l.text_color};"></span>${l.text_color}</span></td>
      <td style="text-align:right;">
        <button class="btn-icon btn-edit" onclick="editLabel('${l.id}')" title="Edit">✏️</button>
        <button class="btn-icon btn-delete" onclick="deleteLabel('${l.id}', '${escapeHtml(l.name).replace(/'/g,"&#39;")}')" title="Hapus">🗑️</button>
      </td>
    </tr>
  `).join("");
  container.innerHTML = `
    <table class="labels-table">
      <thead><tr><th>Nama Label</th><th>Kategori</th><th>Warna Latar</th><th>Warna Teks</th><th style="text-align:right;">Aksi</th></tr></thead>
      <tbody>${rows}</tbody>
    </table>
  `;
}

function escapeHtml(s) {
  return String(s || "").replace(/</g, "&lt;").replace(/>/g, "&gt;").replace(/"/g, "&quot;");
}

function openLabelModal(label) {
  document.getElementById("modalTitle").textContent = label ? "Edit Label" : "Tambah Label";
  document.getElementById("labelId").value = label ? label.id : "";
  document.getElementById("labelName").value = label ? label.name : "";
  document.getElementById("labelCategory").value = label && label.category ? label.category : "";
  document.getElementById("labelBgColor").value = label ? label.bg_color : "#5b3df0";
  document.getElementById("labelBgHex").value = label ? label.bg_color : "#5b3df0";
  document.getElementById("labelTextColor").value = label ? label.text_color : "#ffffff";
  document.getElementById("labelTextHex").value = label ? label.text_color : "#ffffff";
  updatePreview();
  document.getElementById("labelModal").classList.add("show");
  setTimeout(() => document.getElementById("labelName").focus(), 100);
}

function closeLabelModal() {
  document.getElementById("labelModal").classList.remove("show");
}

function syncColor(field) {
  if (field === "bg") {
    document.getElementById("labelBgHex").value = document.getElementById("labelBgColor").value;
  } else {
    document.getElementById("labelTextHex").value = document.getElementById("labelTextColor").value;
  }
  updatePreview();
}

function syncHex(field) {
  if (field === "bg") {
    const hex = document.getElementById("labelBgHex").value;
    if (/^#[0-9a-fA-F]{6}$/.test(hex)) document.getElementById("labelBgColor").value = hex;
  } else {
    const hex = document.getElementById("labelTextHex").value;
    if (/^#[0-9a-fA-F]{6}$/.test(hex)) document.getElementById("labelTextColor").value = hex;
  }
  updatePreview();
}

function updatePreview() {
  const name = document.getElementById("labelName").value || "Label";
  const bg = document.getElementById("labelBgHex").value;
  const text = document.getElementById("labelTextHex").value;
  const pv = document.getElementById("labelPreview");
  pv.style.background = bg;
  pv.style.color = text;
  pv.textContent = name;
}

document.addEventListener("DOMContentLoaded", () => {
  const nameInput = document.getElementById("labelName");
  if (nameInput) nameInput.addEventListener("input", updatePreview);
});

function editLabel(id) {
  const label = allLabels.find(l => l.id === id);
  if (label) openLabelModal(label);
}

async function deleteLabel(id, name) {
  if (!confirm(`Hapus label "${name}"?\n\nLabel ini juga akan dihapus dari semua kontak yang memilikinya.`)) return;
  try {
    const res = await fetch(`/api/labels/${id}`, {method: "DELETE"});
    const data = await res.json();
    if (data.status === "ok") {
      alert(`✓ Label dihapus.${data.cleaned_contacts > 0 ? " Dibersihkan dari " + data.cleaned_contacts + " kontak." : ""}`);
      loadLabels();
    } else {
      alert("Gagal: " + (data.error || "unknown"));
    }
  } catch (e) {
    alert("Error: " + e.message);
  }
}

async function saveLabel() {
  const id = document.getElementById("labelId").value;
  const name = document.getElementById("labelName").value.trim();
  if (!name) {
    alert("Nama label wajib diisi.");
    return;
  }
  const bg_color = document.getElementById("labelBgHex").value;
  const text_color = document.getElementById("labelTextHex").value;
  const category = document.getElementById("labelCategory").value.trim();
  const payload = {name, bg_color, text_color, category};
  const btn = document.getElementById("btnSaveLabel");
  btn.disabled = true;
  btn.textContent = "Menyimpan...";
  try {
    let res;
    if (id) {
      res = await fetch(`/api/labels/${id}`, {method: "PUT", headers: {"Content-Type": "application/json"}, body: JSON.stringify(payload)});
    } else {
      res = await fetch(`/api/labels`, {method: "POST", headers: {"Content-Type": "application/json"}, body: JSON.stringify(payload)});
    }
    const data = await res.json();
    if (res.ok) {
      closeLabelModal();
      loadLabels();
    } else {
      alert("Gagal: " + (data.error || "unknown"));
    }
  } catch (e) {
    alert("Error: " + e.message);
  } finally {
    btn.disabled = false;
    btn.textContent = "Simpan";
  }
}

loadLabels();
</script>
"""
    return Response(render_page("labels", "Label Kontak", "Atur label kontak: nama, warna, dan kategori", body), mimetype="text/html")


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
